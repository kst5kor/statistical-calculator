#!/usr/bin/env python3
"""
Generate an Excel-based Statistical Process Capability Tool — Deep Edition.
Mirrors the original Streamlit application with:
- Auto-linked Data Worksheet (x̄ and σ calculated automatically)
- Manual/Worksheet mode toggle via dropdown
- Full summary panel with centering, capability, robustness, tolerance, hypothesis
- Conditional formatting throughout
- Process distribution bell curve + capability bar chart
- History log with pre-linked formulas
- Reference guide
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import math

OUTPUT_FILE = "SPC_Statistical_Calculator.xlsx"
MAX_DATA_ROWS = 1000

# ===========================================================================
# Color palette
# ===========================================================================
BLUE       = "3B82F6"
DARK_BLUE  = "1E3A8A"
GREEN      = "10B981"
DARK_GREEN = "047857"
RED        = "EF4444"
DARK_RED   = "991B1B"
ORANGE     = "F97316"
AMBER      = "92400E"
GRAY       = "6B7280"
LIGHT_GRAY = "F1F5F9"
WHITE      = "FFFFFF"
DARK_BG    = "1F2937"

# ===========================================================================
# Reusable styles
# ===========================================================================
HEADER_FONT    = Font(name="Calibri", bold=True, size=11, color=WHITE)
HEADER_FILL    = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
SECTION_FONT   = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
SECTION_FILL   = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
INPUT_FILL     = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
RESULT_FILL    = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
CALC_FILL      = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
WARN_FILL      = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
BAD_FILL       = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
GOOD_FILL      = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
VERDICT_GREEN  = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
VERDICT_ORANGE = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
VERDICT_RED    = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
LABEL_FONT     = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
VALUE_FONT     = Font(name="Calibri", size=11)
TITLE_FONT     = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
SUBTITLE_FONT  = Font(name="Calibri", size=10, italic=True, color=GRAY)
THIN_BORDER    = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ===========================================================================
# Helper functions
# ===========================================================================
def header_row(ws, row, start_col, texts, merge_end_col=None):
    for i, text in enumerate(texts):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=merge_end_col)


def section_row(ws, row, start_col, end_col, text):
    ws.merge_cells(start_row=row, start_column=start_col,
                   end_row=row, end_column=end_col)
    c = ws.cell(row=row, column=start_col, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = LEFT
    c.border = THIN_BORDER
    for col in range(start_col, end_col + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=col).fill = SECTION_FILL


def lv(ws, row, lc, label, vc, value=None, formula=None, is_input=False,
       fmt=None, fill=None, font=None, merge_label_to=None):
    """Write label-value pair."""
    cell_l = ws.cell(row=row, column=lc, value=label)
    cell_l.font = LABEL_FONT
    cell_l.alignment = LEFT
    cell_l.border = THIN_BORDER
    if merge_label_to:
        ws.merge_cells(start_row=row, start_column=lc,
                       end_row=row, end_column=merge_label_to)

    cell_v = ws.cell(row=row, column=vc)
    if formula:
        cell_v.value = formula
    elif value is not None:
        cell_v.value = value
    cell_v.font = font or VALUE_FONT
    cell_v.alignment = CENTER
    cell_v.border = THIN_BORDER
    if fill:
        cell_v.fill = fill
    elif is_input:
        cell_v.fill = INPUT_FILL
    else:
        cell_v.fill = RESULT_FILL
    if fmt:
        cell_v.number_format = fmt
    return cell_v


def add_cpk_cond_fmt(ws, cell_range):
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["1.67"],
                   fill=GOOD_FILL, font=Font(color=DARK_GREEN, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="between", formula=["1", "1.669"],
                   fill=WARN_FILL, font=Font(color=AMBER, bold=True)))
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="lessThan", formula=["1"],
                   fill=BAD_FILL, font=Font(color=DARK_RED, bold=True)))


# ===========================================================================
# Cell references (Analysis sheet)
# ===========================================================================
# These are fixed cell addresses used in formulas throughout the workbook.
# Section 1 — Specifications
C_NAME    = "C5"   # Measurement name
C_TM      = "C6"   # Target Mean
C_LSL     = "C7"   # LSL
C_USL     = "C8"   # USL
# Section 2 — Mode & Data
C_MODE    = "C10"  # "Enter Manually" or "Use Data Worksheet"
C_XBAR_M  = "C12"  # Manual x̄
C_SIGMA_M = "C13"  # Manual σ
C_N_M     = "C14"  # Manual n
# Auto-calculated from Data sheet
C_XBAR_D  = "G5"   # = Data!G6 (mean)
C_SIGMA_D = "G6"   # = Data!G7 (std dev)
C_N_D     = "G7"   # = Data!G5 (count)
# Effective values (mode-switched)
C_XBAR    = "G9"   # Effective x̄
C_SIGMA   = "G10"  # Effective σ
C_N       = "G11"  # Effective n
# Targets
C_TARGET  = "C16"  # Target capability index
C_CL      = "C17"  # Confidence level
C_DP      = "C18"  # Decimal places
C_HYPO    = "C19"  # Hypothesis type


# ===========================================================================
# SHEET 1: Analysis
# ===========================================================================
def create_analysis_sheet(wb):
    ws = wb.active
    ws.title = "Analysis"
    ws.sheet_properties.tabColor = BLUE

    # Column widths
    for col, w in [("A",2),("B",28),("C",18),("D",3),("E",3),
                   ("F",28),("G",18),("H",3),("I",28),("J",18)]:
        ws.column_dimensions[col].width = w

    # === TITLE ===
    ws.merge_cells("B1:J1")
    ws.cell(row=1, column=2, value="Statistical Process Capability & Optimization Tool").font = TITLE_FONT
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")
    ws.merge_cells("B2:J2")
    ws.cell(row=2, column=2,
            value="Automotive Dimensional Capability Analysis — Single Characteristic | Excel Edition").font = SUBTITLE_FONT
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")

    # ====================== LEFT COLUMN (B-C): INPUTS ======================
    # --- Section 1: Specifications ---
    section_row(ws, 4, 2, 3, "I. SPECIFICATIONS")
    lv(ws, 5, 2, "Measurement Name", 3, value="Diameter A", is_input=True)
    lv(ws, 6, 2, "Target Mean (Tₘ)", 3, value=10.00, is_input=True, fmt="0.000")
    lv(ws, 7, 2, "Lower Spec Limit (LSL)", 3, value=9.90, is_input=True, fmt="0.000")
    lv(ws, 8, 2, "Upper Spec Limit (USL)", 3, value=10.10, is_input=True, fmt="0.000")

    # --- Section 2: Data & Goals ---
    section_row(ws, 9, 2, 3, "II. DATA & GOALS")

    # Mode dropdown
    lv(ws, 10, 2, "Data Input Mode", 3, value="Enter Manually", is_input=True)
    mode_dv = DataValidation(type="list", formula1='"Enter Manually,Use Data Worksheet"',
                              allow_blank=False, showDropDown=False)
    mode_dv.prompt = "Choose data source"
    mode_dv.promptTitle = "Mode"
    ws.add_data_validation(mode_dv)
    mode_dv.add(ws["C10"])

    # Subtitle
    ws.merge_cells("B11:C11")
    ws.cell(row=11, column=2,
            value="▼ Manual inputs (used when Mode = Enter Manually)").font = Font(
        name="Calibri", size=9, italic=True, color=GRAY)

    lv(ws, 12, 2, "x̄ (Measured Mean)", 3, value=10.00, is_input=True, fmt="0.00000")
    lv(ws, 13, 2, "σ (Standard Deviation)", 3, value=0.015, is_input=True, fmt="0.00000")
    lv(ws, 14, 2, "n (Sample Size)", 3, value=30, is_input=True, fmt="0")

    # --- Section 3: Targets & Settings ---
    section_row(ws, 15, 2, 3, "III. TARGETS & SETTINGS")
    lv(ws, 16, 2, "Target Capability Index", 3, value=1.67, is_input=True, fmt="0.00")
    lv(ws, 17, 2, "Confidence Level (%)", 3, value=95.0, is_input=True, fmt="0.0")
    lv(ws, 18, 2, "Decimal Places", 3, value=3, is_input=True, fmt="0")
    lv(ws, 19, 2, "Hypothesis Type", 3, value="Two-Sided", is_input=True)
    hypo_dv = DataValidation(type="list", formula1='"Two-Sided,Upper-Sided,Lower-Sided"',
                              allow_blank=False, showDropDown=False)
    ws.add_data_validation(hypo_dv)
    hypo_dv.add(ws["C19"])

    # ====================== MIDDLE COLUMN (F-G): EFFECTIVE DATA + RESULTS ======================

    # --- Effective Data (auto-switched by mode) ---
    section_row(ws, 4, 6, 7, "EFFECTIVE DATA (auto-selected by mode)")

    # Data from worksheet (auto-linked)
    lv(ws, 5, 6, "Worksheet x̄", 7, formula="=Data!G6", fmt="0.00000", fill=CALC_FILL)
    lv(ws, 6, 6, "Worksheet σ", 7, formula="=Data!G7", fmt="0.00000", fill=CALC_FILL)
    lv(ws, 7, 6, "Worksheet n", 7, formula="=Data!G5", fmt="0", fill=CALC_FILL)

    # Separator
    ws.cell(row=8, column=6, value="▼ Active values used in calculations:").font = Font(
        name="Calibri", size=9, italic=True, color=ORANGE)
    ws.merge_cells("F8:G8")

    # Effective x̄ — switches based on mode
    lv(ws, 9, 6, "x̄ (ACTIVE)", 7,
       formula=f'=IF({C_MODE}="Use Data Worksheet",G5,{C_XBAR_M})',
       fmt="0.00000", fill=RESULT_FILL,
       font=Font(name="Calibri", size=12, bold=True, color=DARK_BLUE))

    # Effective σ
    lv(ws, 10, 6, "σ (ACTIVE)", 7,
       formula=f'=IF({C_MODE}="Use Data Worksheet",G6,{C_SIGMA_M})',
       fmt="0.00000", fill=RESULT_FILL,
       font=Font(name="Calibri", size=12, bold=True, color=DARK_BLUE))

    # Effective n
    lv(ws, 11, 6, "n (ACTIVE)", 7,
       formula=f'=IF({C_MODE}="Use Data Worksheet",G7,{C_N_M})',
       fmt="0", fill=RESULT_FILL,
       font=Font(name="Calibri", size=12, bold=True, color=DARK_BLUE))

    # *** All result formulas now reference G9, G10, G11 ***
    xbar = "G9"
    sigma = "G10"
    n = "G11"

    # --- Section 4: Calculated Results ---
    section_row(ws, 13, 6, 7, "IV. CALCULATED RESULTS")

    lv(ws, 14, 6, "Tolerance (T = USL − LSL)", 7,
       formula=f"={C_USL}-{C_LSL}", fmt="0.000")
    lv(ws, 15, 6, "6σ Spread", 7,
       formula=f"=6*{sigma}", fmt="0.000")
    lv(ws, 16, 6, "8σ Spread", 7,
       formula=f"=8*{sigma}", fmt="0.000")
    lv(ws, 17, 6, "Cp (Potential Capability)", 7,
       formula=f'=IF({sigma}=0,"∞",({C_USL}-{C_LSL})/(6*{sigma}))', fmt="0.000")
    add_cpk_cond_fmt(ws, "G17")

    lv(ws, 18, 6, "Cpk (Actual Capability)", 7,
       formula=f'=IF({sigma}=0,"∞",MIN(({C_USL}-{xbar})/(3*{sigma}),({xbar}-{C_LSL})/(3*{sigma})))',
       fmt="0.000")
    add_cpk_cond_fmt(ws, "G18")

    lv(ws, 19, 6, "Required Shift (Δ = Tₘ − x̄)", 7,
       formula=f"={C_TM}-{xbar}", fmt="0.000")
    lv(ws, 20, 6, "Shift Direction", 7,
       formula=f'=IF({C_TM}-{xbar}=0,"✅ Centered",IF({C_TM}-{xbar}>0,"⬆ Shift UP","⬇ Shift DOWN"))')
    lv(ws, 21, 6, "Required Tolerance (for target)", 7,
       formula=f"=IF({sigma}=0,0,{C_TARGET}*6*{sigma})", fmt="0.000")
    lv(ws, 22, 6, "x̄ − 3σ", 7, formula=f"={xbar}-3*{sigma}", fmt="0.000")
    lv(ws, 23, 6, "x̄ + 3σ", 7, formula=f"={xbar}+3*{sigma}", fmt="0.000")
    lv(ws, 24, 6, "x̄ − 4σ", 7, formula=f"={xbar}-4*{sigma}", fmt="0.000")
    lv(ws, 25, 6, "x̄ + 4σ", 7, formula=f"={xbar}+4*{sigma}", fmt="0.000")

    # ====================== RIGHT COLUMN (I-J): PROBABILITY + HYPOTHESIS + VERDICT ======================

    # --- Section 5: Probability & Defects ---
    section_row(ws, 4, 9, 10, "V. PROBABILITY & DEFECTS")

    lv(ws, 5, 9, "P(x > USL)", 10,
       formula=f'=IF({sigma}=0,IF({xbar}>{C_USL},1,0),1-NORM.DIST({C_USL},{xbar},{sigma},TRUE))',
       fmt="0.0000%")
    lv(ws, 6, 9, "P(x < LSL)", 10,
       formula=f'=IF({sigma}=0,IF({xbar}<{C_LSL},1,0),NORM.DIST({C_LSL},{xbar},{sigma},TRUE))',
       fmt="0.0000%")
    lv(ws, 7, 9, "P(x < Tₘ)", 10,
       formula=f'=IF({sigma}=0,IF({xbar}<{C_TM},1,0),NORM.DIST({C_TM},{xbar},{sigma},TRUE))',
       fmt="0.00%")
    lv(ws, 8, 9, "PPM Above USL", 10, formula="=J5*1000000", fmt="#,##0.0")
    lv(ws, 9, 9, "PPM Below LSL", 10, formula="=J6*1000000", fmt="#,##0.0")
    lv(ws, 10, 9, "Total PPM", 10, formula="=J8+J9", fmt="#,##0.0",
       fill=WARN_FILL, font=Font(name="Calibri", size=11, bold=True))

    # --- Section 6: Hypothesis Test ---
    section_row(ws, 12, 9, 10, "VI. HYPOTHESIS TEST (μ vs Tₘ)")

    # SE
    lv(ws, 13, 9, "Standard Error (SE)", 10,
       formula=f"=IF({sigma}=0,0,{sigma}/SQRT({n}))", fmt="0.00000")
    # Z-stat
    lv(ws, 14, 9, "Z-statistic", 10,
       formula=f'=IF(J13=0,IF({xbar}={C_TM},0,999),({xbar}-{C_TM})/J13)', fmt="0.000")
    # p-value (adapts to hypothesis type)
    lv(ws, 15, 9, "p-value", 10,
       formula=f'=IF({C_HYPO}="Two-Sided",2*(1-NORM.S.DIST(ABS(J14),TRUE)),IF({C_HYPO}="Upper-Sided",1-NORM.S.DIST(J14,TRUE),NORM.S.DIST(J14,TRUE)))',
       fmt="0.0000")
    # Alpha
    lv(ws, 16, 9, "Alpha (α)", 10,
       formula=f"=1-{C_CL}/100", fmt="0.00")
    # Decision
    lv(ws, 17, 9, "Decision", 10,
       formula='=IF(J15<J16,"❌ Reject H₀ — Mean shifted","✅ Fail to Reject H₀ — On target")',
       font=Font(name="Calibri", size=10, bold=True))
    # CI
    lv(ws, 18, 9, "CI Lower Bound", 10,
       formula=f"={xbar}-NORM.S.INV(1-J16/2)*J13", fmt="0.0000")
    lv(ws, 19, 9, "CI Upper Bound", 10,
       formula=f"={xbar}+NORM.S.INV(1-J16/2)*J13", fmt="0.0000")

    # --- Section 7: Assessment Summary ---
    section_row(ws, 21, 9, 10, "VII. ASSESSMENT SUMMARY")

    # Centering
    lv(ws, 22, 9, "1. Centering", 10,
       formula=f'=IF({sigma}=0,IF({C_TM}-{xbar}=0,"✅ Perfectly centered (σ=0)","⚠ σ=0 but shift of "&TEXT({C_TM}-{xbar},"0.000")&" needed"),IF(ABS({C_TM}-{xbar})<{sigma}*0.05,"✅ Well-centered","⚠ Off-target by "&TEXT({C_TM}-{xbar},"0.000")))',
       font=Font(name="Calibri", size=10))
    # Capability
    lv(ws, 23, 9, "2. Capability", 10,
       formula=f'=IF({sigma}=0,"✅ Perfect (σ=0, index=∞)",IF(G18>={C_TARGET},"✅ Capable: Cpk "&TEXT(G18,"0.000")&" ≥ "&TEXT({C_TARGET},"0.00"),IF(G18>=1,"⚠ Marginal: Cpk "&TEXT(G18,"0.000"),"❌ Not Capable: Cpk "&TEXT(G18,"0.000"))))',
       font=Font(name="Calibri", size=10))
    # Robustness
    lv(ws, 24, 9, "3. Robustness", 10,
       formula=f'=IF({sigma}=0,"✅ Robust (σ=0)",IF(AND({C_LSL}<G24,{C_USL}>G25),"✅ Robust: ±4σ within specs",IF(AND({C_LSL}<G22,{C_USL}>G23),"⚠ Marginal: ±3σ OK, ±4σ NOT","❌ Not Robust: ±3σ breaches specs")))',
       font=Font(name="Calibri", size=10))
    # Tolerance
    lv(ws, 25, 9, "4. Tolerance", 10,
       formula=f'=IF({sigma}=0,"✅ Adequate (σ=0)",IF(G21<=G14,"✅ Adequate: Tol "&TEXT(G14,"0.000")&" ≥ required "&TEXT(G21,"0.000"),"❌ Inadequate: needs "&TEXT(G21,"0.000")&" but only "&TEXT(G14,"0.000")))',
       font=Font(name="Calibri", size=10))
    # Hypothesis
    lv(ws, 26, 9, "5. Hypothesis", 10,
       formula='=J17', font=Font(name="Calibri", size=10))

    # --- OVERALL VERDICT ---
    section_row(ws, 28, 2, 10, "")
    ws.merge_cells("B28:J28")
    vc = ws.cell(row=28, column=2)
    vc.value = f'=IF(OR(G18<1,G21>G14),"❌ ACTION REQUIRED — Process not capable",IF(OR(G18<{C_TARGET},ABS({C_TM}-{xbar})>{sigma}*0.05,J15<J16),"⚠ MARGINAL — Review recommendations","✅ PROCESS HEALTH: GOOD"))'
    vc.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")

    # Conditional formatting on verdict row
    ws.conditional_formatting.add("B28",
        FormulaRule(formula=['ISNUMBER(SEARCH("GOOD",B28))'],
                    fill=VERDICT_GREEN, font=Font(color=WHITE, bold=True, size=14)))
    ws.conditional_formatting.add("B28",
        FormulaRule(formula=['ISNUMBER(SEARCH("MARGINAL",B28))'],
                    fill=VERDICT_ORANGE, font=Font(color=WHITE, bold=True, size=14)))
    ws.conditional_formatting.add("B28",
        FormulaRule(formula=['ISNUMBER(SEARCH("ACTION",B28))'],
                    fill=VERDICT_RED, font=Font(color=WHITE, bold=True, size=14)))

    # --- Recommendations ---
    section_row(ws, 30, 2, 10, "RECOMMENDATIONS")
    ws.merge_cells("B31:J31")
    ws.cell(row=31, column=2).value = (
        f'=IF(ABS({C_TM}-{xbar})>{sigma}*0.05,"• Adjust process mean by "&TEXT({C_TM}-{xbar},"0.000")&" to center on target.","")&'
        f'IF(G18<{C_TARGET},CHAR(10)&"• Reduce variation (σ) to meet capability target of "&TEXT({C_TARGET},"0.00")&".","")'
        f'&IF(G21>G14,CHAR(10)&"• Widen tolerance or reduce σ. Required: "&TEXT(G21,"0.000")&", current: "&TEXT(G14,"0.000")&".","")'
        f'&IF(J15<J16,CHAR(10)&"• Hypothesis test shows significant mean shift (p="&TEXT(J15,"0.0000")&" < α="&TEXT(J16,"0.00")&").","")'
        f'&IF(AND(G18>={C_TARGET},ABS({C_TM}-{xbar})<={sigma}*0.05,G21<=G14,J15>=J16),"• Process meets all criteria. Monitor for stability.","")'
    )
    ws.cell(row=31, column=2).font = Font(name="Calibri", size=10)
    ws.cell(row=31, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[31].height = 70

    # Conditional formatting: highlight mode-disabled manual input cells
    ws.conditional_formatting.add("C12:C14",
        FormulaRule(formula=[f'$C$10="Use Data Worksheet"'],
                    fill=PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"),
                    font=Font(color="9CA3AF")))

    ws.sheet_view.showGridLines = False
    ws.print_area = "A1:J31"


# ===========================================================================
# SHEET 2: Data Worksheet
# ===========================================================================
def create_data_sheet(wb):
    ws = wb.create_sheet("Data")
    ws.sheet_properties.tabColor = GREEN

    for col, w in [("A",2),("B",8),("C",24),("D",18),("E",4),("F",26),("G",18)]:
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("B1:G1")
    ws.cell(row=1, column=2, value="Data Worksheet — Part-by-Part Entry").font = TITLE_FONT
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:G2")
    ws.cell(row=2, column=2,
            value=f'Enter DMC and measured Value below (max {MAX_DATA_ROWS} parts). Statistics auto-link to Analysis sheet.').font = SUBTITLE_FONT
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")

    # Characteristic info + capacity indicator
    ws.merge_cells("B3:D3")
    ws.cell(row=3, column=2,
            value='=Analysis!C5&" | Tₘ="&TEXT(Analysis!C6,"0.000")&" | LSL="&TEXT(Analysis!C7,"0.000")&" | USL="&TEXT(Analysis!C8,"0.000")').font = Font(
        name="Calibri", size=10, bold=True, color=BLUE)
    # Capacity bar
    ws.merge_cells("F3:G3")
    ws.cell(row=3, column=6,
            value=f'=G5&" of {MAX_DATA_ROWS} rows used ("&TEXT(G5/{MAX_DATA_ROWS},"0%")&" capacity)"').font = Font(
        name="Calibri", size=10, bold=True, color=ORANGE)
    ws.cell(row=3, column=6).alignment = CENTER

    # Data table headers
    row = 4
    header_row(ws, row, 2, ["#", "DMC / Serial Number", "Value"])

    # ===== SAMPLE TEMPLATE DATA (10 realistic parts) =====
    sample_data = [
        ("DMC-2024-001", 10.005),
        ("DMC-2024-002", 9.998),
        ("DMC-2024-003", 10.012),
        ("DMC-2024-004", 9.985),
        ("DMC-2024-005", 10.008),
        ("DMC-2024-006", 9.992),
        ("DMC-2024-007", 10.015),
        ("DMC-2024-008", 10.001),
        ("DMC-2024-009", 9.990),
        ("DMC-2024-010", 10.010),
    ]

    # Data entry rows
    for i in range(1, MAX_DATA_ROWS + 1):
        r = row + i
        ws.cell(row=r, column=2, value=i).font = Font(size=9, color=GRAY)
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).fill = INPUT_FILL
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).fill = INPUT_FILL
        ws.cell(row=r, column=4).number_format = "0.000"
        # Pre-fill sample data
        if i <= len(sample_data):
            dmc, val = sample_data[i - 1]
            ws.cell(row=r, column=3, value=dmc)
            ws.cell(row=r, column=4, value=val)
        # Alternating rows
        if i % 2 == 0:
            ws.cell(row=r, column=3).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
            ws.cell(row=r, column=4).fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    last_row = row + MAX_DATA_ROWS
    data_range = f"D5:D{last_row}"

    # --- Statistics Panel ---
    section_row(ws, 4, 6, 7, "AUTO-CALCULATED STATISTICS")

    stats = [
        ("Count (n)",       f"=COUNTA({data_range})",                                          "0"),
        ("Mean (x̄)",       f'=IF(G5>=2,AVERAGE({data_range}),"")',                            "0.00000"),
        ("Std Dev (σ)",     f'=IF(G5>=2,STDEV.S({data_range}),"")',                            "0.00000"),
        ("Min",             f'=IF(G5>=1,MIN({data_range}),"")',                                "0.000"),
        ("Max",             f'=IF(G5>=1,MAX({data_range}),"")',                                "0.000"),
        ("Range",           f'=IF(G5>=2,G9-G8,"")',                                            "0.000"),
        ("Median",          f'=IF(G5>=1,MEDIAN({data_range}),"")',                             "0.000"),
        ("Q1 (25th %ile)",  f'=IF(G5>=4,PERCENTILE({data_range},0.25),"")',                   "0.000"),
        ("Q3 (75th %ile)",  f'=IF(G5>=4,PERCENTILE({data_range},0.75),"")',                   "0.000"),
        ("IQR",             f'=IF(G5>=4,G13-G12,"")',                                          "0.000"),
        ("Skewness",        f'=IF(G5>=3,SKEW({data_range}),"")',                               "0.000"),
        ("Kurtosis",        f'=IF(G5>=4,KURT({data_range}),"")',                               "0.000"),
    ]

    for i, (label, formula, fmt) in enumerate(stats):
        lv(ws, 5 + i, 6, label, 7, formula=formula, fmt=fmt, fill=CALC_FILL)

    # Status
    status_row = 5 + len(stats) + 1
    ws.merge_cells(f"F{status_row}:G{status_row}")
    ws.cell(row=status_row, column=6).value = (
        f'=IF(G5>=2,"✅ "&G5&" of {MAX_DATA_ROWS} data points ready. Analysis auto-linked.","⚠ Need ≥2 data points ({MAX_DATA_ROWS} max allowed).")')
    ws.cell(row=status_row, column=6).font = Font(name="Calibri", size=10, bold=True, color=DARK_GREEN)

    # Quick Cp/Cpk preview
    preview_row = status_row + 2
    section_row(ws, preview_row, 6, 7, "QUICK PREVIEW (from worksheet data)")
    lv(ws, preview_row + 1, 6, "Cp (preview)", 7,
       formula=f'=IF(G7=""," — ", IF(G7=0,"∞",(Analysis!C8-Analysis!C7)/(6*G7)))', fmt="0.000")
    add_cpk_cond_fmt(ws, f"G{preview_row+1}")
    lv(ws, preview_row + 2, 6, "Cpk (preview)", 7,
       formula=f'=IF(G7=""," — ", IF(G7=0,"∞",MIN((Analysis!C8-G6)/(3*G7),(G6-Analysis!C7)/(3*G7))))', fmt="0.000")
    add_cpk_cond_fmt(ws, f"G{preview_row+2}")

    # Sample data note
    note_row = preview_row + 4
    ws.merge_cells(f"F{note_row}:G{note_row}")
    ws.cell(row=note_row, column=6,
            value="💡 10 sample data points pre-filled as template. Replace with your actual measurements.").font = Font(
        name="Calibri", size=9, italic=True, color=GRAY)

    ws.sheet_view.showGridLines = False


# ===========================================================================
# SHEET 3: Charts
# ===========================================================================
def create_charts_sheet(wb):
    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = ORANGE

    for col, w in [("A",2),("B",12),("C",14),("D",14),("E",14),("F",14),
                   ("G",2),("H",8),("I",14),("J",14),("K",14),("L",14),("M",14)]:
        ws.column_dimensions[col].width = w

    ws.merge_cells("B1:M1")
    ws.cell(row=1, column=2, value="Visualization — Process Capability & Control Charts").font = TITLE_FONT
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    # ======================== LEFT SIDE: Distribution Charts ========================

    # --- Bell curve helper data (z from -4 to +4 in 0.25 steps) ---
    section_row(ws, 3, 2, 5, "Normal Distribution Curve Data")

    header_row(ws, 4, 2, ["z", "x Value", "Current PDF", "Centered PDF"])

    z_values = [i * 0.25 for i in range(-16, 17)]  # -4 to +4 in 0.25 steps
    xbar_ref = "Analysis!G9"
    sigma_ref = "Analysis!G10"
    tm_ref = "Analysis!C6"

    for i, z in enumerate(z_values):
        r = 5 + i
        ws.cell(row=r, column=2, value=z).number_format = "0.00"
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3,
                value=f"={xbar_ref}+B{r}*{sigma_ref}").number_format = "0.0000"
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=4,
                value=f'=IF({sigma_ref}>0,(1/({sigma_ref}*SQRT(2*PI())))*EXP(-0.5*B{r}^2),0)').number_format = "0.0000"
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=5,
                value=f'=IF({sigma_ref}>0,(1/({sigma_ref}*SQRT(2*PI())))*EXP(-0.5*((C{r}-{tm_ref})/{sigma_ref})^2),0)').number_format = "0.0000"
        ws.cell(row=r, column=5).border = THIN_BORDER

    last_data_row = 5 + len(z_values) - 1

    # Chart 1: Process Distribution (Current + Centered)
    chart1 = LineChart()
    chart1.title = "1. Process Distribution — Current vs Centered"
    chart1.y_axis.title = "Density"
    chart1.x_axis.title = "Measurement Value"
    chart1.style = 10
    chart1.width = 24
    chart1.height = 15
    chart1.legend.position = "b"

    x_cats = Reference(ws, min_col=3, min_row=5, max_row=last_data_row)
    y_current = Reference(ws, min_col=4, min_row=4, max_row=last_data_row)
    y_centered = Reference(ws, min_col=5, min_row=4, max_row=last_data_row)

    chart1.add_data(y_current, titles_from_data=True)
    chart1.add_data(y_centered, titles_from_data=True)
    chart1.set_categories(x_cats)

    chart1.series[0].graphicalProperties.line.width = 25000
    chart1.series[0].graphicalProperties.line.solidFill = "B91C1C"  # Red curve
    chart1.series[1].graphicalProperties.line.solidFill = "007BC5"  # Blue curve
    chart1.series[1].graphicalProperties.line.dashStyle = "dash"

    ws.add_chart(chart1, "G3")

    # ======================== RIGHT SIDE: Control Charts ========================

    # --- I-MR (Individual & Moving Range) Control Chart Data ---
    ctrl_start = 3
    section_row(ws, ctrl_start, 8, 13, "I-MR Control Chart Data (from Data Worksheet)")
    header_row(ws, ctrl_start + 1, 8, ["#", "Value", "MR", "x̄ (CL)", "UCL", "LCL"])

    # I-MR constants: for individual values, d2=1.128, D4=3.267, D3=0
    # UCL = x̄ + 2.66*MR̄,  LCL = x̄ - 2.66*MR̄  (where 2.66 = 3/d2 = 3/1.128)
    # MR UCL = D4*MR̄ = 3.267*MR̄
    data_ref = "Data!D"  # Column D in Data sheet has the values
    n_ref = "Data!G5"     # Count
    mean_ref = "Data!G6"  # Mean

    CTRL_ROWS = 50  # Show up to 50 points in control chart
    for i in range(1, CTRL_ROWS + 1):
        r = ctrl_start + 1 + i
        data_row = 4 + i  # Row in Data sheet
        # Part #
        ws.cell(row=r, column=8, value=i).font = Font(size=9, color=GRAY)
        ws.cell(row=r, column=8).alignment = CENTER
        ws.cell(row=r, column=8).border = THIN_BORDER
        # Individual Value (linked from Data sheet)
        ws.cell(row=r, column=9,
                value=f'=IF({data_ref}{data_row}="""",NA(),{data_ref}{data_row})').number_format = "0.000"
        ws.cell(row=r, column=9).border = THIN_BORDER
        # Moving Range: |Xi - Xi-1|
        if i == 1:
            ws.cell(row=r, column=10, value='=NA()').number_format = "0.000"
        else:
            prev_data_row = data_row - 1
            ws.cell(row=r, column=10,
                    value=f'=IF(OR({data_ref}{data_row}="""",{data_ref}{prev_data_row}=""""),NA(),ABS({data_ref}{data_row}-{data_ref}{prev_data_row}))').number_format = "0.000"
        ws.cell(row=r, column=10).border = THIN_BORDER
        # Center Line (Mean)
        ws.cell(row=r, column=11,
                value=f'=IF({n_ref}>=2,{mean_ref},NA())').number_format = "0.000"
        ws.cell(row=r, column=11).border = THIN_BORDER
        # UCL = x̄ + 2.66 * MR̄ (where MR̄ = average of MR column)
        ws.cell(row=r, column=12,
                value=f'=IF({n_ref}>=2,{mean_ref}+2.66*AVERAGE(J{ctrl_start+3}:J{ctrl_start+1+CTRL_ROWS}),NA())').number_format = "0.000"
        ws.cell(row=r, column=12).border = THIN_BORDER
        # LCL = x̄ - 2.66 * MR̄
        ws.cell(row=r, column=13,
                value=f'=IF({n_ref}>=2,{mean_ref}-2.66*AVERAGE(J{ctrl_start+3}:J{ctrl_start+1+CTRL_ROWS}),NA())').number_format = "0.000"
        ws.cell(row=r, column=13).border = THIN_BORDER

    ctrl_last_row = ctrl_start + 1 + CTRL_ROWS

    # Chart 2: I-Chart (Individual Values with Control Limits)
    chart_i = LineChart()
    chart_i.title = "2. I-Chart — Individual Values with Control Limits"
    chart_i.y_axis.title = "Value"
    chart_i.x_axis.title = "Part Number"
    chart_i.style = 10
    chart_i.width = 24
    chart_i.height = 14
    chart_i.legend.position = "b"

    cats_ctrl = Reference(ws, min_col=8, min_row=ctrl_start+2, max_row=ctrl_last_row)
    y_val = Reference(ws, min_col=9, min_row=ctrl_start+1, max_row=ctrl_last_row)
    y_cl  = Reference(ws, min_col=11, min_row=ctrl_start+1, max_row=ctrl_last_row)
    y_ucl = Reference(ws, min_col=12, min_row=ctrl_start+1, max_row=ctrl_last_row)
    y_lcl = Reference(ws, min_col=13, min_row=ctrl_start+1, max_row=ctrl_last_row)

    chart_i.add_data(y_val, titles_from_data=True)
    chart_i.add_data(y_cl, titles_from_data=True)
    chart_i.add_data(y_ucl, titles_from_data=True)
    chart_i.add_data(y_lcl, titles_from_data=True)
    chart_i.set_categories(cats_ctrl)

    # Style series: Value=blue dots+line, CL=green solid, UCL=red dash, LCL=red dash
    chart_i.series[0].graphicalProperties.line.solidFill = "3B82F6"  # Blue
    chart_i.series[0].graphicalProperties.line.width = 18000
    chart_i.series[1].graphicalProperties.line.solidFill = "059669"  # Green CL
    chart_i.series[1].graphicalProperties.line.width = 18000
    chart_i.series[2].graphicalProperties.line.solidFill = "DC2626"  # Red UCL
    chart_i.series[2].graphicalProperties.line.dashStyle = "dash"
    chart_i.series[2].graphicalProperties.line.width = 15000
    chart_i.series[3].graphicalProperties.line.solidFill = "DC2626"  # Red LCL
    chart_i.series[3].graphicalProperties.line.dashStyle = "dash"
    chart_i.series[3].graphicalProperties.line.width = 15000

    ws.add_chart(chart_i, "G20")

    # Chart 3: MR Chart (Moving Range)
    chart_mr = LineChart()
    chart_mr.title = "3. Moving Range (MR) Chart"
    chart_mr.y_axis.title = "Moving Range"
    chart_mr.x_axis.title = "Part Number"
    chart_mr.style = 10
    chart_mr.width = 24
    chart_mr.height = 14
    chart_mr.legend.position = "b"

    y_mr = Reference(ws, min_col=10, min_row=ctrl_start+1, max_row=ctrl_last_row)
    chart_mr.add_data(y_mr, titles_from_data=True)
    chart_mr.set_categories(cats_ctrl)

    chart_mr.series[0].graphicalProperties.line.solidFill = "F97316"  # Orange
    chart_mr.series[0].graphicalProperties.line.width = 18000

    ws.add_chart(chart_mr, "G36")

    # ======================== BOTTOM: Capability & Spec Reference ========================

    # --- Capability bar chart data ---
    cap_start = last_data_row + 3
    section_row(ws, cap_start, 2, 5, "Capability Index Comparison")
    header_row(ws, cap_start + 1, 2, ["Metric", "Value", "Target", "Min (1.0)"])
    for col_idx in [2,3,4,5]:
        ws.cell(row=cap_start + 2, column=col_idx).border = THIN_BORDER
        ws.cell(row=cap_start + 3, column=col_idx).border = THIN_BORDER
    ws.cell(row=cap_start + 2, column=2, value="Cp")
    ws.cell(row=cap_start + 2, column=3, value="=Analysis!G17").number_format = "0.00"
    ws.cell(row=cap_start + 2, column=4, value=f"=Analysis!{C_TARGET}").number_format = "0.00"
    ws.cell(row=cap_start + 2, column=5, value=1.0).number_format = "0.00"
    ws.cell(row=cap_start + 3, column=2, value="Cpk")
    ws.cell(row=cap_start + 3, column=3, value="=Analysis!G18").number_format = "0.00"
    ws.cell(row=cap_start + 3, column=4, value=f"=Analysis!{C_TARGET}").number_format = "0.00"
    ws.cell(row=cap_start + 3, column=5, value=1.0).number_format = "0.00"

    chart2 = BarChart()
    chart2.title = "4. Capability Index vs Target"
    chart2.type = "col"
    chart2.style = 10
    chart2.width = 16
    chart2.height = 12
    chart2.y_axis.title = "Index Value"
    chart2.legend.position = "b"

    cats = Reference(ws, min_col=2, min_row=cap_start + 2, max_row=cap_start + 3)
    vals = Reference(ws, min_col=3, min_row=cap_start + 1, max_row=cap_start + 3)
    tgts = Reference(ws, min_col=4, min_row=cap_start + 1, max_row=cap_start + 3)
    mins = Reference(ws, min_col=5, min_row=cap_start + 1, max_row=cap_start + 3)
    chart2.add_data(vals, titles_from_data=True)
    chart2.add_data(tgts, titles_from_data=True)
    chart2.add_data(mins, titles_from_data=True)
    chart2.set_categories(cats)

    # Color bars
    chart2.series[0].graphicalProperties.solidFill = "3B82F6"  # Blue actual
    chart2.series[1].graphicalProperties.solidFill = "059669"  # Green target
    chart2.series[2].graphicalProperties.solidFill = "DC2626"  # Red minimum

    ws.add_chart(chart2, f"A{cap_start}")

    # Spec line reference
    ref_start = cap_start + 6
    section_row(ws, ref_start, 2, 5, "Specification Summary")
    lv(ws, ref_start+1, 2, "LSL", 3, formula="=Analysis!C7", fmt="0.000")
    lv(ws, ref_start+1, 4, "Tₘ", 5, formula="=Analysis!C6", fmt="0.000")
    lv(ws, ref_start+2, 2, "USL", 3, formula="=Analysis!C8", fmt="0.000")
    lv(ws, ref_start+2, 4, "Mean (x̄)", 5, formula="=Analysis!G9", fmt="0.000")
    lv(ws, ref_start+3, 2, "σ (ACTIVE)", 3, formula="=Analysis!G10", fmt="0.00000")
    lv(ws, ref_start+3, 4, "n (ACTIVE)", 5, formula="=Analysis!G11", fmt="0")
    lv(ws, ref_start+4, 2, "6σ Spread", 3, formula="=Analysis!G15", fmt="0.000")
    lv(ws, ref_start+4, 4, "8σ Spread", 5, formula="=Analysis!G16", fmt="0.000")

    ws.sheet_view.showGridLines = False


# ===========================================================================
# SHEET 4: History
# ===========================================================================
def create_history_sheet(wb):
    ws = wb.create_sheet("History")
    ws.sheet_properties.tabColor = "8B5CF6"

    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:O1")
    ws.cell(row=1, column=2, value="Analysis History Log").font = TITLE_FONT
    ws.cell(row=1, column=2).alignment = Alignment(horizontal="center")

    ws.merge_cells("B2:O2")
    ws.cell(row=2, column=2,
            value="Row 4 is auto-linked to the current Analysis. To save a run, copy-paste Row 4 as VALUES into the next empty row.").font = SUBTITLE_FONT
    ws.cell(row=2, column=2).alignment = Alignment(horizontal="center")

    headers = [
        "Date", "Name", "Mode", "Tₘ", "LSL", "USL", "x̄", "σ", "n",
        "Cp", "Cpk", "PPM Total", "Shift (Δ)", "Verdict"
    ]
    col_widths = [12, 16, 16, 10, 10, 10, 12, 12, 6, 10, 10, 12, 12, 36]

    for i, (h, w) in enumerate(zip(headers, col_widths)):
        col = i + 2
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Pre-linked first row
    formulas = [
        ('=TODAY()', "YYYY-MM-DD"),
        ('=Analysis!C5', None),
        ('=Analysis!C10', None),
        ('=Analysis!C6', "0.000"),
        ('=Analysis!C7', "0.000"),
        ('=Analysis!C8', "0.000"),
        ('=Analysis!G9', "0.00000"),
        ('=Analysis!G10', "0.00000"),
        ('=Analysis!G11', "0"),
        ('=Analysis!G17', "0.000"),
        ('=Analysis!G18', "0.000"),
        ('=Analysis!J10', "#,##0.0"),
        ('=Analysis!G19', "0.000"),
        ('=Analysis!B28', None),
    ]

    for i, (f, fmt) in enumerate(formulas):
        col = i + 2
        cell = ws.cell(row=4, column=col, value=f)
        cell.border = THIN_BORDER
        cell.alignment = CENTER
        cell.fill = RESULT_FILL
        if fmt:
            cell.number_format = fmt

    # Empty rows
    for row_idx in range(5, 55):
        for col in range(2, 16):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.fill = INPUT_FILL if row_idx % 2 == 0 else PatternFill(
                start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    add_cpk_cond_fmt(ws, "L4:L55")
    add_cpk_cond_fmt(ws, "K4:K55")

    ws.sheet_view.showGridLines = False


# ===========================================================================
# SHEET 5: Reference
# ===========================================================================
def create_reference_sheet(wb):
    ws = wb.create_sheet("Reference")
    ws.sheet_properties.tabColor = GRAY

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 90

    ws.cell(row=1, column=2, value="Reference Guide — Formulas, Definitions & How to Use").font = TITLE_FONT

    content = [
        ("", False),
        ("How to Use This Workbook", True),
        ("1. Enter specifications (Tₘ, LSL, USL) in the Analysis sheet (blue cells)", False),
        ("2. Choose Mode: 'Enter Manually' or 'Use Data Worksheet' (dropdown in C10)", False),
        ("3. If Manual: enter x̄, σ, n directly in the Analysis sheet", False),
        ("4. If Worksheet: enter part data in the Data sheet — mean & σ auto-link", False),
        ("5. All results, probability, hypothesis test, and verdict update INSTANTLY", False),
        ("6. View charts in the Charts sheet (bell curve + capability bars)", False),
        ("7. To save a run: go to History sheet, copy Row 4, paste as VALUES into Row 5+", False),
        ("", False),
        ("Core Capability Formulas", True),
        ("  Cp = (USL − LSL) / 6σ  — Potential capability if process is perfectly centered", False),
        ("  Cpk = min[(USL − x̄) / 3σ,  (x̄ − LSL) / 3σ]  — Actual capability with centering error", False),
        ("  Required Shift (Δ) = Tₘ − x̄  — How far to adjust the process mean", False),
        ("  Required Tolerance = Target Index × 6σ  — Minimum tolerance band needed", False),
        ("", False),
        ("Capability Index Interpretation (Automotive Standards)", True),
        ("  Cpk ≥ 1.67  →  ✅ GOOD — Capable (standard automotive target)", False),
        ("  1.33 ≤ Cpk < 1.67  →  ⚠ ACCEPTABLE — Meets minimum but below target", False),
        ("  1.00 ≤ Cpk < 1.33  →  ⚠ MARGINAL — High risk, improvement needed", False),
        ("  Cpk < 1.00  →  ❌ NOT CAPABLE — Produces significant defects", False),
        ("", False),
        ("Probability & PPM", True),
        ("  P(x > USL) — Probability a part exceeds upper spec limit", False),
        ("  P(x < LSL) — Probability a part falls below lower spec limit", False),
        ("  PPM = Probability × 1,000,000 — Parts Per Million defective", False),
        ("", False),
        ("Hypothesis Testing", True),
        ("  H₀: μ = Tₘ (process is on target)", False),
        ("  Z = (x̄ − Tₘ) / (σ / √n)", False),
        ("  p-value < α → Reject H₀ → Significant shift detected", False),
        ("  p-value ≥ α → Fail to Reject H₀ → No significant shift", False),
        ("  Two-Sided: tests if mean differs in either direction", False),
        ("  Upper-Sided: tests if mean is significantly above target", False),
        ("  Lower-Sided: tests if mean is significantly below target", False),
        ("", False),
        ("Robustness Assessment", True),
        ("  ROBUST: ±4σ spread contained within specification limits", False),
        ("  MARGINAL: ±3σ contained but ±4σ is NOT — low tolerance for future shifts", False),
        ("  NOT ROBUST: ±3σ breaches specification limits", False),
        ("", False),
        ("Spread Metrics", True),
        ("  6σ Spread — Contains ~99.73% of process output (±3σ)", False),
        ("  8σ Spread — Contains ~99.9937% of process output (±4σ)", False),
        ("", False),
        ("Color Coding", True),
        ("  🟢 Green cells = Good / Capable / Within target", False),
        ("  🟡 Yellow cells = Warning / Marginal", False),
        ("  🔴 Red cells = Bad / Action Required / Not capable", False),
        ("  🔵 Blue cells = Input cells (enter your data here)", False),
    ]

    row = 3
    for text, is_header in content:
        cell = ws.cell(row=row, column=2, value=text)
        if is_header:
            cell.font = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
            cell.fill = SECTION_FILL
        else:
            cell.font = Font(name="Calibri", size=10)
        cell.alignment = LEFT
        row += 1

    ws.sheet_view.showGridLines = False


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()

    create_analysis_sheet(wb)
    create_data_sheet(wb)
    create_charts_sheet(wb)
    create_history_sheet(wb)
    create_reference_sheet(wb)

    # Set Analysis as the default active sheet
    wb.active = 0

    wb.save(OUTPUT_FILE)
    print(f"✅ Excel tool generated: {OUTPUT_FILE}")
    print(f"   Sheets: Analysis | Data | Charts | History | Reference")
    print(f"   Features:")
    print(f"     - Mode toggle: Enter Manually / Use Data Worksheet (dropdown)")
    print(f"     - Auto-linked x̄, σ, n from Data sheet")
    print(f"     - Full assessment: Centering, Capability, Robustness, Tolerance, Hypothesis")
    print(f"     - Overall verdict with conditional color formatting")
    print(f"     - Recommendations formula")
    print(f"     - Bell curve (current + centered) chart")
    print(f"     - Capability bar chart (Cp/Cpk vs Target)")
    print(f"     - History log with pre-linked first row")
    print(f"     - {MAX_DATA_ROWS} data entry rows in worksheet")


if __name__ == "__main__":
    main()
