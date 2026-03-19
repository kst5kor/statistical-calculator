import streamlit as st
import pandas as pd
import numpy as np
import scipy.stats as stats
import math
import plotly.graph_objects as go
import io
import datetime
import re
from functools import lru_cache
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# --- Statistical Calculator Logic ---
# Ported from the 'statisticalCalculator' JavaScript object
class StatisticalCalculator:
    def erf(self, x):
        return math.erf(x)

    def standard_normal_cdf(self, z):
        return 0.5 * (1 + self.erf(z / math.sqrt(2)))

    def get_critical_value(self, cl, type):
        alpha = 1 - (cl / 100.0)
        if type == "Two-Sided":
            return stats.norm.ppf(1 - alpha / 2)
        elif type == "Upper-Sided":
            return stats.norm.ppf(1 - alpha)
        else:  # Lower-Sided
            return stats.norm.ppf(
                alpha
            )  # Note: This will be negative, handle in CI calculation

    def validate(self, params):
        LSL = params.get("lsl")
        USL = params.get("usl")
        s = params.get("s")
        n_samples = params.get("n_samples")
        confidenceLevel = params.get("confidence_level")
        distribution = params.get("distribution")
        Tm = params.get("tm")
        importedData = params.get("importedData", [])
        mode = params.get("mode")

        try:
            if any(v is None for v in [LSL, USL, Tm]):
                return "All specification values (Tm, LSL, USL) must be valid numbers."
            if USL <= LSL:
                return "USL must be greater than LSL."
            if mode == "manual" and (s is None or s < 0):
                return "Standard Deviation must be zero or positive for manual input."
            if n_samples is None or n_samples < 2:
                return "Sample Size (n) must be at least 2."
            if (
                confidenceLevel is None
                or confidenceLevel <= 0
                or confidenceLevel >= 100
            ):
                return "Confidence Level must be between 0 and 100."
            if distribution == "Lognormal" and (LSL <= 0 or USL <= 0 or Tm <= 0):
                return "For Lognormal distribution, all specification limits and the target mean must be positive."
            if mode == "import" and importedData:
                if len(importedData) < 2:
                    return "Imported data must contain at least 2 valid numeric points."
                if distribution == "Lognormal" and any(d <= 0 for d in importedData):
                    return "Lognormal distribution requires all imported data points to be positive."
            elif mode == "import" and not importedData:
                return "Import mode selected, but no data found in Data Worksheet."
            if mode == "import" and s is None:
                return "Could not calculate Standard Deviation from imported data. Check data format."
            if mode == "import" and s < 0:
                return "Calculated Standard Deviation from imported data cannot be negative."
            return None
        except Exception as e:
            return f"Validation error: {e}"

    class Normal:
        def pdf(self, x, mean, stdDev):
            if stdDev <= 0 or not np.isfinite(stdDev):
                return 0
            return (1 / (stdDev * np.sqrt(2 * np.pi))) * np.exp(
                -0.5 * ((x - mean) / stdDev) ** 2
            )

        def calculate(self, params):
            x_bar, s, USL, LSL, Tm = (
                params["x_bar"],
                params["s"],
                params["usl"],
                params["lsl"],
                params["tm"],
            )
            if s == 0:
                prob_above = 1 if x_bar > USL else 0
                prob_below = 1 if x_bar < LSL else 0
                prob_below_target = 1 if x_bar < Tm else 0
                cpk_s0 = np.inf if LSL <= x_bar <= USL else -np.inf
                return {
                    **params,
                    "T_drawing": USL - LSL,
                    "sixSigmaSpread": 0,
                    "Cp": np.inf,
                    "CpkCurrent": cpk_s0,
                    "prob_above": prob_above,
                    "prob_below": prob_below,
                    "prob_below_target": prob_below_target,
                }
            if s < 0 or not np.isfinite(s):
                return {
                    **params,
                    "T_drawing": np.nan,
                    "sixSigmaSpread": np.nan,
                    "Cp": np.nan,
                    "CpkCurrent": np.nan,
                    "prob_above": np.nan,
                    "prob_below": np.nan,
                    "prob_below_target": np.nan,
                }

            T_drawing = USL - LSL
            sixSigmaSpread = 6 * s
            Cp = T_drawing / sixSigmaSpread
            CpkCurrent = min((USL - x_bar) / (3 * s), (x_bar - LSL) / (3 * s))
            z_usl = (USL - x_bar) / s
            z_lsl = (LSL - x_bar) / s
            z_target = (Tm - x_bar) / s

            prob_above = 1 - stats.norm.cdf(z_usl)
            prob_below = stats.norm.cdf(z_lsl)
            prob_below_target = stats.norm.cdf(z_target)
            return {
                **params,
                "T_drawing": T_drawing,
                "sixSigmaSpread": sixSigmaSpread,
                "Cp": Cp,
                "CpkCurrent": CpkCurrent,
                "prob_above": prob_above,
                "prob_below": prob_below,
                "prob_below_target": prob_below_target,
            }

    class Lognormal:
        def pdf(self, x, mu_log, sigma_log):
            if x <= 0 or sigma_log <= 0 or not np.isfinite(sigma_log):
                return 0
            term1 = 1 / (x * sigma_log * np.sqrt(2 * np.pi))
            term2 = np.exp(-((np.log(x) - mu_log) ** 2) / (2 * sigma_log**2))
            return term1 * term2

        def calculate(self, params):
            x_bar, s, USL, LSL, Tm = (
                params["x_bar"],
                params["s"],
                params["usl"],
                params["lsl"],
                params["tm"],
            )

            if s == 0 or x_bar <= 0 or not np.isfinite(x_bar):
                prob_above = 1 if x_bar > USL else 0
                prob_below = 1 if x_bar < LSL else 0
                prob_below_target = 1 if x_bar < Tm else 0
                cpk_s0 = np.inf if LSL <= x_bar <= USL and x_bar > 0 else -np.inf
                return {
                    **params,
                    "T_drawing": np.nan,
                    "sixSigmaSpread": 0,
                    "Cp": np.inf,
                    "CpkCurrent": cpk_s0,
                    "prob_above": prob_above,
                    "prob_below": prob_below,
                    "prob_below_target": prob_below_target,
                    "mu_log": np.nan,
                    "sigma_log": 0,
                }
            if s < 0 or not np.isfinite(s):
                return {
                    **params,
                    "T_drawing": np.nan,
                    "sixSigmaSpread": np.nan,
                    "Cp": np.nan,
                    "CpkCurrent": np.nan,
                    "prob_above": np.nan,
                    "prob_below": np.nan,
                    "prob_below_target": np.nan,
                    "mu_log": np.nan,
                    "sigma_log": np.nan,
                }

            if LSL <= 0 or USL <= 0 or Tm <= 0:
                return {
                    **params,
                    "error": "LSL, USL, and Tm must be positive for Lognormal distribution.",
                }

            try:
                sigma_log_sq = np.log(1 + (s**2 / x_bar**2))
                sigma_log = np.sqrt(sigma_log_sq)
                mu_log = np.log(x_bar) - 0.5 * sigma_log_sq
            except ValueError:
                return {
                    **params,
                    "error": "Failed to calculate lognormal parameters. Check data.",
                }

            LSL_log, USL_log, Tm_log = np.log(LSL), np.log(USL), np.log(Tm)

            if sigma_log <= 0 or not np.isfinite(sigma_log):
                prob_above = 1 if x_bar > USL else 0
                prob_below = 1 if x_bar < LSL else 0
                prob_below_target = 1 if x_bar < Tm else 0
                cpk_s0 = np.inf if LSL <= x_bar <= USL else -np.inf
                return {
                    **params,
                    "T_drawing": np.nan,
                    "sixSigmaSpread": 0,
                    "Cp": np.inf,
                    "CpkCurrent": cpk_s0,
                    "prob_above": prob_above,
                    "prob_below": prob_below,
                    "prob_below_target": prob_below_target,
                    "mu_log": mu_log,
                    "sigma_log": sigma_log,
                }

            Cp = (USL_log - LSL_log) / (6 * sigma_log)
            CpkCurrent = min(
                (USL_log - mu_log) / (3 * sigma_log),
                (mu_log - LSL_log) / (3 * sigma_log),
            )

            z_usl_log = (USL_log - mu_log) / sigma_log
            z_lsl_log = (LSL_log - mu_log) / sigma_log
            z_target_log = (Tm_log - mu_log) / sigma_log

            prob_above = 1 - stats.norm.cdf(z_usl_log)
            prob_below = stats.norm.cdf(z_lsl_log)
            prob_below_target = stats.norm.cdf(z_target_log)

            return {
                **params,
                "T_drawing": USL - LSL,
                "sixSigmaSpread": 6 * s,
                "Cp": Cp,
                "CpkCurrent": CpkCurrent,
                "prob_above": prob_above,
                "prob_below": prob_below,
                "prob_below_target": prob_below_target,
                "mu_log": mu_log,
                "sigma_log": sigma_log,
            }

    def parse_raw_data(self, data_string):
        if not data_string:
            return []
        values = re.split(r"[\s,;\n]+", data_string.strip())
        return [float(v) for v in values if v and self.is_numeric(v)]

    def is_numeric(self, s):
        try:
            float(s)
            return True
        except (ValueError, TypeError):
            return False

    def calculate(self, inputs):
        params = {
            "tm": inputs.get("tm"),
            "lsl": inputs.get("lsl"),
            "usl": inputs.get("usl"),
            "target_index_value": inputs.get("target_index_value", 1.67),
            "target_index_type": inputs.get("target_index_type", "Cpk"),
            "confidence_level": inputs.get("confidence_level", 95.0),
            "distribution": "Normal",
            "dp": inputs.get("decimal_places", 3),
            "hypothesis_type": inputs.get("hypothesis_type", "Two-Sided"),
            "mode": inputs.get("mode", "manual"),
            "measurement_name": inputs.get("measurement_name", "Unnamed") or "Unnamed",
        }

        if params["mode"] == "import":
            data = self.parse_raw_data(inputs.get("raw_data", ""))
            params["importedData"] = data
            if len(data) >= 2:
                params["n_samples"] = len(data)
                params["x_bar"] = np.mean(data)
                params["s"] = np.std(data, ddof=1) if len(data) > 1 else 0
            else:
                params["n_samples"] = len(data)
                params["x_bar"] = np.nan
                params["s"] = np.nan
                params["importedData"] = []
        else:
            params["x_bar"] = inputs.get("x_bar")
            params["s"] = inputs.get("s")
            params["n_samples"] = inputs.get("n_samples")
            params["importedData"] = []

        validationError = self.validate(params)
        if validationError:
            return {**params, "error": validationError}

        # Check for non-numeric essential values
        essential_keys = ["tm", "lsl", "usl", "x_bar", "s", "n_samples"]
        if any(not self.is_numeric(params[k]) for k in essential_keys):
            if params.get("s") == 0 and all(
                self.is_numeric(params[k])
                for k in ["tm", "lsl", "usl", "x_bar", "n_samples"]
            ):
                pass  # s=0 is a valid case
            else:
                return {
                    **params,
                    "error": "Essential inputs (Tm, LSL, USL, x_bar, s, n) must be valid numbers.",
                }

        results = {}
        if params["distribution"] == "Lognormal":
            results = self.Lognormal().calculate(params)
        else:
            results = self.Normal().calculate(params)

        if results.get("error"):
            return results

        results["shiftValue"] = results["tm"] - results["x_bar"]
        results["newToleranceTotal"] = (
            results["target_index_value"] * 6 * results["s"]
            if results["s"] > 0 and np.isfinite(results["target_index_value"])
            else (0 if results["s"] == 0 else np.nan)
        )
        results["eightSigmaSpread"] = 8 * results["s"] if results["s"] >= 0 else np.nan
        results["minus3s"] = results["x_bar"] - 3 * results["s"]
        results["plus3s"] = results["x_bar"] + 3 * results["s"]
        results["minus4s"] = results["x_bar"] - 4 * results["s"]
        results["plus4s"] = results["x_bar"] + 4 * results["s"]
        results["ppm_above"] = results.get("prob_above", np.nan) * 1e6
        results["ppm_below"] = results.get("prob_below", np.nan) * 1e6

        alpha = 1 - (results["confidence_level"] / 100)

        if results["n_samples"] >= 2 and results["s"] >= 0:
            std_error = (
                results["s"] / np.sqrt(results["n_samples"]) if results["s"] > 0 else 0
            )
            z_stat = (
                (results["x_bar"] - results["tm"]) / std_error
                if std_error > 0
                else (
                    0
                    if results["x_bar"] == results["tm"]
                    else np.inf * np.sign(results["x_bar"] - results["tm"])
                )
            )

            p_value = np.nan
            if not np.isfinite(z_stat):
                p_value = 0.0
            elif results["hypothesis_type"] == "Two-Sided":
                p_value = 2 * (1 - stats.norm.cdf(abs(z_stat)))
            elif results["hypothesis_type"] == "Upper-Sided":  # mu > Tm
                p_value = 1 - stats.norm.cdf(z_stat)
            else:  # Lower-Sided, mu < Tm
                p_value = stats.norm.cdf(z_stat)

            # Use more precise ppf function
            criticalValue_ppf = (
                abs(stats.norm.ppf(alpha / 2))
                if results["hypothesis_type"] == "Two-Sided"
                else abs(stats.norm.ppf(alpha))
            )
            marginOfError = criticalValue_ppf * std_error

            if results["hypothesis_type"] == "Two-Sided":
                results["ci_lower"] = results["x_bar"] - marginOfError
                results["ci_upper"] = results["x_bar"] + marginOfError
            elif (
                results["hypothesis_type"] == "Upper-Sided"
            ):  # Test is mu > Tm, CI is for mu
                results["ci_lower"] = results["x_bar"] - marginOfError  # One-sided CI
                results["ci_upper"] = np.inf
            else:  # Lower-Sided
                results["ci_lower"] = -np.inf
                results["ci_upper"] = results["x_bar"] + marginOfError  # One-sided CI

            results["hypothesisResult"] = {
                "z_stat": z_stat,
                "p_value": p_value,
                "alpha": alpha,
            }
        else:
            results["ci_lower"] = np.nan
            results["ci_upper"] = np.nan
            results["hypothesisResult"] = {
                "z_stat": np.nan,
                "p_value": np.nan,
                "alpha": alpha,
            }

        return {**results, "error": None}


# --- Plotting Logic ---
# Ported from 'plotManager'
class PlotManager:
    # Interactive plot configuration
    PLOT_CONFIG = {
        "displayModeBar": True,
        "displaylogo": False,
        "modeBarButtonsToAdd": ["drawline", "eraseshape"],
        "modeBarButtonsToRemove": ["lasso2d", "select2d"],
        "toImageButtonOptions": {
            "format": "png",
            "filename": "capability_chart",
            "height": 600,
            "width": 1000,
            "scale": 2,
        },
        "scrollZoom": True,
    }

    def generate_pdf_data(self, dist_type, params, x_min, x_max, points=200):
        x = np.linspace(x_min, x_max, points)
        y = np.zeros_like(x)

        calc = StatisticalCalculator()

        if dist_type == "Normal" and params.get("stdDev", 0) > 0:
            y = [calc.Normal().pdf(val, params["mean"], params["stdDev"]) for val in x]
        elif dist_type == "Lognormal" and params.get("sigma_log", 0) > 0:
            y = [
                calc.Lognormal().pdf(val, params["mu_log"], params["sigma_log"])
                for val in x
            ]

        return x, np.nan_to_num(y)

    def update_plots(self, results):
        LSL, USL, x_bar, s, Tm, _target_index_value, dp = (
            results.get("lsl"),
            results.get("usl"),
            results.get("x_bar"),
            results.get("s"),
            results.get("tm"),
            results.get("target_index_value"),
            results.get("dp"),
        )
        ci_lower, ci_upper, confidenceLevel, distribution = (
            results.get("ci_lower"),
            results.get("ci_upper"),
            results.get("confidence_level"),
            results.get("distribution"),
        )
        mu_log, sigma_log, importedData = (
            results.get("mu_log"),
            results.get("sigma_log"),
            results.get("importedData", []),
        )

        cannot_plot = (
            any(not np.isfinite(v) for v in [LSL, USL, x_bar, Tm])
            or s < 0
            or not np.isfinite(s)
        )
        if cannot_plot:
            return None, None, None  # Return empty figures

        newToleranceTotal = results.get("newToleranceTotal", np.nan)
        newLSL = (
            Tm - (newToleranceTotal / 2) if np.isfinite(newToleranceTotal) else np.nan
        )
        newUSL = (
            Tm + (newToleranceTotal / 2) if np.isfinite(newToleranceTotal) else np.nan
        )

        data_min = (
            min(importedData)
            if importedData
            else (x_bar - 4.5 * s if s > 0 else x_bar - 1)
        )
        data_max = (
            max(importedData)
            if importedData
            else (x_bar + 4.5 * s if s > 0 else x_bar + 1)
        )

        x_points = [
            LSL,
            USL,
            newLSL,
            newUSL,
            x_bar,
            Tm,
            ci_lower,
            ci_upper,
            x_bar - 4.5 * s if s > 0 else None,
            x_bar + 4.5 * s if s > 0 else None,
            data_min,
            data_max,
        ]

        finite_x = [p for p in x_points if p is not None and np.isfinite(p)]

        if not finite_x:
            x_min, x_max = Tm - 1, Tm + 1
        else:
            raw_min, raw_max = min(finite_x), max(finite_x)
            range_val = raw_max - raw_min
            min_range = max(
                s * 0.5 if s > 0 else 0.1,
                abs(Tm - x_bar) or 0.1,
                (USL - LSL) * 0.1 or 0.1,
                0.1,
            )
            if range_val < min_range or range_val == 0:
                range_val = min_range
            x_min = raw_min - range_val * 0.2
            x_max = raw_max + range_val * 0.2

        pdf_data_before_x, pdf_data_before_y = [], []
        pdf_data_after_x, pdf_data_after_y = [], []
        max_pdf_y = 1

        if s > 0 and np.isfinite(s):
            if distribution == "Lognormal":
                pdf_params = {"mu_log": mu_log, "sigma_log": sigma_log}
                pdf_params_centered = {
                    "mu_log": np.log(Tm) - 0.5 * (sigma_log**2),
                    "sigma_log": sigma_log,
                }
            else:
                pdf_params = {"mean": x_bar, "stdDev": s}
                pdf_params_centered = {"mean": Tm, "stdDev": s}

            pdf_data_before_x, pdf_data_before_y = self.generate_pdf_data(
                distribution, pdf_params, x_min, x_max
            )
            pdf_data_after_x, pdf_data_after_y = self.generate_pdf_data(
                distribution, pdf_params_centered, x_min, x_max
            )

            valid_y = [
                y
                for y in np.concatenate((pdf_data_before_y, pdf_data_after_y))
                if np.isfinite(y) and y > 0
            ]
            if valid_y:
                max_pdf_y = max(valid_y) * 1.1

        # Theme-adaptive font color (readable in both light and dark mode)
        _fc = "#8b95a5"
        layout_defaults = {
            "xaxis": {
                "title": {"text": "Measurement Value", "font": {"color": _fc, "size": 11}},
                "range": [x_min, x_max],
                "zeroline": False,
                "tickformat": f".{dp}f",
                "tickfont": {"size": 10, "color": _fc},
                "gridcolor": "rgba(128,128,128,0.15)",
                "linecolor": "rgba(128,128,128,0.25)",
                "showspikes": True,
                "spikemode": "across",
                "spikesnap": "cursor",
                "spikecolor": "#888",
                "spikethickness": 0.5,
                "spikedash": "dot",
            },
            "yaxis": {
                "title": {"text": "Density" if s > 0 else "", "font": {"color": _fc, "size": 11}},
                "tickformat": ".2f" if s > 0 else "",
                "fixedrange": False,
                "range": [0, max_pdf_y],
                "tickfont": {"size": 10, "color": _fc},
                "showticklabels": bool(s > 0),
                "gridcolor": "rgba(128,128,128,0.15)",
                "linecolor": "rgba(128,128,128,0.25)",
                "showspikes": True,
                "spikemode": "across",
                "spikesnap": "cursor",
                "spikecolor": "#888",
                "spikethickness": 0.5,
                "spikedash": "dot",
            },
            "height": 380,
            "margin": {"t": 55, "b": 65, "l": 55, "r": 25},
            "showlegend": True,
            "legend": {
                "orientation": "h",
                "y": -0.22,
                "x": 0.5,
                "xanchor": "center",
                "bgcolor": "rgba(128,128,128,0.08)",
                "bordercolor": "rgba(128,128,128,0.2)",
                "borderwidth": 1,
                "font": {"size": 10, "color": _fc},
            },
            "hovermode": "x unified",
            "hoverlabel": {
                "font_size": 11,
                "namelength": -1,
                "bgcolor": "rgba(30,41,59,0.92)",
                "font_color": "#e2e8f0",
                "bordercolor": "rgba(128,128,128,0.3)",
            },
            "dragmode": "zoom",
            "modebar": {
                "orientation": "v",
                "bgcolor": "rgba(0,0,0,0)",
                "color": _fc,
            },
            "paper_bgcolor": "rgba(0,0,0,0)",
            "plot_bgcolor": "rgba(0,0,0,0)",
            "font": {"color": _fc},
        }
        # Note: PLOT_CONFIG is now defined at class level

        # Plot 1: Current Process
        fig_before = go.Figure()
        if s > 0:
            fig_before.add_trace(
                go.Scatter(
                    x=pdf_data_before_x,
                    y=pdf_data_before_y,
                    mode="lines",
                    name=f"Current PDF (x̄={x_bar:.{dp}f})",
                    fill="tozeroy",
                    fillcolor="rgba(185, 28, 28, 0.1)",
                    line={"color": "#B91C1C", "width": 2},
                )
            )
        fig_before.add_trace(
            go.Scatter(
                x=[LSL, LSL],
                y=[0, max_pdf_y * 0.95],
                mode="lines",
                name="LSL",
                line={"color": "#047857", "dash": "dash", "width": 1.5},
            )
        )
        fig_before.add_trace(
            go.Scatter(
                x=[USL, USL],
                y=[0, max_pdf_y * 0.95],
                mode="lines",
                name="USL",
                line={"color": "#047857", "dash": "dash", "width": 1.5},
            )
        )
        fig_before.add_trace(
            go.Scatter(
                x=[x_bar, x_bar],
                y=[0, max_pdf_y * 0.9],
                mode="lines",
                name="Mean (x̄)",
                line={
                    "color": "#DC2626",
                    "width": 3 if s == 0 else 1.5,
                    "dash": "solid",
                },
            )
        )
        fig_before.add_trace(
            go.Scatter(
                x=[Tm, Tm],
                y=[0, max_pdf_y * 0.8],
                mode="lines",
                name="Target (Tm)",
                line={"color": "#4B5563", "dash": "dot", "width": 1.5},
            )
        )

        shapes_before = [
            {
                "type": "rect",
                "xref": "x",
                "yref": "paper",
                "x0": x_min,
                "y0": 0,
                "x1": LSL,
                "y1": 1,
                "fillcolor": "rgba(239, 68, 68, 0.15)",
                "line": {"width": 0},
                "layer": "below",
            },
            {
                "type": "rect",
                "xref": "x",
                "yref": "paper",
                "x0": USL,
                "y0": 0,
                "x1": x_max,
                "y1": 1,
                "fillcolor": "rgba(239, 68, 68, 0.15)",
                "line": {"width": 0},
                "layer": "below",
            },
        ]
        annotations_before = []
        if s > 0 and np.isfinite(ci_lower) and np.isfinite(ci_upper):
            shapes_before.append(
                {
                    "type": "line",
                    "xref": "x",
                    "yref": "paper",
                    "x0": ci_lower,
                    "y0": 0.05,
                    "x1": ci_upper,
                    "y1": 0.05,
                    "line": {"color": "#F97316", "width": 4},
                    "layer": "above",
                }
            )
            annotations_before.append(
                {
                    "x": (ci_lower + ci_upper) / 2,
                    "y": 0.02,
                    "xref": "x",
                    "yref": "paper",
                    "text": f"{confidenceLevel}% CI",
                    "showarrow": False,
                    "font": {"size": 9, "color": "#F97316"},
                    "yanchor": "top",
                }
            )

        ci_text = (
            f"[{ci_lower:.{dp}f}, {ci_upper:.{dp}f}]"
            if np.isfinite(ci_lower) and np.isfinite(ci_upper)
            else (
                f"[{ci_lower:.{dp}f}, +∞)"
                if np.isfinite(ci_lower)
                else f"(-∞, {ci_upper:.{dp}f}]"
                if np.isfinite(ci_upper)
                else ""
            )
        )
        title_before = f"1. Current Process Distribution {'& CI (' + ci_text + ')' if s > 0 and ci_text else ('(σ=0)' if s == 0 else '')}"
        fig_before.update_layout(
            **layout_defaults,
            title={"text": title_before, "font": {"size": 12}},
            shapes=shapes_before,
            annotations=annotations_before,
        )

        # Plot 2: Centered Process
        fig_after = go.Figure()
        if s > 0:
            fig_after.add_trace(
                go.Scatter(
                    x=pdf_data_after_x,
                    y=pdf_data_after_y,
                    mode="lines",
                    name=f"Centered PDF (at Tm={Tm:.{dp}f})",
                    fill="tozeroy",
                    fillcolor="rgba(0, 123, 197, 0.1)",
                    line={"color": "#007BC5", "width": 2},
                )
            )
        if np.isfinite(newLSL):
            fig_after.add_trace(
                go.Scatter(
                    x=[newLSL, newLSL],
                    y=[0, max_pdf_y * 0.95],
                    mode="lines",
                    name="Req. LSL",
                    line={"color": "#004A86", "width": 1.5, "dash": "dot"},
                )
            )
        if np.isfinite(newUSL):
            fig_after.add_trace(
                go.Scatter(
                    x=[newUSL, newUSL],
                    y=[0, max_pdf_y * 0.95],
                    mode="lines",
                    name="Req. USL",
                    line={"color": "#004A86", "width": 1.5, "dash": "dot"},
                )
            )

        fig_after.add_trace(
            go.Scatter(
                x=[Tm, Tm],
                y=[0, max_pdf_y * 0.9],
                mode="lines",
                name="Target (Tm)",
                line={
                    "color": "#007BC5",
                    "width": 3 if s == 0 else 1.5,
                    "dash": "solid",
                },
            )
        )
        fig_after.add_trace(
            go.Scatter(
                x=[LSL, LSL],
                y=[0, max_pdf_y * 0.95],
                mode="lines",
                name="Orig. LSL",
                line={"color": "rgba(4, 120, 87, 0.5)", "dash": "dash", "width": 1},
            )
        )
        fig_after.add_trace(
            go.Scatter(
                x=[USL, USL],
                y=[0, max_pdf_y * 0.95],
                mode="lines",
                name="Orig. USL",
                line={"color": "rgba(4, 120, 87, 0.5)", "dash": "dash", "width": 1},
            )
        )

        shapes_after = []
        if np.isfinite(newLSL) and np.isfinite(newUSL):
            shapes_after.append(
                {
                    "type": "rect",
                    "xref": "x",
                    "yref": "paper",
                    "x0": x_min,
                    "y0": 0,
                    "x1": newLSL,
                    "y1": 1,
                    "fillcolor": "rgba(0, 123, 197, 0.1)",
                    "line": {"width": 0},
                    "layer": "below",
                }
            )
            shapes_after.append(
                {
                    "type": "rect",
                    "xref": "x",
                    "yref": "paper",
                    "x0": newUSL,
                    "y0": 0,
                    "x1": x_max,
                    "y1": 1,
                    "fillcolor": "rgba(0, 123, 197, 0.1)",
                    "line": {"width": 0},
                    "layer": "below",
                }
            )

        title_after = (
            f"2. Centered Process vs. Required Specs (Tol: {newToleranceTotal:.{dp}f})"
            if np.isfinite(newToleranceTotal)
            else (
                "2. Centered Process (σ=0)"
                if s == 0
                else "2. Centered Process Distribution"
            )
        )
        fig_after.update_layout(
            **layout_defaults,
            title={"text": title_after, "font": {"size": 12}},
            shapes=shapes_after,
        )

        # Plot 3: Frequency Histogram
        fig_hist = None
        if importedData and len(importedData) >= 2:
            fig_hist = go.Figure()
            fig_hist.add_trace(
                go.Histogram(
                    x=importedData,
                    name="Data Count",
                    marker={
                        "color": "rgba(0, 123, 197, 0.7)",
                        "line": {"color": "rgba(0, 70, 130, 0.8)", "width": 0.5},
                    },
                )
            )

            shapes_hist = [
                {
                    "type": "line",
                    "x0": x_bar,
                    "x1": x_bar,
                    "y0": 0,
                    "y1": 1,
                    "yref": "paper",
                    "line": {"color": "#DC2626", "width": 1.5, "dash": "dash"},
                },
                {
                    "type": "line",
                    "x0": LSL,
                    "x1": LSL,
                    "y0": 0,
                    "y1": 1,
                    "yref": "paper",
                    "line": {"color": "#059669", "width": 1.5, "dash": "dot"},
                },
                {
                    "type": "line",
                    "x0": USL,
                    "x1": USL,
                    "y0": 0,
                    "y1": 1,
                    "yref": "paper",
                    "line": {"color": "#059669", "width": 1.5, "dash": "dot"},
                },
            ]
            annotations_hist = [
                {
                    "x": x_bar,
                    "y": 1.02,
                    "yref": "paper",
                    "text": "Mean",
                    "showarrow": False,
                    "font": {"size": 10, "color": "#DC2626"},
                },
                {
                    "x": LSL,
                    "y": 1.02,
                    "yref": "paper",
                    "text": "LSL",
                    "showarrow": False,
                    "font": {"size": 10, "color": "#059669"},
                    "xanchor": "right",
                },
                {
                    "x": USL,
                    "y": 1.02,
                    "yref": "paper",
                    "text": "USL",
                    "showarrow": False,
                    "font": {"size": 10, "color": "#059669"},
                    "xanchor": "left",
                },
            ]

            fig_hist.update_layout(
                title={"text": "3. Data Frequency Distribution", "font": {"size": 12, "color": _fc}},
                xaxis={
                    "title": {"text": "Value", "font": {"color": _fc, "size": 11}},
                    "range": [x_min, x_max],
                    "zeroline": False,
                    "tickfont": {"size": 10, "color": _fc},
                    "gridcolor": "rgba(128,128,128,0.15)",
                    "linecolor": "rgba(128,128,128,0.25)",
                },
                yaxis={
                    "title": {"text": "Frequency (Count)", "font": {"color": _fc, "size": 11}},
                    "fixedrange": True,
                    "tickfont": {"size": 10, "color": _fc},
                    "gridcolor": "rgba(128,128,128,0.15)",
                    "linecolor": "rgba(128,128,128,0.25)",
                },
                height=380,
                bargap=0.05,
                shapes=shapes_hist,
                annotations=annotations_hist,
                margin={"t": 55, "b": 65, "l": 55, "r": 25},
                showlegend=True,
                legend={
                    "orientation": "h",
                    "y": -0.22,
                    "x": 0.5,
                    "xanchor": "center",
                    "bgcolor": "rgba(128,128,128,0.08)",
                    "bordercolor": "rgba(128,128,128,0.2)",
                    "borderwidth": 1,
                    "font": {"size": 10, "color": _fc},
                },
                hovermode="x unified",
                hoverlabel={
                    "font_size": 11,
                    "namelength": -1,
                    "bgcolor": "rgba(30,41,59,0.92)",
                    "font_color": "#e2e8f0",
                    "bordercolor": "rgba(128,128,128,0.3)",
                },
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                font={"color": _fc},
            )

        return fig_before, fig_after, fig_hist


# --- Summary Panel Logic ---
# Ported from 'updateSummaryPanel'
def get_summary_panel_content(results):
    shiftValue = results.get("shiftValue", np.nan)
    s = results.get("s", np.nan)
    CpkCurrent = results.get("CpkCurrent", np.nan)
    target_index_value = results.get("target_index_value", np.nan)
    LSL, USL = results.get("lsl", np.nan), results.get("usl", np.nan)
    minus3s, plus3s = results.get("minus3s", np.nan), results.get("plus3s", np.nan)
    minus4s, plus4s = results.get("minus4s", np.nan), results.get("plus4s", np.nan)
    newToleranceTotal = results.get("newToleranceTotal", np.nan)
    T_drawing = results.get("T_drawing", np.nan)
    dp = results.get("dp", 3)
    hypothesisResult = results.get("hypothesisResult", {})

    calculation_invalid = (
        not np.isfinite(shiftValue)
        or (not np.isfinite(CpkCurrent) and s != 0)
        or not np.isfinite(newToleranceTotal)
        or not np.isfinite(s)
        or s < 0
    )

    if calculation_invalid:
        return {
            "verdict": "INVALID INPUTS",
            "verdict_color": "red",
            "centering": "Calculation failed due to invalid or incomplete inputs (e.g., negative Std Dev, NaN values).",
            "capability": "",
            "robustness": "",
            "robustness_class": "status-red",
            "tolerance": "",
            "hypothesis": "",
            "recommendations": [
                "<li>Enter valid numeric inputs to see recommendations. Ensure Standard Deviation is not negative.</li>"
            ],
        }

    recommendations = []
    is_good = True
    is_marginal = False

    # Centering
    if s == 0:
        centering_text = f'<span style="color: green; font-weight: bold;">Excellent:</span> Process has zero variation and is centered{f" (but requires shift of {shiftValue:.{dp}f})." if shiftValue != 0 else "."}'
        if shiftValue != 0:
            recommendations.append(
                f"Adjust process mean by <b>{shiftValue:.{dp}f}</b> to align with T<sub>m</sub>."
            )
            is_marginal = True
    elif abs(shiftValue) < (s * 0.05):
        centering_text = '<span style="color: green; font-weight: bold;">Excellent:</span> Process is well-centered.'
    else:
        centering_text = f'<span style="color: orange; font-weight: bold;">Needs Adjustment:</span> Mean is off-target by <b>{shiftValue:.{dp}f}</b>. Adjustment of <b>{abs(shiftValue):.{dp}f} {"UP (+)" if shiftValue < 0 else "DOWN (-)"}</b> is required.'
        recommendations.append(
            f"Adjust process mean by <b>{shiftValue:.{dp}f}</b> to align with T<sub>m</sub>."
        )
        is_marginal = True

    # Capability
    if s == 0:
        capability_text = '<span style="color: green; font-weight: bold;">Perfect Capability (σ=0):</span> Index is effectively infinite (∞).'
    elif np.isfinite(CpkCurrent) and CpkCurrent >= target_index_value:
        capability_text = f'<span style="color: green; font-weight: bold;">Capable:</span> Current index of <b>{CpkCurrent:.{dp}f}</b> meets target of <b>{target_index_value:.2f}</b>.'
    elif np.isfinite(CpkCurrent) and CpkCurrent >= 1.33:
        capability_text = f'<span style="color: orange; font-weight: bold;">Marginally Capable:</span> Index of <b>{CpkCurrent:.{dp}f}</b> is acceptable but below target ({target_index_value:.2f}).'
        recommendations.append(
            "Improve stability or reduce variation (σ) to meet capability target."
        )
        is_marginal = True
    else:
        cpk_display = f"{CpkCurrent:.{dp}f}" if np.isfinite(CpkCurrent) else "N/A"
        capability_text = f'<span style="color: red; font-weight: bold;">Not Capable:</span> Index of <b>{cpk_display}</b> is below target ({target_index_value:.2f}). High risk of defects.'
        recommendations.append(
            "Urgent action required to reduce variation (σ) and/or re-center mean."
        )
        is_good = False

    # Robustness
    robustness_text = ""
    robustness_class = ""
    if s == 0:
        robustness_text = "ROBUST: Process has zero variation."
        robustness_class = "status-green"
    elif all(np.isfinite(v) for v in [LSL, USL, minus3s, plus3s, minus4s, plus4s]):
        if LSL < minus3s and USL > plus3s:
            if LSL < minus4s and USL > plus4s:
                robustness_text = "ROBUST: The ±4σ process spread is contained within specification limits."
                robustness_class = "status-green"
            else:
                robustness_text = "MARGINAL: The ±3σ spread is contained, but ±4σ is NOT. Low tolerance for future shifts."
                robustness_class = "status-yellow"
        else:
            robustness_text = (
                "NOT ROBUST: The ±3σ process spread breaches the specification limits."
            )
            robustness_class = "status-red"
    else:
        robustness_text = "Robustness check skipped due to invalid limits/spread."

    # Tolerance
    if s == 0:
        tolerance_text = '<span style="color: green; font-weight: bold;">Adequate:</span> Zero variation requires zero tolerance.'
    elif np.isfinite(T_drawing) and newToleranceTotal <= T_drawing:
        tolerance_text = f'<span style="color: green; font-weight: bold;">Adequate:</span> Current tolerance of <b>{T_drawing:.{dp}f}</b> is sufficient.'
    elif np.isfinite(T_drawing):
        tolerance_text = f'<span style="color: red; font-weight: bold;">Inadequate:</span> Tolerance is too tight. Requires minimum of <b>{newToleranceTotal:.{dp}f}</b>.'
        recommendations.append(
            "Widen specification range or fundamentally reduce process variation (σ)."
        )
        is_good = False
    else:
        tolerance_text = "Tolerance check skipped due to invalid limits."

    # Hypothesis Test
    p_value = hypothesisResult.get("p_value", np.nan)
    alpha = hypothesisResult.get("alpha", np.nan)
    z_stat = hypothesisResult.get("z_stat", np.nan)

    if np.isfinite(p_value) and np.isfinite(alpha) and np.isfinite(z_stat):
        if p_value < alpha:
            hypothesis_text = f'<span style="color: orange; font-weight: bold;">Reject H₀:</span> With a p-value of <b>{p_value:.3e}</b> (which is < α={alpha:.2f}), there is significant evidence that the process mean has shifted from the target.'
            is_marginal = True
        else:
            hypothesis_text = f'<span style="color: green; font-weight: bold;">Fail to Reject H₀:</span> With a p-value of <b>{p_value:.3e}</b> (which is >= α={alpha:.2f}), there is no significant evidence that the mean has shifted from the target.'
        hypothesis_text += f" (Z-statistic: {z_stat:.3f})"
    else:
        hypothesis_text = "Hypothesis test skipped (requires n>=2 and valid inputs)."

    # Final Verdict
    if not is_good:
        verdict_text = "ACTION REQUIRED"
        verdict_color = "red"
    elif is_marginal:
        verdict_text = "MARGINAL"
        verdict_color = "orange"
    else:
        verdict_text = "PROCESS HEALTH: GOOD"
        verdict_color = "green"

    if not recommendations:
        recommendations.append(
            "Process appears to meet target criteria based on current data. Monitor for stability."
        )

    return {
        "verdict": verdict_text,
        "verdict_color": verdict_color,
        "centering": centering_text,
        "capability": capability_text,
        "robustness": robustness_text,
        "robustness_class": robustness_class,  # This part is tricky to style in st.markdown
        "tolerance": tolerance_text,
        "hypothesis": hypothesis_text,
        "recommendations": [f"<li>{r}</li>" for r in recommendations],
    }


# --- Export Logic ---
# Ported from 'exportManager'
class ExportManager:
    def __init__(self):
        self.styles = {
            "title": {
                "font": Font(bold=True, sz=16, color="1F2937"),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "subtitle": {
                "font": Font(sz=10, color="6B7280"),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "header": {
                "font": Font(bold=True, color="FFFFFFFF"),
                "fill": PatternFill(
                    start_color="4B5563", end_color="4B5563", fill_type="solid"
                ),
                "alignment": Alignment(
                    horizontal="center", vertical="center", wrapText=True
                ),
                "border": Border(
                    bottom=Side(style="thin"),
                    top=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                ),
            },
            "subheader": {
                "font": Font(bold=True, color="111827", sz=12),
                "fill": PatternFill(
                    start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"
                ),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": Border(bottom=Side(style="medium", color="007BC5")),
            },
            "metricLabel": {
                "font": Font(bold=True),
                "alignment": Alignment(horizontal="right", vertical="center"),
            },
            "good": {
                "fill": PatternFill(
                    start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"
                ),
                "font": Font(color="065F46", bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "marginal": {
                "fill": PatternFill(
                    start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"
                ),
                "font": Font(color="92400E", bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "bad": {
                "fill": PatternFill(
                    start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
                ),
                "font": Font(color="991B1B", bold=True),
                "alignment": Alignment(horizontal="center", vertical="center"),
            },
            "dataCell": {
                "border": Border(
                    bottom=Side(style="dotted", color="D1D5DB"),
                    top=Side(style="dotted", color="D1D5DB"),
                    left=Side(style="dotted", color="D1D5DB"),
                    right=Side(style="dotted", color="D1D5DB"),
                ),
                "alignment": Alignment(vertical="center"),
            },
            "wrap": {"alignment": Alignment(wrapText=True, vertical="top")},
            "infinity": {
                "font": Font(sz=14),
                "alignment": Alignment(horizontal="right", vertical="center"),
            },
        }
        self.number_formats = {
            "integer": "0",
            "ppm": "#,##0",
            "scientific": "0.00E+00",
            "dateTime": "yyyy-mm-dd hh:mm:ss",
        }

    def _get_num_style(self, dp=3):
        dp = int(dp) if dp is not None else 3
        return {
            "number_format": f"0.{'0' * dp}",
            "alignment": Alignment(horizontal="right", vertical="center"),
        }

    def _get_perc_style(self, dp=3):
        dp = dp if dp is not None and isinstance(dp, int) else 3
        return {
            "number_format": f"0.{'0' * dp}%",
            "alignment": Alignment(horizontal="right", vertical="center"),
        }

    def _apply_styles(self, ws, data_with_styles):
        max_col_width = {}
        for r_idx, row in enumerate(data_with_styles, 1):
            for c_idx, cell_data in enumerate(row, 1):
                if not cell_data:
                    continue

                cell = ws.cell(row=r_idx, column=c_idx, value=cell_data["v"])
                cell.style = "Normal"  # Reset style

                style_dict = {**self.styles.get("dataCell", {})}

                if "s" in cell_data:
                    style_dict.update(cell_data["s"])

                if "font" in style_dict:
                    cell.font = style_dict["font"]
                if "fill" in style_dict:
                    cell.fill = style_dict["fill"]
                if "alignment" in style_dict:
                    cell.alignment = style_dict["alignment"]
                if "border" in style_dict:
                    cell.border = style_dict["border"]
                if "number_format" in style_dict:
                    cell.number_format = style_dict["number_format"]

                # Auto-fit columns
                cell_len = len(str(cell.value))
                if c_idx not in max_col_width or cell_len > max_col_width[c_idx]:
                    max_col_width[c_idx] = cell_len

        for c_idx, width in max_col_width.items():
            ws.column_dimensions[get_column_letter(c_idx)].width = min(
                max(width + 2, 10), 60
            )

    def _create_cell(self, value, style_keys=None, extra_styles=None):
        if style_keys is None:
            style_keys = []
        if extra_styles is None:
            extra_styles = {}

        final_style = {}
        for key in style_keys:
            if key in self.styles:
                final_style.update(self.styles[key])
        final_style.update(extra_styles)

        display_value = value
        if isinstance(value, (int, float)):
            if not np.isfinite(value):
                display_value = "∞" if value > 0 else "-∞"
                final_style.update(self.styles.get("infinity", {}))
        elif value is None:
            display_value = ""

        return {"v": display_value, "s": final_style}

    def export_current_results(self, results, summary):
        dp = results.get("dp", 3)
        dp = int(dp) if dp is not None else 3  # Ensure dp is an integer
        num_style = self._get_num_style(dp)
        num_style_s = self._get_num_style(dp + 2)
        num_style_pval = {
            "number_format": self.number_formats["scientific"],
            "alignment": self._get_num_style(dp)["alignment"],
        }
        ppm_style = {
            "number_format": self.number_formats["ppm"],
            "alignment": self._get_num_style(dp)["alignment"],
        }
        int_style = {
            "number_format": self.number_formats["integer"],
            "alignment": self._get_num_style(dp)["alignment"],
        }
        perc_style = self._get_perc_style(3)

        verdict_style_key = (
            "bad"
            if summary["verdict_color"] == "red"
            else (
                "marginal"
                if summary["verdict_color"] == "orange"
                else ("good" if summary["verdict_color"] == "green" else "dataCell")
            )
        )
        _verdict_style = self.styles.get(verdict_style_key, {})  # noqa: F841

        cpk_meets_target = (
            np.isfinite(results.get("CpkCurrent", np.nan))
            and np.isfinite(results.get("target_index_value", np.nan))
            and results["CpkCurrent"] >= results["target_index_value"]
        )
        cpk_style = (
            self.styles["good"]
            if results.get("s") == 0 and results.get("CpkCurrent", 0) > 0
            else (self.styles["good"] if cpk_meets_target else self.styles["bad"])
        )

        shift_significant = (
            np.isfinite(results.get("shiftValue", np.nan))
            and results.get("s", 0) > 0
            and abs(results["shiftValue"]) >= (results["s"] * 0.05)
        )
        shift_style = {
            **num_style,
            **(self.styles["marginal"] if shift_significant else {}),
        }

        ppm_below_style = {
            **ppm_style,
            **(self.styles["bad"] if results.get("ppm_below", 0) > 0 else {}),
        }
        ppm_above_style = {
            **ppm_style,
            **(self.styles["bad"] if results.get("ppm_above", 0) > 0 else {}),
        }

        hypo = results.get("hypothesisResult", {})
        hypo_conclusion = "N/A"
        hypo_style = self.styles["wrap"]
        if np.isfinite(hypo.get("p_value", np.nan)) and np.isfinite(
            hypo.get("alpha", np.nan)
        ):
            hypo_conclusion = (
                "Reject Null Hypothesis (Significant Shift)"
                if hypo["p_value"] < hypo["alpha"]
                else "Fail to Reject Null Hypothesis (No Significant Shift)"
            )
        elif results.get("s") == 0:
            hypo_conclusion = (
                "Reject Null Hypothesis (Significant Shift)"
                if results.get("shiftValue") != 0
                else "Fail to Reject Null Hypothesis (No Shift)"
            )

        if "Reject" in hypo_conclusion:
            hypo_style.update(self.styles["marginal"])
        elif "Fail" in hypo_conclusion:
            hypo_style.update(self.styles["good"])

        data = [
            [
                self._create_cell(
                    f"Capability Analysis Report: {results.get('measurement_name', 'Unnamed')}",
                    ["title"],
                ),
                None,
                None,
            ],
            [
                self._create_cell(
                    f"Analysis Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                    ["subtitle"],
                ),
                None,
                None,
            ],
            [],
            [
                self._create_cell("OVERALL ASSESSMENT", ["subheader"]),
                self._create_cell(summary["verdict"], [verdict_style_key]),
                None,
            ],
            [],
            [self._create_cell("INPUT PARAMETERS", ["subheader"]), None, None],
            [
                self._create_cell("Metric", ["header"]),
                self._create_cell("Value", ["header"]),
                self._create_cell("Notes", ["header"]),
            ],
            [
                self._create_cell("Measurement Name", ["metricLabel"]),
                self._create_cell(results.get("measurement_name")),
                None,
            ],
            [
                self._create_cell("Target Mean (Tm)", ["metricLabel"]),
                self._create_cell(results.get("tm"), extra_styles=num_style),
                None,
            ],
            [
                self._create_cell("LSL", ["metricLabel"]),
                self._create_cell(results.get("lsl"), extra_styles=num_style),
                None,
            ],
            [
                self._create_cell("USL", ["metricLabel"]),
                self._create_cell(results.get("usl"), extra_styles=num_style),
                None,
            ],
        ]

        if results.get("mode") == "manual":
            data.extend(
                [
                    [
                        self._create_cell("Measured Mean (x̄)", ["metricLabel"]),
                        self._create_cell(results.get("x_bar"), extra_styles=num_style),
                        None,
                    ],
                    [
                        self._create_cell("Std Deviation (s)", ["metricLabel"]),
                        self._create_cell(results.get("s"), extra_styles=num_style_s),
                        None,
                    ],
                    [
                        self._create_cell("Sample Size (n)", ["metricLabel"]),
                        self._create_cell(
                            results.get("n_samples"), extra_styles=int_style
                        ),
                        None,
                    ],
                ]
            )
        else:
            data.extend(
                [
                    [
                        self._create_cell("Data Source", ["metricLabel"]),
                        self._create_cell("Imported Data"),
                        None,
                    ],
                    [
                        self._create_cell("(Calculated) Mean (x̄)", ["metricLabel"]),
                        self._create_cell(results.get("x_bar"), extra_styles=num_style),
                        self._create_cell("From imported data", ["wrap"]),
                    ],
                    [
                        self._create_cell("(Calculated) Std Dev (s)", ["metricLabel"]),
                        self._create_cell(results.get("s"), extra_styles=num_style_s),
                        self._create_cell("From imported data", ["wrap"]),
                    ],
                    [
                        self._create_cell(
                            "(Calculated) Sample Size (n)", ["metricLabel"]
                        ),
                        self._create_cell(
                            results.get("n_samples"), extra_styles=int_style
                        ),
                        self._create_cell("From imported data", ["wrap"]),
                    ],
                ]
            )

        data.extend(
            [
                [
                    self._create_cell("Target Index Type", ["metricLabel"]),
                    self._create_cell(results.get("target_index_type")),
                    None,
                ],
                [
                    self._create_cell("Target Index Value", ["metricLabel"]),
                    self._create_cell(
                        results.get("target_index_value"),
                        extra_styles=self._get_num_style(2),
                    ),
                    None,
                ],
                [
                    self._create_cell("Confidence Level (%)", ["metricLabel"]),
                    self._create_cell(
                        results.get("confidence_level"), extra_styles=int_style
                    ),
                    None,
                ],
                [
                    self._create_cell("Distribution", ["metricLabel"]),
                    self._create_cell(results.get("distribution")),
                    None,
                ],
                [
                    self._create_cell("Hypothesis Type", ["metricLabel"]),
                    self._create_cell(results.get("hypothesis_type")),
                    None,
                ],
                [],
                [self._create_cell("CALCULATED RESULTS", ["subheader"]), None, None],
                [
                    self._create_cell("Metric", ["header"]),
                    self._create_cell("Value", ["header"]),
                    self._create_cell("Notes", ["header"]),
                ],
                [
                    self._create_cell("Shift Required (Tm - x̄)", ["metricLabel"]),
                    self._create_cell(
                        results.get("shiftValue"), extra_styles=shift_style
                    ),
                    None,
                ],
                [
                    self._create_cell("Drawing Tolerance (USL - LSL)", ["metricLabel"]),
                    self._create_cell(results.get("T_drawing"), extra_styles=num_style),
                    None,
                ],
                [
                    self._create_cell("6σ Spread (6 * s)", ["metricLabel"]),
                    self._create_cell(
                        results.get("sixSigmaSpread"), extra_styles=num_style
                    ),
                    None,
                ],
                [
                    self._create_cell("8σ Spread (8 * s)", ["metricLabel"]),
                    self._create_cell(
                        results.get("eightSigmaSpread"), extra_styles=num_style
                    ),
                    None,
                ],
                [
                    self._create_cell("Cp (Potential Capability)", ["metricLabel"]),
                    self._create_cell(results.get("Cp"), extra_styles=num_style),
                    None,
                ],
                [
                    self._create_cell(
                        f"{results.get('target_index_type')} (Actual Capability)",
                        ["metricLabel"],
                    ),
                    self._create_cell(
                        results.get("CpkCurrent"),
                        extra_styles={**num_style, **cpk_style},
                    ),
                    self._create_cell(
                        f"Target: {results.get('target_index_value', 0):.2f}", ["wrap"]
                    ),
                ],
                [
                    self._create_cell("Required Tolerance", ["metricLabel"]),
                    self._create_cell(
                        results.get("newToleranceTotal"), extra_styles=num_style
                    ),
                    self._create_cell(
                        f"For target {results.get('target_index_type')} of {results.get('target_index_value', 0):.2f}",
                        ["wrap"],
                    ),
                ],
                [
                    self._create_cell("Confidence Interval Lower", ["metricLabel"]),
                    self._create_cell(results.get("ci_lower"), extra_styles=num_style),
                    None,
                ],
                [
                    self._create_cell("Confidence Interval Upper", ["metricLabel"]),
                    self._create_cell(results.get("ci_upper"), extra_styles=num_style),
                    None,
                ],
                [],
                [self._create_cell("PROBABILITY & DEFECTS", ["subheader"]), None, None],
                [
                    self._create_cell("Metric", ["header"]),
                    self._create_cell("Value", ["header"]),
                    None,
                ],
                [
                    self._create_cell("Probability < LSL (%)", ["metricLabel"]),
                    self._create_cell(
                        results.get("prob_below"), extra_styles=perc_style
                    ),
                    None,
                ],
                [
                    self._create_cell("Probability > USL (%)", ["metricLabel"]),
                    self._create_cell(
                        results.get("prob_above"), extra_styles=perc_style
                    ),
                    None,
                ],
                [
                    self._create_cell("PPM < LSL", ["metricLabel"]),
                    self._create_cell(
                        results.get("ppm_below"), extra_styles=ppm_below_style
                    ),
                    None,
                ],
                [
                    self._create_cell("PPM > USL", ["metricLabel"]),
                    self._create_cell(
                        results.get("ppm_above"), extra_styles=ppm_above_style
                    ),
                    None,
                ],
                [],
                [
                    self._create_cell(
                        "HYPOTHESIS TEST (Mean vs Target)", ["subheader"]
                    ),
                    None,
                    None,
                ],
                [
                    self._create_cell("Metric", ["header"]),
                    self._create_cell("Value", ["header"]),
                    None,
                ],
                [
                    self._create_cell("Z-Statistic", ["metricLabel"]),
                    self._create_cell(
                        hypo.get("z_stat"), extra_styles=self._get_num_style(4)
                    ),
                    None,
                ],
                [
                    self._create_cell("P-Value", ["metricLabel"]),
                    self._create_cell(hypo.get("p_value"), extra_styles=num_style_pval),
                    None,
                ],
                [
                    self._create_cell("Alpha", ["metricLabel"]),
                    self._create_cell(
                        hypo.get("alpha"), extra_styles=self._get_num_style(2)
                    ),
                    None,
                ],
                [
                    self._create_cell("Conclusion", ["metricLabel"]),
                    self._create_cell(hypo_conclusion, extra_styles=hypo_style),
                    None,
                ],
            ]
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Capability Report"

        self._apply_styles(ws, data)

        # Apply merges
        ws.merge_cells("A1:C1")
        ws.merge_cells("A2:C2")
        ws.merge_cells("B4:C4")
        ws.merge_cells("A6:C6")
        ws.merge_cells("A24:C24")
        ws.merge_cells("A35:C35")
        ws.merge_cells("A41:C41")

        # Apply row heights
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 6
        ws.row_dimensions[4].height = 20

        # Save to memory buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def export_selected_history(self, history_data):
        headers = [
            "Timestamp",
            "Characteristic",
            "Measurement_Name",
            "Verdict",
            "Cp",
            "Cpk/Ppk",
            "Target_Index_Type",
            "Target_Index_Value",
            "Required_Shift",
            "Target_Mean_Tm",
            "LSL",
            "USL",
            "Measured_Mean_xbar",
            "Std_Dev_s",
            "Sample_Size_n",
            "PPM_Below_LSL",
            "PPM_Above_USL",
            "Distribution",
            "Confidence_Level",
            "Hypothesis_Type",
            "Z_Stat",
            "P_Value",
            "Alpha",
            "Hypo_Conclusion",
        ]

        data = [[self._create_cell(h, ["header"]) for h in headers]]

        for entry in history_data:
            dp = entry.get("dp", 3)
            dp = int(dp) if dp is not None else 3  # Ensure dp is an integer
            verdict = entry.get("verdict", "N/A")
            verdict_style_key = (
                "bad"
                if "ACTION" in verdict or "INVALID" in verdict
                else (
                    "marginal"
                    if "MARGINAL" in verdict
                    else ("good" if "GOOD" in verdict else "dataCell")
                )
            )

            num_style = self._get_num_style(dp)
            num_style_more = self._get_num_style(dp + 1)
            ppm_style = {
                "number_format": self.number_formats["ppm"],
                "alignment": self._get_num_style(dp)["alignment"],
            }
            int_style = {
                "number_format": self.number_formats["integer"],
                "alignment": self._get_num_style(dp)["alignment"],
            }
            num_style_pval = {
                "number_format": self.number_formats["scientific"],
                "alignment": self._get_num_style(dp)["alignment"],
            }

            cpk_meets = (
                np.isfinite(entry.get("CpkCurrent", np.nan))
                and np.isfinite(entry.get("target_index_value", np.nan))
                and entry["CpkCurrent"] >= entry["target_index_value"]
            )
            cpk_style = (
                self.styles["good"]
                if entry.get("s") == 0 and entry.get("CpkCurrent", 0) > 0
                else (self.styles["good"] if cpk_meets else self.styles["bad"])
            )

            shift_sig = (
                np.isfinite(entry.get("shiftValue", np.nan))
                and entry.get("s", 0) > 0
                and abs(entry["shiftValue"]) >= (entry["s"] * 0.05)
            )
            shift_style = {
                **num_style,
                **(self.styles["marginal"] if shift_sig else {}),
            }

            ppm_below_style = {
                **ppm_style,
                **(self.styles["bad"] if entry.get("ppm_below", 0) > 0 else {}),
            }
            ppm_above_style = {
                **ppm_style,
                **(self.styles["bad"] if entry.get("ppm_above", 0) > 0 else {}),
            }

            hypo = entry.get("hypothesisResult", {})
            hypo_conclusion = ""
            hypo_style_key = "dataCell"
            if np.isfinite(hypo.get("p_value", np.nan)) and np.isfinite(
                hypo.get("alpha", np.nan)
            ):
                hypo_conclusion = (
                    "Reject H0"
                    if hypo["p_value"] < hypo["alpha"]
                    else "Fail to Reject H0"
                )
            elif entry.get("s") == 0:
                hypo_conclusion = (
                    "Reject H0" if entry.get("shiftValue") != 0 else "Fail to Reject H0"
                )

            if "Reject" in hypo_conclusion:
                hypo_style_key = "marginal"
            elif "Fail" in hypo_conclusion:
                hypo_style_key = "good"

            row = [
                self._create_cell(
                    datetime.datetime.fromisoformat(entry.get("id"))
                    if entry.get("id")
                    else None,
                    extra_styles={"number_format": self.number_formats["dateTime"]},
                ),
                self._create_cell(
                    entry.get("characteristic_name", entry.get("measurement_name", ""))
                ),
                self._create_cell(entry.get("measurement_name", "")),
                self._create_cell(verdict, [verdict_style_key]),
                self._create_cell(entry.get("Cp"), extra_styles=num_style),
                self._create_cell(
                    entry.get("CpkCurrent"), extra_styles={**num_style, **cpk_style}
                ),
                self._create_cell(entry.get("target_index_type", "")),
                self._create_cell(
                    entry.get("target_index_value"), extra_styles=self._get_num_style(2)
                ),
                self._create_cell(entry.get("shiftValue"), extra_styles=shift_style),
                self._create_cell(entry.get("tm"), extra_styles=num_style),
                self._create_cell(entry.get("lsl"), extra_styles=num_style),
                self._create_cell(entry.get("usl"), extra_styles=num_style),
                self._create_cell(entry.get("x_bar"), extra_styles=num_style),
                self._create_cell(entry.get("s"), extra_styles=num_style_more),
                self._create_cell(entry.get("n_samples"), extra_styles=int_style),
                self._create_cell(entry.get("ppm_below"), extra_styles=ppm_below_style),
                self._create_cell(entry.get("ppm_above"), extra_styles=ppm_above_style),
                self._create_cell(entry.get("distribution", "")),
                self._create_cell(
                    entry.get("confidence_level"), extra_styles=int_style
                ),
                self._create_cell(entry.get("hypothesis_type", "")),
                self._create_cell(
                    hypo.get("z_stat"), extra_styles=self._get_num_style(4)
                ),
                self._create_cell(hypo.get("p_value"), extra_styles=num_style_pval),
                self._create_cell(
                    hypo.get("alpha"), extra_styles=self._get_num_style(2)
                ),
                self._create_cell(hypo_conclusion, [hypo_style_key]),
            ]
            data.append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = "History Selection"

        self._apply_styles(ws, data)
        ws.row_dimensions[1].height = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# --- Sigma Assistant Mascot (Clippy-Style) ---
class SigmaAssistant:
    """Generates a Clippy-style floating Sigma Assistant with CSS animations."""

    # Messages for each state
    STATE_MESSAGES = {
        "idle": "Hello! I'm here to help. Run an analysis to see results!",
        "happy": "Excellent! Process is stable and capable. Great work!",
        "sad": "Action required. The process needs improvement.",
        "thinking": "Analyzing the data... Please wait!",
        "concerned": "The process is marginal. Review recommendations.",
    }

    # State colors (border color for the body)
    STATE_COLORS = {
        "idle": "#3B82F6",  # Blue
        "happy": "#10B981",  # Green
        "sad": "#EF4444",  # Red
        "thinking": "#FBBF24",  # Yellow
        "concerned": "#F97316",  # Orange
    }

    @classmethod
    def render_fixed(cls, state="idle", message=None):
        """
        Render the Clippy-style mascot using st.markdown for TRUE fixed positioning.
        This injects CSS/HTML directly into Streamlit's main page, not an iframe.
        """
        if state not in cls.STATE_MESSAGES:
            state = "idle"

        msg = message if message else cls.STATE_MESSAGES.get(state)
        color = cls.STATE_COLORS.get(state, "#3B82F6")

        # Animation name based on state
        animation_map = {
            "idle": "sigma-bob",
            "happy": "sigma-happy-dance",
            "sad": "sigma-sad-slump",
            "thinking": "sigma-thinking",
            "concerned": "sigma-bob",
        }
        animation = animation_map.get(state, "sigma-bob")

        # Mouth path based on state
        mouth_map = {
            "idle": "M 55 90 Q 70 95 85 90",
            "happy": "M 55 85 C 60 105, 80 105, 85 85",
            "sad": "M 55 95 Q 70 85 85 95",
            "thinking": "M 60 90 L 80 90",
            "concerned": "M 55 93 Q 70 88 85 93",
        }
        mouth = mouth_map.get(state, mouth_map["idle"])

        # Eyebrow transforms based on state
        eyebrow_left = "translate(50, 45)"
        eyebrow_right = "translate(90, 45)"
        if state == "sad":
            eyebrow_left = "translate(45, 43) rotate(15)"
            eyebrow_right = "translate(95, 43) rotate(-15)"
        elif state == "concerned":
            eyebrow_left = "translate(50, 45) rotate(10)"
            eyebrow_right = "translate(90, 45) rotate(-10)"
        elif state == "thinking":
            eyebrow_left = "translate(50, 45) rotate(-10)"
            eyebrow_right = "translate(90, 45) rotate(10)"

        html = f'''
<style>
/* Sigma Assistant Fixed Positioning - injected into Streamlit main page */
.sigma-fixed-container {{
    position: fixed !important;
    bottom: 80px !important;
    right: 20px !important;
    z-index: 999999 !important;
    display: flex;
    flex-direction: column;
    align-items: center;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    pointer-events: none;
}}

.sigma-speech-bubble {{
    background-color: #1F2937;
    color: white;
    padding: 10px 14px;
    border-radius: 10px;
    margin-bottom: 8px;
    max-width: 200px;
    font-size: 0.8rem;
    box-shadow: 0 4px 12px rgba(0,0,0,0.25);
    text-align: center;
    line-height: 1.4;
}}

.sigma-mascot {{
    width: 100px;
    height: 100px;
    cursor: pointer;
    pointer-events: auto;
}}

.sigma-mascot svg {{
    filter: drop-shadow(2px 3px 3px rgba(0, 0, 0, 0.2));
}}

/* Animations */
@keyframes sigma-bob {{
    0%, 100% {{ transform: translateY(0); }}
    50% {{ transform: translateY(-3px); }}
}}

@keyframes sigma-happy-dance {{
    0%, 100% {{ transform: translateY(0) rotate(0); }}
    15% {{ transform: translateY(-5px) rotate(3deg); }}
    30% {{ transform: translateY(0) rotate(0); }}
    45% {{ transform: translateY(-5px) rotate(-3deg); }}
    60% {{ transform: translateY(0) rotate(0); }}
}}

@keyframes sigma-sad-slump {{
    0%, 100% {{ transform: translateY(0) rotate(0); }}
    50% {{ transform: translateY(3px) rotate(-1deg) scaleY(0.96); }}
}}

@keyframes sigma-thinking {{
    0%, 100% {{ transform: rotate(0deg); }}
    25% {{ transform: rotate(1deg); }}
    75% {{ transform: rotate(-1deg); }}
}}

.sigma-animate {{
    animation: {animation} 2s ease-in-out infinite;
}}
</style>

<div class="sigma-fixed-container">
    <div class="sigma-speech-bubble">{msg}</div>
    <div class="sigma-mascot sigma-animate">
        <svg viewBox="-20 -30 150 150" xmlns="http://www.w3.org/2000/svg">
            <defs>
                <filter id="sigma-shadow"><feGaussianBlur in="SourceAlpha" stdDeviation="2"/></filter>
            </defs>
            <!-- Shadow -->
            <ellipse cx="55" cy="115" rx="35" ry="8" fill="black" opacity="0.15" filter="url(#sigma-shadow)"/>
            <!-- Body -->
            <path d="M 20 110 C 20 110, 5 20, 55 20 C 105 20, 90 110, 90 110 Z" 
                  fill="#F9FAFB" stroke="{color}" stroke-width="4" stroke-linejoin="round"/>
            <!-- Face -->
            <g transform="translate(0, -10)">
                <!-- Eyebrows -->
                <g transform="{eyebrow_left}">
                    <path d="M -8 0 Q 0 -5 8 0" fill="none" stroke="#4B5563" stroke-width="3" stroke-linecap="round"/>
                </g>
                <g transform="{eyebrow_right}">
                    <path d="M -8 0 Q 0 -5 8 0" fill="none" stroke="#4B5563" stroke-width="3" stroke-linecap="round"/>
                </g>
                <!-- Eyes -->
                <g transform="translate(50, 60)">
                    <ellipse rx="10" ry="8" fill="white" stroke="#1F2937" stroke-width="1.5"/>
                    <circle cx="1" cy="1" r="4" fill="#1F2937"/>
                    <circle cx="3" cy="-1" r="1.5" fill="white"/>
                </g>
                <g transform="translate(90, 60)">
                    <ellipse rx="10" ry="8" fill="white" stroke="#1F2937" stroke-width="1.5"/>
                    <circle cx="1" cy="1" r="4" fill="#1F2937"/>
                    <circle cx="3" cy="-1" r="1.5" fill="white"/>
                </g>
                <!-- Mouth -->
                <path d="{mouth}" fill="none" stroke="#1E3A8A" stroke-width="2.5" stroke-linecap="round"/>
            </g>
        </svg>
    </div>
</div>
'''
        return html


# --- Chatbot Logic ---
# Ported from 'sigmaAssistant'
class Chatbot:
    def __init__(self):
        # Prepare reference content
        self.reference_content_sections = self._prepare_reference_content()
        self.common_words = set(
            [
                "a",
                "an",
                "the",
                "is",
                "are",
                "what",
                "how",
                "when",
                "where",
                "for",
                "to",
                "of",
                "in",
                "and",
                "or",
                "do",
                "does",
                "can",
                "explain",
                "tell",
                "me",
                "about",
            ]
        )

    def _prepare_reference_content(self):
        raw_sections = [
            # --- Application Context ---
            {"context": "Application Context", "text": "This tool is used in Six Sigma and SPC environments for Process Centering and Tolerance Verification. It helps engineers quantify process drift, predict initial settings, and verify tolerance adequacy."},
            {"context": "Application Context", "text": "Quantify Process Drift: Calculate the exact Required Shift (Δ) = Tₘ − x̄ to move the measured process mean back to the engineering target."},
            {"context": "Application Context", "text": "Predict Initial State: Infer the required initial dimension/setting to achieve the target (Tₘ) after process variables exert their influence."},
            {"context": "Application Context", "text": "Verify Tolerance Adequacy: Determine the minimum Required Tolerance (USL − LSL) for the existing process variation (σ) to meet a desired Cpk/Ppk."},

            # --- Step-by-Step Workflow ---
            {"context": "How to Use This Tool", "text": "Step 1: Go to 'Data Worksheet' tab. Enter measurement label, description, Tₘ (Target Mean), LSL (Lower Spec Limit), and USL (Upper Spec Limit)."},
            {"context": "How to Use This Tool", "text": "Step 2: Enter data — type DMC/serial + measured values directly in the worksheet grid, or upload a CSV/Excel file, or click 'Sample Data' for 1000 test points."},
            {"context": "How to Use This Tool", "text": "Step 3: Go to 'Analysis & Report' tab. Select 'Use Data Worksheet' mode, then click 'ANALYZE & PLOT' to run the statistical analysis."},
            {"context": "How to Use This Tool", "text": "Step 4: Review results in the 'Visualization' tab — histogram, box plot, capability curves, I-Chart, MR-Chart, and statistics summary."},
            {"context": "How to Use This Tool", "text": "Step 5: Check the 'History' tab to compare with previous runs. Export results as Excel for reports."},
            {"context": "How to Use This Tool", "text": "You can download a pre-formatted Excel template using the '📥 Download Template' button in the Data Worksheet tab. Fill it in Excel and upload it back to avoid formatting issues."},
            {"context": "How to Use This Tool", "text": "Data Worksheet supports up to 1000 data rows. The capacity indicator shows how many rows are used."},

            # --- Capability Indices ---
            {"context": "Capability Indices (Cp/Cpk)", "text": "Cp = (USL − LSL) / 6σ measures POTENTIAL capability — how well the process COULD perform if perfectly centered. A higher Cp means the tolerance band is wider relative to process spread."},
            {"context": "Capability Indices (Cp/Cpk)", "text": "Cpk = min[(USL − x̄) / 3σ, (x̄ − LSL) / 3σ] measures ACTUAL capability with real centering included. Cpk ≤ Cp always. If Cpk = Cp, the process is perfectly centered."},
            {"context": "Capability Indices (Cp/Cpk)", "text": "Cpk interpretation: Cpk < 1.0 = NOT capable (process spread exceeds tolerance). Cpk 1.0–1.33 = Marginally capable. Cpk 1.33–1.67 = Capable. Cpk ≥ 1.67 = Highly capable. Cpk ≥ 2.0 = Six Sigma level."},
            {"context": "Capability Indices (Cp/Cpk)", "text": "Automotive industry (IATF 16949) typically requires Cpk ≥ 1.33 for ongoing production and Cpk ≥ 1.67 for new processes or critical characteristics."},
            {"context": "Capability Indices (Cp/Cpk)", "text": "If Cp is high but Cpk is low, the process is capable but not centered — you need to shift the mean toward the target. The Required Shift (Δ) tells you exactly how much."},
            {"context": "Capability Indices (Cp/Cpk)", "text": "Pp = (USL − LSL) / 6σ_overall and Ppk = min[(USL − x̄) / 3σ_overall, (x̄ − LSL) / 3σ_overall] use OVERALL (long-term) standard deviation including between-subgroup variation. Ppk ≤ Cpk always."},

            # --- Control Charts ---
            {"context": "Control Charts (I-MR)", "text": "The I-Chart (Individuals Chart) plots each measured value in sequence. The center line (CL) is x̄, UCL = x̄ + 3σ, LCL = x̄ − 3σ. Points outside UCL/LCL indicate the process is out of statistical control."},
            {"context": "Control Charts (I-MR)", "text": "The MR-Chart (Moving Range) plots |Xᵢ − Xᵢ₋₁| — the absolute difference between consecutive measurements. MR̄ is the average moving range, MR UCL = 3.267 × MR̄ (D4 constant for n=2). No LCL for MR chart."},
            {"context": "Control Charts (I-MR)", "text": "Use the 'Show Points' filter (10/25/50/100/250/500/All) to control how many data points are displayed on the control charts. This helps focus on recent data or see the full picture."},
            {"context": "Control Charts (I-MR)", "text": "Zone A is between ±2σ and ±3σ (~4.28% of data). Zone B is between ±1σ and ±2σ (~27.18% of data). Zone C is within ±1σ (~68.27% of data). The percentages assume a normal distribution."},
            {"context": "Control Charts (I-MR)", "text": "Warning limits (±2σ) are shown as amber dotted lines. These are optional — toggle them with the 'Show Warning Limits (±2σ)' checkbox. Points between 2σ and 3σ are early warnings of potential issues."},
            {"context": "Control Charts (I-MR)", "text": "±1σ zone lines (purple) are available in the legend — click their name in the legend to toggle visibility on the I-Chart."},
            {"context": "Control Charts (I-MR)", "text": "LSL, USL, and Target Mean are displayed on the I-Chart as green dash-dot and purple long-dash lines respectively, if specification limits are set."},

            # --- Western Electric Rules ---
            {"context": "Western Electric Rules", "text": "Rule 1: Any single point beyond ±3σ (UCL or LCL) = Out of Control. This is the most basic rule and is always checked."},
            {"context": "Western Electric Rules", "text": "Rule 2: Two out of three consecutive points beyond ±2σ on the same side = Warning signal of a process shift."},
            {"context": "Western Electric Rules", "text": "Rule 3: Four out of five consecutive points beyond ±1σ on the same side = Developing trend."},
            {"context": "Western Electric Rules", "text": "Rule 4: Eight or more consecutive points on the same side of the center line = Process shift (mean has moved)."},
            {"context": "Western Electric Rules", "text": "Rule 5: Six consecutive points steadily increasing or decreasing = Trend (tool wear, drift, temperature change)."},

            # --- Standard Deviation ---
            {"context": "Standard Deviation (σ)", "text": "σ (sigma) measures process variation — the spread of measured values around the mean. Calculated as STDEV.S (sample standard deviation with n−1 degrees of freedom)."},
            {"context": "Standard Deviation (σ)", "text": "6σ Spread contains 99.73% of all data in a normal distribution. 8σ Spread (±4σ) contains 99.9937%. These define the natural process limits."},
            {"context": "Standard Deviation (σ)", "text": "To reduce σ, investigate and control sources of variation: machine vibration, tool wear, material inconsistency, temperature fluctuations, operator technique."},
            {"context": "Standard Deviation (σ)", "text": "The relationship between σ and PPM: at ±3σ = 2,700 PPM defective. At ±4σ = 63 PPM. At ±5σ = 0.57 PPM. At ±6σ = 0.002 PPM (3.4 with 1.5σ shift)."},

            # --- PPM & Sigma Level ---
            {"context": "PPM & Sigma Level", "text": "PPM (Parts Per Million) is the defect rate. PPM = P(defect) × 1,000,000. For example, 0.1% defect rate = 1,000 PPM."},
            {"context": "PPM & Sigma Level", "text": "Sigma level table: 1σ = 691,462 PPM. 2σ = 308,538 PPM. 3σ = 66,807 PPM. 4σ = 6,210 PPM. 5σ = 233 PPM. 6σ = 3.4 PPM (with 1.5σ shift)."},
            {"context": "PPM & Sigma Level", "text": "PPM below LSL means parts that are too small/thin/weak. PPM above USL means parts that are too large/thick. Both come from the tail probabilities of the normal distribution."},

            # --- Hypothesis Testing ---
            {"context": "Hypothesis Testing", "text": "Z-test determines if the process mean (μ) is statistically different from Target Mean (Tₘ). Z = (x̄ − Tₘ) / (σ / √n). Large |Z| = significant shift."},
            {"context": "Hypothesis Testing", "text": "Null Hypothesis (H₀): μ = Tₘ (process is on target). Alternative Hypothesis (H₁): μ ≠ Tₘ (two-sided), μ > Tₘ (upper), or μ < Tₘ (lower)."},
            {"context": "Hypothesis Testing", "text": "p-value < 0.05 → Reject H₀ (significant evidence the mean has shifted). p-value ≥ 0.05 → Fail to Reject H₀ (no significant evidence of shift)."},
            {"context": "Hypothesis Testing", "text": "Two-Sided test: Use when deviations in BOTH directions are concerning (e.g., diameter must be exactly Tₘ). Upper-Sided: concern is mean ABOVE target. Lower-Sided: concern is mean BELOW target."},
            {"context": "Hypothesis Testing", "text": "Confidence Level = 1 − α. At 95% confidence (α = 0.05), if p < 0.05 you reject H₀. At 99% confidence (α = 0.01), you need p < 0.01 to reject."},

            # --- Distribution ---
            {"context": "Distribution", "text": "Normal distribution is the default assumption for most machining processes (turning, grinding, milling, stamping, injection molding) when the process is stable."},
            {"context": "Distribution", "text": "Lognormal distribution applies when values are always positive and right-skewed (e.g., surface roughness, concentricity, contamination levels). Log-transform the data and check normality."},
            {"context": "Distribution", "text": "A skewness near 0 and kurtosis near 0 (excess) indicate a distribution close to normal. High positive skewness suggests lognormal may fit better."},

            # --- Additional Metrics ---
            {"context": "Additional Metrics", "text": "x̄ ± 4σ Spread contains ~99.9937% of process output — a conservative view. x̄ ± 3σ Spread (99.73%) is the standard for control limits."},
            {"context": "Additional Metrics", "text": "P(x < LSL) and P(x > USL) are the tail probabilities from the normal CDF. They represent the fraction of parts expected to fall outside specifications."},
            {"context": "Additional Metrics", "text": "P(x < Tₘ) should be 50% for a perfectly centered process. Values far from 50% indicate the process mean is shifted from target."},
            {"context": "Additional Metrics", "text": "Required Shift (Δ) = Tₘ − x̄. Positive Δ means increase the mean (parts are too small). Negative Δ means decrease the mean (parts are too large)."},
            {"context": "Additional Metrics", "text": "Required Tolerance = Target Index × 6σ. This tells you the minimum tolerance band needed for the current process variation to meet your capability target."},

            # --- Troubleshooting ---
            {"context": "Troubleshooting", "text": "If Cpk is low but Cp is good: Your process spread is fine but the mean is off-center. Apply the Required Shift (Δ) to recenter."},
            {"context": "Troubleshooting", "text": "If both Cp and Cpk are low: Process variation is too high. Investigate sources of variation — machine condition, material, tooling, environment."},
            {"context": "Troubleshooting", "text": "If control chart shows trends or runs: Process is drifting. Check tool wear, temperature drift, material batch changes. Implement SPC monitoring."},
            {"context": "Troubleshooting", "text": "If many points are outside control limits: The process is unstable. Do NOT calculate capability indices until the process is in statistical control first."},
            {"context": "Troubleshooting", "text": "If data is not normal: Check skewness and kurtosis. Consider lognormal transformation. Remove outliers if justified. Increase sample size."},
            {"context": "Troubleshooting", "text": "If you see 'Need ≥2 data points': Enter at least 2 measurements in the Data Worksheet or in the manual entry fields."},
            {"context": "Troubleshooting", "text": "If the worksheet upload doesn't work: Ensure your file has at least one numeric column. The tool uses the first numeric column as measurement values."},

            # --- History & Export ---
            {"context": "History & Export", "text": "Every time you click 'ANALYZE & PLOT', the results are saved to the History tab with a timestamp. Up to 250 runs are kept."},
            {"context": "History & Export", "text": "Use the History tab filters to search by name, verdict, or characteristic. Select rows with checkboxes and click 'Export Selected' to download as Excel."},
            {"context": "History & Export", "text": "The History Cpk Trend chart shows how Cpk has changed over time. A declining trend indicates the process is degrading and needs attention."},
            {"context": "History & Export", "text": "You can download the full history as CSV using the 'Download History CSV' button for further analysis in Excel or other tools."},

            # --- Industry Standards ---
            {"context": "Industry Standards", "text": "IATF 16949 (Automotive): Cpk ≥ 1.33 for production, Cpk ≥ 1.67 for new processes/safety-critical features. Some OEMs require Cpk ≥ 2.0 for critical dimensions."},
            {"context": "Industry Standards", "text": "ISO 22514 defines capability study procedures. Short-term (Cm/Cmk) uses ≥50 parts from one setup. Long-term (Pp/Ppk) uses ≥100 parts across multiple setups/shifts."},
            {"context": "Industry Standards", "text": "Aerospace (AS9100): Process capability requirements vary by customer but typically Cpk ≥ 1.33. Some require Cpk ≥ 1.5 for critical features."},
            {"context": "Industry Standards", "text": "Medical devices (ISO 13485): Capability requirements depend on risk classification. Critical-to-quality features often require Cpk ≥ 1.33 minimum."},
        ]

        return [
            {"original": s["text"], "lower": s["text"].lower(), "context": s["context"]}
            for s in raw_sections
            if len(s["text"]) > 10
        ]

    def get_response(self, user_input):
        user_input_lower = user_input.lower()
        if not user_input_lower:
            return None

        # Simple keyword extraction
        keywords = [
            word
            for word in re.split(r"[\s,?\-.();:]+", user_input_lower)
            if word and len(word) > 2 and word not in self.common_words
        ]

        if not keywords:
            return "Please ask a more specific question using keywords like 'Cp', 'Cpk', 'control chart', 'histogram', 'PPM', 'hypothesis', 'troubleshooting', etc."

        best_match = None
        highest_score = 0

        for section in self.reference_content_sections:
            current_score = 0
            for keyword in keywords:
                if keyword in section["lower"]:
                    current_score += 1
                    if keyword in section["context"].lower():
                        current_score += 2  # Stronger context bonus

            if current_score > highest_score:
                highest_score = current_score
                best_match = section
            elif (
                current_score == highest_score
                and best_match
                and len(section["original"]) < len(best_match["original"])
            ):
                best_match = section

        if best_match:
            return f'**{best_match["context"]}**\n\n{best_match["original"]}'
        else:
            return ("Sorry, I couldn't find that specific topic. Try keywords like:\n"
                    "- **Capability:** Cp, Cpk, Pp, Ppk, capability, sigma\n"
                    "- **Charts:** control chart, I-chart, MR chart, zones, filter\n"
                    "- **Stats:** standard deviation, PPM, histogram, normal\n"
                    "- **Testing:** hypothesis, p-value, Z-test, confidence\n"
                    "- **Help:** troubleshooting, workflow, steps, industry\n"
                    "- **Data:** template, upload, download, worksheet")


# --- Main App ---

# Initialize calculators and managers
calc = StatisticalCalculator()
plotter = PlotManager()
exporter = ExportManager()
bot = Chatbot()


def coerce_valid_numeric_values(values):
    valid_values = []
    for value in values:
        if isinstance(value, (int, float, np.integer, np.floating)) and np.isfinite(
            value
        ):
            valid_values.append(float(value))
    return valid_values


CHARACTERISTIC_FIELDS = [
    "tm",
    "lsl",
    "usl",
    "target_index_value",
    "target_index_type",
    "confidence_level",
    "distribution",
    "hypothesis_type",
    "x_bar",
    "s",
    "n_samples",
    "decimal_places",
    "mode",
    "measurement_name",
    "description",
    "raw_data",
    "transform_dirty",
]


def default_characteristic_state(name="Characteristic 1"):
    worksheet_df = pd.DataFrame({"Value": [None] * 100})
    return {
        "tm": 10.00,
        "lsl": 9.90,
        "usl": 10.10,
        "target_index_value": 1.67,
        "target_index_type": "Cpk",
        "confidence_level": 95.0,
        "distribution": "Normal",
        "hypothesis_type": "Two-Sided",
        "x_bar": 10.00,
        "s": 0.015,
        "n_samples": 30,
        "decimal_places": 3,
        "mode": "Use Data Worksheet",
        "measurement_name": name,
        "description": "",
        "raw_data": "",
        "transform_dirty": False,
        "results": {},
        "summary": {},
        "figs": {},
        "worksheet_data": worksheet_df.copy(),
        "original_worksheet_data": worksheet_df.copy(),
    }


def sanitize_characteristic_name(name):
    cleaned = re.sub(r"\s+", " ", str(name or "").strip())
    return cleaned[:80] if cleaned else ""


def characteristic_from_flat_state(name):
    state = default_characteristic_state(name)
    for key in CHARACTERISTIC_FIELDS:
        if key in st.session_state:
            state[key] = st.session_state[key]
    state["measurement_name"] = sanitize_characteristic_name(
        st.session_state.get("measurement_name", name)
    ) or name
    state["results"] = dict(st.session_state.get("results", {}))
    state["summary"] = dict(st.session_state.get("summary", {}))
    state["figs"] = dict(st.session_state.get("figs", {}))
    worksheet = st.session_state.get("worksheet_data")
    if isinstance(worksheet, pd.DataFrame):
        state["worksheet_data"] = worksheet.copy()
    original = st.session_state.get("original_worksheet_data")
    if isinstance(original, pd.DataFrame):
        state["original_worksheet_data"] = original.copy()
    else:
        state["original_worksheet_data"] = state["worksheet_data"].copy()
    return state


def ensure_characteristics_state():
    if "characteristics" not in st.session_state or not st.session_state.characteristics:
        initial_name = sanitize_characteristic_name(
            st.session_state.get("measurement_name", "Characteristic 1")
        ) or "Characteristic 1"
        st.session_state.characteristics = {
            initial_name: characteristic_from_flat_state(initial_name)
        }
        st.session_state.active_characteristic_name = initial_name
        st.session_state.loaded_characteristic_name = None
        st.session_state.new_characteristic_name = ""
    else:
        if "active_characteristic_name" not in st.session_state:
            st.session_state.active_characteristic_name = next(
                iter(st.session_state.characteristics)
            )
        if "loaded_characteristic_name" not in st.session_state:
            st.session_state.loaded_characteristic_name = None
        if "new_characteristic_name" not in st.session_state:
            st.session_state.new_characteristic_name = ""


def sync_characteristic_from_global(name):
    if name not in st.session_state.characteristics:
        st.session_state.characteristics[name] = default_characteristic_state(name)
    state = st.session_state.characteristics[name]
    for key in CHARACTERISTIC_FIELDS:
        state[key] = st.session_state.get(key, state.get(key))
    state["measurement_name"] = sanitize_characteristic_name(
        st.session_state.get("measurement_name", name)
    ) or name
    state["results"] = dict(st.session_state.get("results", {}))
    state["summary"] = dict(st.session_state.get("summary", {}))
    state["figs"] = dict(st.session_state.get("figs", {}))
    worksheet = st.session_state.get("worksheet_data")
    if isinstance(worksheet, pd.DataFrame):
        state["worksheet_data"] = worksheet.copy()
    original = st.session_state.get("original_worksheet_data")
    if isinstance(original, pd.DataFrame):
        state["original_worksheet_data"] = original.copy()


def sync_global_from_characteristic(name):
    if name not in st.session_state.characteristics:
        st.session_state.characteristics[name] = default_characteristic_state(name)
    state = st.session_state.characteristics[name]
    for key in CHARACTERISTIC_FIELDS:
        try:
            st.session_state[key] = state.get(key)
        except Exception:
            # Widget-bound keys (e.g. tm, lsl, usl, mode) cannot be modified
            # after the widget has been instantiated in the current run; skip.
            pass
    try:
        st.session_state.measurement_name = state.get("measurement_name", name)
    except Exception:
        pass
    st.session_state.results = dict(state.get("results", {}))
    st.session_state.summary = dict(state.get("summary", {}))
    st.session_state.figs = dict(state.get("figs", {}))
    ws = state.get("worksheet_data")
    if isinstance(ws, pd.DataFrame):
        st.session_state.worksheet_data = ws.copy()
    else:
        st.session_state.worksheet_data = pd.DataFrame({"Value": [None] * 20})
    ows = state.get("original_worksheet_data", st.session_state.worksheet_data)
    if isinstance(ows, pd.DataFrame):
        st.session_state.original_worksheet_data = ows.copy()
    else:
        st.session_state.original_worksheet_data = st.session_state.worksheet_data.copy()



def sync_characteristic_state_machine():
    ensure_characteristics_state()
    active = st.session_state.active_characteristic_name
    loaded = st.session_state.loaded_characteristic_name
    if loaded is None:
        sync_global_from_characteristic(active)
        st.session_state.loaded_characteristic_name = active
    elif loaded != active:
        sync_characteristic_from_global(loaded)
        sync_global_from_characteristic(active)
        st.session_state.loaded_characteristic_name = active
    else:
        sync_characteristic_from_global(active)


def simplify_to_single_characteristic():
    ensure_characteristics_state()
    active_name = st.session_state.get("active_characteristic_name")
    if active_name not in st.session_state.characteristics:
        active_name = next(iter(st.session_state.characteristics))
    # Preserve any widget-driven state (e.g. mode radio button) by syncing
    # current global values INTO the characteristic BEFORE loading back.
    # This ensures the user's radio selection is not overwritten.
    if active_name in st.session_state.characteristics:
        for key in CHARACTERISTIC_FIELDS:
            if key in st.session_state:
                st.session_state.characteristics[active_name][key] = st.session_state[key]
    active_state = st.session_state.characteristics[active_name]
    st.session_state.characteristics = {active_name: active_state}
    st.session_state.active_characteristic_name = active_name
    st.session_state.loaded_characteristic_name = active_name
    sync_global_from_characteristic(active_name)


def set_active_characteristic(name):
    if name not in st.session_state.characteristics:
        st.session_state.characteristics[name] = default_characteristic_state(name)
    current_loaded = st.session_state.get("loaded_characteristic_name")
    if current_loaded:
        sync_characteristic_from_global(current_loaded)
    st.session_state.active_characteristic_name = name
    sync_global_from_characteristic(name)
    st.session_state.loaded_characteristic_name = name


def reset_active_characteristic_state():
    active_name = st.session_state.get("active_characteristic_name", "Characteristic 1")
    st.session_state.characteristics[active_name] = default_characteristic_state(
        active_name
    )
    st.session_state.loaded_characteristic_name = None
    for key in [
        "tm",
        "lsl",
        "usl",
        "target_index_value",
        "target_index_type",
        "confidence_level",
        "distribution",
        "hypothesis_type",
        "x_bar",
        "s",
        "n_samples",
        "decimal_places",
        "mode",
        "measurement_name",
        "description",
        "raw_data",
        "worksheet_measurement_name",
        "worksheet_description",
        "worksheet_tm",
        "worksheet_lsl",
        "worksheet_usl",
        "worksheet_data",
        "original_worksheet_data",
        "results",
        "summary",
        "figs",
    ]:
        st.session_state.pop(key, None)


def create_characteristic(name):
    new_name = sanitize_characteristic_name(name)
    if not new_name:
        return False, "Enter a characteristic name."
    if new_name in st.session_state.characteristics:
        return False, "That characteristic already exists."
    st.session_state.characteristics[new_name] = default_characteristic_state(new_name)
    set_active_characteristic(new_name)
    st.session_state.new_characteristic_name = ""
    return True, new_name


def delete_active_characteristic():
    if len(st.session_state.characteristics) <= 1:
        return False, "At least one characteristic must remain."
    active = st.session_state.active_characteristic_name
    st.session_state.characteristics.pop(active, None)
    next_name = next(iter(st.session_state.characteristics))
    set_active_characteristic(next_name)
    return True, next_name


def get_max_parts_count():
    max_count = 0
    for state in st.session_state.characteristics.values():
        worksheet = state.get("worksheet_data")
        if isinstance(worksheet, pd.DataFrame):
            max_count = max(max_count, len(worksheet))
    return max(max_count, len(st.session_state.get("part_ids", [])), 12)


def ensure_part_ids():
    target_len = get_max_parts_count()
    part_ids = list(st.session_state.get("part_ids", []))
    if len(part_ids) < target_len:
        part_ids.extend([""] * (target_len - len(part_ids)))
    st.session_state.part_ids = part_ids[:target_len]


def build_characteristic_matrix():
    ensure_part_ids()
    row_count = len(st.session_state.part_ids)
    matrix = {"DMC": st.session_state.part_ids[:row_count]}
    for name, state in st.session_state.characteristics.items():
        worksheet = state.get("worksheet_data")
        values = []
        if isinstance(worksheet, pd.DataFrame) and "Value" in worksheet.columns:
            values = worksheet["Value"].tolist()
        padded = values + [None] * max(0, row_count - len(values))
        matrix[name] = padded[:row_count]
    return pd.DataFrame(matrix)


def save_characteristic_matrix(matrix_df):
    cleaned_df = matrix_df.copy()
    st.session_state.part_ids = cleaned_df["DMC"].fillna("").astype(str).tolist()
    for name in list(st.session_state.characteristics.keys()):
        if name not in cleaned_df.columns:
            st.session_state.characteristics.pop(name, None)
    for column in cleaned_df.columns:
        if column == "DMC":
            continue
        if column not in st.session_state.characteristics:
            st.session_state.characteristics[column] = default_characteristic_state(column)
        values = cleaned_df[column].tolist()
        worksheet_df = pd.DataFrame({"Value": values})
        state = st.session_state.characteristics[column]
        state["worksheet_data"] = worksheet_df
        state["raw_data"] = ", ".join(
            map(str, worksheet_df["Value"].dropna().tolist())
        )
        if not state.get("transform_dirty", False):
            state["original_worksheet_data"] = worksheet_df.copy()
    if st.session_state.active_characteristic_name not in st.session_state.characteristics:
        st.session_state.active_characteristic_name = next(iter(st.session_state.characteristics))
    set_active_characteristic(st.session_state.active_characteristic_name)


def build_characteristic_metadata():
    rows = []
    for name, state in st.session_state.characteristics.items():
        rows.append(
            {
                "Characteristic": name,
                "Description": state.get("description", ""),
                "Target Mean": state.get("tm", 10.0),
                "LSL": state.get("lsl", 9.9),
                "USL": state.get("usl", 10.1),
            }
        )
    return pd.DataFrame(rows)


def save_characteristic_metadata(metadata_df):
    updated = {}
    for _, row in metadata_df.iterrows():
        raw_name = sanitize_characteristic_name(row.get("Characteristic"))
        if not raw_name:
            continue
        prior_state = st.session_state.characteristics.get(
            raw_name, default_characteristic_state(raw_name)
        )
        prior_state["measurement_name"] = raw_name
        prior_state["description"] = str(row.get("Description", "") or "")
        prior_state["tm"] = row.get("Target Mean", prior_state["tm"])
        prior_state["lsl"] = row.get("LSL", prior_state["lsl"])
        prior_state["usl"] = row.get("USL", prior_state["usl"])
        updated[raw_name] = prior_state
    if updated:
        st.session_state.characteristics = updated
        if st.session_state.active_characteristic_name not in updated:
            st.session_state.active_characteristic_name = next(iter(updated))
        set_active_characteristic(st.session_state.active_characteristic_name)


def run_characteristic_analysis(characteristic_name):
    state = st.session_state.characteristics[characteristic_name]
    analysis_inputs = dict(state)
    worksheet_values = []
    if state.get("mode") == "Use Data Worksheet":
        worksheet = state.get("worksheet_data")
        if isinstance(worksheet, pd.DataFrame) and "Value" in worksheet.columns:
            worksheet_values = [
                float(v) for v in worksheet["Value"].dropna().tolist()
                if calc.is_numeric(v)
            ]
        analysis_inputs["raw_data"] = ", ".join(map(str, worksheet_values))
        analysis_inputs["mode"] = "import"
    else:
        analysis_inputs["mode"] = "manual"

    results = calc.calculate(analysis_inputs)

    # Ensure importedData is always present when using worksheet mode
    if worksheet_values and not results.get("importedData"):
        results["importedData"] = worksheet_values

    summary = {}
    figs = {}
    if not results.get("error"):
        summary = get_summary_panel_content(results)
        results["verdict"] = summary.get("verdict", "N/A")
        fig_before, fig_after, fig_hist = plotter.update_plots(results)
        figs = {"before": fig_before, "after": fig_after, "hist": fig_hist}
    state["results"] = results
    state["summary"] = summary
    state["figs"] = figs
    return results, summary, figs


def analyze_all_characteristics():
    # Propagate the global data input mode to all characteristics
    global_mode = st.session_state.get("mode", "Use Data Worksheet")
    summaries = []
    for name in st.session_state.characteristics:
        # Set each characteristic's mode to match the global setting
        st.session_state.characteristics[name]["mode"] = global_mode
        results, summary, _ = run_characteristic_analysis(name)
        summaries.append(
            {
                "Characteristic": name,
                "Mode": global_mode,
                "Samples": results.get("n_samples"),
                "Cpk/Ppk": results.get("CpkCurrent", np.nan),
                "Verdict": summary.get("verdict", results.get("error", "Error")),
            }
        )
    active_name = st.session_state.active_characteristic_name
    set_active_characteristic(active_name)
    st.session_state.batch_results_df = pd.DataFrame(summaries)


@lru_cache(maxsize=64)
def calculate_descriptive_stats(values):
    data_array = np.asarray(values, dtype=float)
    if data_array.size < 2:
        return None

    q1, q2, q3 = np.percentile(data_array, [25, 50, 75])
    return {
        "count": int(data_array.size),
        "mean": float(np.mean(data_array)),
        "std": float(np.std(data_array, ddof=1)),
        "min": float(np.min(data_array)),
        "max": float(np.max(data_array)),
        "range": float(np.max(data_array) - np.min(data_array)),
        "q1": float(q1),
        "q2": float(q2),
        "q3": float(q3),
        "iqr": float(q3 - q1),
    }


def get_outlier_bounds(stats_summary, method):
    if method == "IQR (1.5×)":
        return (
            stats_summary["q1"] - 1.5 * stats_summary["iqr"],
            stats_summary["q3"] + 1.5 * stats_summary["iqr"],
        )

    sigma_multiplier = 3 if method == "3-Sigma" else 2
    return (
        stats_summary["mean"] - sigma_multiplier * stats_summary["std"],
        stats_summary["mean"] + sigma_multiplier * stats_summary["std"],
    )


def set_worksheet_data(values):
    worksheet_df = pd.DataFrame({"Value": list(values)})
    st.session_state.worksheet_data = worksheet_df
    st.session_state.raw_data = ", ".join(map(str, worksheet_df["Value"].dropna()))
    st.session_state.original_worksheet_data = worksheet_df.copy()
    st.session_state.transform_dirty = False
    active = st.session_state.get("active_characteristic_name")
    if active:
        sync_characteristic_from_global(active)


def apply_data_transformation(values, transform_type, **kwargs):
    data_arr = np.asarray(values, dtype=float)

    if transform_type == "Review & Remove Outliers (IQR)":
        q1, q3 = np.percentile(data_arr, [25, 75])
        iqr = q3 - q1
        mask = (data_arr >= q1 - 1.5 * iqr) & (data_arr <= q3 + 1.5 * iqr)
        return data_arr[mask], None

    if transform_type == "Gauge Rounding":
        return np.round(data_arr, kwargs.get("round_decimals", 3)), None

    if transform_type == "Offset Correction":
        return data_arr + kwargs.get("shift_value", 0.0), None

    if transform_type == "Unit Conversion / Scale":
        return data_arr * kwargs.get("scale_factor", 1.0), None

    return None, "Select a transformation before applying changes."


# Set page configuration
st.set_page_config(
    page_title="Statistical Process Capability & Data Analytics",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={"Get Help": None, "Report a bug": None, "About": None},
)

# Hide Streamlit's default Deploy button and menu, reduce top padding
st.markdown(
    """
<style>
    /* === HIDE STREAMLIT DEFAULTS === */
    .stDeployButton {display: none !important;}
    #MainMenu {display: none !important;}
    header {display: none !important;}
    footer {display: none !important;}
    .stMainBlockContainer {padding-top: 0.75rem !important;}
    .block-container {padding-top: 0.75rem !important; padding-left: 1.25rem !important; padding-right: 1.25rem !important;}
    
    html, body {
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        font-size: 15px;
    }
    
    /* === METRICS === */
    [data-testid="stMetric"] {
        padding: 0.85rem 1rem !important;
        border-radius: 10px;
        border-left: 3px solid #3b82f6;
    }
    
    /* === THEME-AWARE METRIC BACKGROUNDS === */
    @media (prefers-color-scheme: light) {
        [data-testid="stMetric"] {
            background: #f8fafc;
        }
    }
    @media (prefers-color-scheme: dark) {
        [data-testid="stMetric"] {
            background: #1e293b !important;
        }
        [data-testid="stMetric"] label,
        [data-testid="stMetric"] [data-testid="stMetricValue"],
        [data-testid="stMetric"] [data-testid="stMetricDelta"] {
            color: #e2e8f0 !important;
        }
    }
    
    /* Explicit dark mode overrides using all known Streamlit selectors */
    [data-theme="dark"] [data-testid="stMetric"],
    .stApp[data-theme="dark"] [data-testid="stMetric"],
    [data-testid="stAppViewContainer"][data-theme="dark"] [data-testid="stMetric"],
    .stApp.appview-container [data-testid="stMetric"] {
        background: #1e293b !important;
    }
    [data-theme="dark"] [data-testid="stMetric"] label,
    [data-theme="dark"] [data-testid="stMetric"] [data-testid="stMetricValue"],
    [data-theme="dark"] [data-testid="stMetric"] [data-testid="stMetricDelta"],
    .stApp[data-theme="dark"] [data-testid="stMetric"] label,
    .stApp[data-theme="dark"] [data-testid="stMetric"] [data-testid="stMetricValue"],
    .stApp[data-theme="dark"] [data-testid="stMetric"] [data-testid="stMetricDelta"] {
        color: #e2e8f0 !important;
    }
    
    /* === PROFESSIONAL NAVIGATION BAR === */
    .stTabs [data-baseweb="tab-list"] {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%) !important;
        border-radius: 12px !important;
        padding: 6px 8px !important;
        gap: 4px !important;
        box-shadow: 0 4px 16px rgba(0, 0, 0, 0.2), inset 0 1px 0 rgba(255,255,255,0.05) !important;
        border: 1px solid rgba(255,255,255,0.08) !important;
        margin-bottom: 1.25rem !important;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 0.65rem 1.25rem !important;
        font-weight: 500 !important;
        font-size: 0.92rem !important;
        letter-spacing: 0.02em !important;
        color: #94a3b8 !important;
        border-radius: 8px !important;
        transition: all 0.25s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border: 1px solid transparent !important;
        background: transparent !important;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: #e2e8f0 !important;
        background: rgba(59, 130, 246, 0.12) !important;
        border-color: rgba(59, 130, 246, 0.2) !important;
    }
    .stTabs [aria-selected="true"] {
        color: #ffffff !important;
        background: linear-gradient(135deg, #2563eb 0%, #3b82f6 100%) !important;
        border-color: rgba(96, 165, 250, 0.3) !important;
        box-shadow: 0 2px 8px rgba(37, 99, 235, 0.35), inset 0 1px 0 rgba(255,255,255,0.15) !important;
        font-weight: 600 !important;
    }
    .stTabs [data-baseweb="tab-highlight"] {
        display: none !important;
    }
    .stTabs [data-baseweb="tab-border"] {
        display: none !important;
    }
    
    .stButton > button,
    .stDownloadButton > button {
        min-height: 2.8rem;
    }
    
    [data-testid="stDataFrame"] {
        border-radius: 8px !important;
    }
    
    div[data-testid="stHorizontalBlock"] {
        align-items: stretch;
    }
    
    .stSelectbox, .stNumberInput, .stTextInput, .stRadio {
        margin-bottom: 0.25rem;
    }
</style>
""",
    unsafe_allow_html=True,
)


# --- Session State Initialization ---
def init_session_state(clear_form=False):
    defaults = {
        "tm": 10.00,
        "lsl": 9.90,
        "usl": 10.10,
        "target_index_value": 1.67,
        "target_index_type": "Cpk",
        "confidence_level": 95.0,
        "distribution": "Normal",
        "hypothesis_type": "Two-Sided",
        "x_bar": 10.00,
        "s": 0.015,
        "n_samples": 30,
        "decimal_places": 3,
        "mode": "Use Data Worksheet",
        "measurement_name": "",
        "raw_data": "",
        "transform_dirty": False,
        "last_uploaded_signature": None,
    }

    if "history" not in st.session_state:
        st.session_state.history = []

    if "chat_messages" not in st.session_state:
        st.session_state.chat_messages = []

    if "part_ids" not in st.session_state:
        st.session_state.part_ids = []

    if "batch_results_df" not in st.session_state:
        st.session_state.batch_results_df = pd.DataFrame()

    if "results" not in st.session_state:
        st.session_state.results = {}
        st.session_state.summary = {}
        st.session_state.figs = {}

    # Sigma Assistant mascot state
    if "mascot_state" not in st.session_state:
        st.session_state.mascot_state = "idle"
        st.session_state.mascot_cp = 1.0
        st.session_state.mascot_message = None

    if clear_form:
        active_name = sanitize_characteristic_name(
            st.session_state.get("active_characteristic_name", "Characteristic 1")
        ) or "Characteristic 1"
        st.session_state.results = {}
        st.session_state.summary = {}
        st.session_state.figs = {}
        st.session_state.chat_messages = []
        st.session_state.part_ids = []
        st.session_state.batch_results_df = pd.DataFrame()
        for key, value in defaults.items():
            st.session_state[key] = value
        st.session_state.measurement_name = active_name
        st.session_state.characteristics = {
            active_name: default_characteristic_state(active_name)
        }
        st.session_state.active_characteristic_name = active_name
        st.session_state.loaded_characteristic_name = None
    else:
        for key, value in defaults.items():
            if key not in st.session_state:
                st.session_state[key] = value


init_session_state()
ensure_characteristics_state()
sync_characteristic_state_machine()

# --- Main App UI ---
st.title("Statistical Process Capability & Data Analytics")

# Define Tabs
tab_analysis, tab_data, tab_viz, tab_history, tab_ref = st.tabs(
    ["Analysis & Report", "Data Worksheet", "Visualization", "History", "Reference"]
)

# --- Tab 1: Analysis & Report ---
with tab_analysis:
    # Display Error Messages
    if st.session_state.results and st.session_state.results.get("error"):
        st.error(f"**Analysis Error:** {st.session_state.results['error']}")

    # --- Characteristic Selector ---
    char_names = list(st.session_state.characteristics.keys())
    selector_cols = st.columns([2, 1])
    with selector_cols[0]:
        selected_char = st.selectbox(
            "Active Characteristic",
            char_names,
            index=char_names.index(st.session_state.active_characteristic_name) if st.session_state.active_characteristic_name in char_names else 0,
            key="analysis_char_selector",
            help="Select the characteristic to view results for. Add new characteristics in the Data Worksheet tab.",
        )
        if selected_char != st.session_state.active_characteristic_name:
            set_active_characteristic(selected_char)
            st.rerun()
    with selector_cols[1]:
        st.metric("Total Characteristics", len(char_names))

    active_characteristic = st.session_state.active_characteristic_name

    main_cols = st.columns([1.2, 1, 1])

    # --- Column 1: Input Parameters ---
    with main_cols[0]:
        st.header("I. Input Parameters")
        st.markdown(
            """
            <p style="font-size: 0.9rem; font-style: italic; color: inherit; opacity: 0.7;">
            Define product <b>specifications</b> and <b>measured process performance</b> data for the selected characteristic.
            </p>
            """,
            unsafe_allow_html=True,
        )

        with st.container(border=True):
            st.subheader("1. Specifications")
            spec_cols = st.columns(3)
            with spec_cols[0]:
                st.number_input(
                    "Tₘ (Target Mean)",
                    step=0.01,
                    key="tm",
                    help="The desired, ideal center of your process distribution.",
                )
            with spec_cols[1]:
                st.number_input(
                    "LSL (Lower Spec)",
                    step=0.01,
                    key="lsl",
                    help="The minimum acceptable value for your measurement.",
                )
            with spec_cols[2]:
                st.number_input(
                    "USL (Upper Spec)",
                    step=0.01,
                    key="usl",
                    help="The maximum acceptable value for your measurement.",
                )

        with st.container(border=True):
            st.subheader("2. Data & Goals")
            st.radio(
                "Data Input Mode",
                ["Use Data Worksheet", "Enter Manually"],
                key="mode",
                horizontal=True,
            )

            if st.session_state.mode == "Enter Manually":
                data_cols = st.columns(2)
                with data_cols[0]:
                    st.number_input(
                        "x̄ (Measured Mean)",
                        step=0.01,
                        key="x_bar",
                        help="The average value calculated from your sample data.",
                    )
                with data_cols[1]:
                    st.number_input(
                        "σ (Std Dev)",
                        step=0.001,
                        min_value=0.0,
                        format="%.5f",
                        key="s",
                        help="A measure of the amount of variation or dispersion of a set of values.",
                    )
            else:
                # Safely handle worksheet data that may not be initialized yet
                _ws_data = st.session_state.get("worksheet_data")
                if isinstance(_ws_data, pd.DataFrame) and "Value" in _ws_data.columns:
                    active_count = len(
                        coerce_valid_numeric_values(
                            _ws_data["Value"].dropna().tolist()
                        )
                    )
                else:
                    active_count = 0
                if active_count > 0:
                    st.success(
                        f"📊 Worksheet mode: **{active_count}** valid data points for `{active_characteristic}`."
                    )
                else:
                    st.warning(
                        f"⚠️ No valid data in worksheet for `{active_characteristic}`. Go to the **Data Worksheet** tab to enter values."
                    )

            st.number_input(
                "Target Index",
                step=0.01,
                key="target_index_value",
                help="The minimum capability value (e.g., Cpk 1.67) you aim for your process to achieve.",
            )
            st.selectbox(
                "Index Type",
                ["Cpk", "Cmk", "Ppk"],
                key="target_index_type",
                help="Capability Index Type: Cpk (short-term) or Ppk (long-term).",
            )

        with st.container(border=True):
            st.subheader("3. Statistical Settings")
            stat_cols_1 = st.columns(2)
            with stat_cols_1[0]:
                st.number_input(
                    "n (Samples)",
                    step=1,
                    min_value=2,
                    key="n_samples",
                    help="The number of data points in your sample. Must be >= 2.",
                )
                st.number_input(
                    "CL (%)",
                    min_value=1.0,
                    max_value=99.9,
                    step=0.1,
                    key="confidence_level",
                    help="Confidence Level for the Mean's Confidence Interval. 95% is common.",
                )
            with stat_cols_1[1]:
                st.number_input(
                    "Decimals",
                    min_value=1,
                    max_value=6,
                    step=1,
                    key="decimal_places",
                )
                st.text_input(
                    "Distribution",
                    value="Normal (automotive dimensional data default)",
                    disabled=True,
                    help="Dimensional capability calculations in this tool use the standard normal-process assumption.",
                )

            st.selectbox(
                "Hypothesis (μ vs Tₘ)",
                options=["Two-Sided", "Upper-Sided", "Lower-Sided"],
                format_func=lambda x: (
                    f"{x} (μ ≠ Tₘ)"
                    if x == "Two-Sided"
                    else (f"{x} (μ > Tₘ)" if x == "Upper-Sided" else f"{x} (μ < Tₘ)")
                ),
                key="hypothesis_type",
            )

        # Other buttons outside the form
        btn_cols = st.columns(3)
        with btn_cols[0]:
            submitted = st.button(
                "ANALYZE & PLOT", use_container_width=True, type="primary"
            )
        with btn_cols[1]:
            submitted_all = st.button(
                "⚡ ANALYZE ALL", use_container_width=True, type="secondary",
                help="Run analysis on ALL characteristics simultaneously.",
            )
        with btn_cols[2]:
            st.button(
                "RESET ACTIVE",
                use_container_width=True,
                on_click=reset_active_characteristic_state,
            )
    # --- Analysis Logic ---
    if submitted:
        # User clicked Analyze, so we run calculations for the active characteristic
        st.session_state.results, st.session_state.summary, st.session_state.figs = (
            run_characteristic_analysis(active_characteristic)
        )

        if not st.session_state.results.get("error"):
            # Update Sigma Assistant mascot state based on verdict
            verdict = st.session_state.summary.get("verdict", "")
            cp_value = st.session_state.results.get("Cp", 1.0)
            if "GOOD" in verdict:
                st.session_state.mascot_state = "happy"
                st.session_state.mascot_message = None  # Use default happy message
            elif "MARGINAL" in verdict:
                st.session_state.mascot_state = "concerned"
                st.session_state.mascot_message = None  # Use default concerned message
            elif "ACTION" in verdict or "INVALID" in verdict:
                st.session_state.mascot_state = "sad"
                st.session_state.mascot_message = None  # Use default sad message
            else:
                st.session_state.mascot_state = "idle"
                st.session_state.mascot_message = None
            st.session_state.mascot_cp = cp_value if cp_value and cp_value > 0 else 1.0

            # Save to history
            history_entry = st.session_state.results.copy()
            history_entry["id"] = datetime.datetime.now().isoformat()
            history_entry["characteristic_name"] = active_characteristic
            if "importedData" in history_entry:
                del history_entry["importedData"]  # Don't save large data array
            st.session_state.history.insert(0, history_entry)
            st.session_state.history = st.session_state.history[:250]  # Limit history

            # Generate plots
            fig_before, fig_after, fig_hist = plotter.update_plots(
                st.session_state.results
            )
            st.session_state.figs = {
                "before": fig_before,
                "after": fig_after,
                "hist": fig_hist,
            }
            sync_characteristic_from_global(active_characteristic)

        else:
            # Clear previous results if new run has errors
            st.session_state.summary = {}
            st.session_state.figs = {}
            sync_characteristic_from_global(active_characteristic)

        st.rerun()  # Rerun to display the new results

    # --- ANALYZE ALL Logic ---
    if submitted_all:
        analyze_all_characteristics()
        # Save all to history
        for char_name in st.session_state.characteristics:
            char_state = st.session_state.characteristics[char_name]
            char_results = char_state.get("results", {})
            if char_results and not char_results.get("error"):
                history_entry = char_results.copy()
                history_entry["id"] = datetime.datetime.now().isoformat()
                history_entry["characteristic_name"] = char_name
                if "importedData" in history_entry:
                    del history_entry["importedData"]
                st.session_state.history.insert(0, history_entry)
        st.session_state.history = st.session_state.history[:250]
        # Update global from active
        sync_global_from_characteristic(st.session_state.active_characteristic_name)
        st.rerun()

    # --- Batch Results Summary ---
    if "batch_results_df" in st.session_state and isinstance(st.session_state.batch_results_df, pd.DataFrame) and not st.session_state.batch_results_df.empty:
        st.divider()
        st.subheader("📊 Bulk Analysis Summary")
        batch_df = st.session_state.batch_results_df.copy()
        # Color-code the verdict column
        def _verdict_color(v):
            if "GOOD" in str(v).upper():
                return "background-color: #22c55e20; color: #22c55e;"
            elif "MARGINAL" in str(v).upper():
                return "background-color: #f59e0b20; color: #f59e0b;"
            elif "ACTION" in str(v).upper() or "INVALID" in str(v).upper():
                return "background-color: #ef444420; color: #ef4444;"
            return ""

        st.dataframe(
            batch_df.style.map(_verdict_color, subset=["Verdict"]),
            use_container_width=True,
            hide_index=True,
            column_config={
                "Cpk/Ppk": st.column_config.NumberColumn(format="%.3f"),
            },
        )

    # --- Column 2: Calculated Results ---
    with main_cols[1]:
        st.header("II. Calculated Results")
        st.markdown(
            """
            <p style="font-size: 0.9rem; font-style: italic; color: inherit; opacity: 0.7;">
            Key metrics based on the input data, including capability, spread, and recommended adjustments.
            </p>
            """,
            unsafe_allow_html=True,
        )

        res = st.session_state.results
        dp = res.get("dp", 3)

        def format_num(val, default="N/A", dps=None):
            if dps is None:
                dps = dp
            if val is None or not np.isfinite(val):
                return "∞" if val == np.inf else ("-∞" if val == -np.inf else default)
            return f"{val:.{dps}f}"

        if res and not res.get("error"):
            with st.container(border=True):
                st.markdown("**Current Process Metrics**")
                res_cols_1 = st.columns(2)
                with res_cols_1[0]:
                    st.metric(
                        label="Cₚ (Potential)",
                        value=format_num(res.get("Cp")),
                        help="Process Potential (Cp): Measures how capable the process would be if it were perfectly centered.",
                    )
                with res_cols_1[1]:
                    st.metric(
                        label=f"Current Index ({res.get('target_index_type', 'Cpk')})",
                        value=format_num(res.get("CpkCurrent")),
                        help="Process Capability (Cpk/Ppk): Measures the actual process capability, accounting for how centered it is.",
                    )

                st.markdown(
                    f"**6σ Spread:** **`{format_num(res.get('sixSigmaSpread'))}`**",
                    help="The range that contains approximately 99.73% of your process data (Mean ± 3 standard deviations).",
                )
                st.markdown(
                    f"_(x̄ ± 3σ): [ {format_num(res.get('minus3s'))}, {format_num(res.get('plus3s'))} ]_"
                )

                st.markdown(
                    f"**8σ Spread:** **`{format_num(res.get('eightSigmaSpread'))}`**",
                    help="A wider range containing about 99.9937% of data (Mean ± 4 standard deviations).",
                )
                st.markdown(
                    f"_(x̄ ± 4σ): [ {format_num(res.get('minus4s'))}, {format_num(res.get('plus4s'))} ]_"
                )

            res_cols_2 = st.columns(2)
            with res_cols_2[0]:
                with st.container(border=True):
                    st.metric(
                        label="Required Shift (Δ)",
                        value=format_num(res.get("shiftValue")),
                        help="The exact adjustment needed to move the measured process mean to the target mean (Tm).",
                    )
            with res_cols_2[1]:
                with st.container(border=True):
                    st.metric(
                        label=f"Req. Tolerance (Target {res.get('target_index_type')})",
                        value=format_num(res.get("newToleranceTotal")),
                        help="The minimum specification width (USL - LSL) your process needs to achieve its target capability index, given its current standard deviation.",
                    )

            with st.container(border=True):
                ci_label = f"Mean CI @ {res.get('confidence_level')}% ({res.get('hypothesis_type')})"
                ci_value = f"[{format_num(res.get('ci_lower'))}, {format_num(res.get('ci_upper'))}]"
                st.metric(
                    label=ci_label,
                    value=ci_value,
                    help="Confidence Interval (CI) for the Mean: The range within which the true population mean is likely to fall.",
                )

            if res.get("importedData"):
                with st.container(border=True):
                    st.markdown("**Data Summary (Import)**")
                    data_sum_cols = st.columns(3)
                    with data_sum_cols[0]:
                        st.metric("Mean", format_num(res.get("x_bar")))
                    with data_sum_cols[1]:
                        st.metric(
                            "Min", format_num(min(res.get("importedData", [np.nan])))
                        )
                    with data_sum_cols[2]:
                        st.metric(
                            "Max", format_num(max(res.get("importedData", [np.nan])))
                        )

                    if res.get("distribution") == "Lognormal" and np.isfinite(
                        res.get("mu_log", np.nan)
                    ):
                        log_cols = st.columns(2)
                        with log_cols[0]:
                            st.metric("Log-Mean (μ')", format_num(res.get("mu_log")))
                        with log_cols[1]:
                            st.metric(
                                "Log-Std Dev (σ')", format_num(res.get("sigma_log"))
                            )

            with st.container(border=True):
                st.markdown("**Probability & Defect Analysis**")
                prob_cols_1 = st.columns(2)
                with prob_cols_1[0]:
                    st.metric(
                        "P(x < LSL)",
                        f"{res.get('prob_below', 0) * 100:.3f}%",
                        help="The calculated chance that a single part will be produced below the Lower Specification Limit.",
                    )
                    st.metric(
                        "PPM < LSL",
                        f"{res.get('ppm_below', 0):,.0f}",
                        help="The expected number of defective parts per million that will fall below the Lower Specification Limit.",
                    )
                with prob_cols_1[1]:
                    st.metric(
                        "P(x > USL)",
                        f"{res.get('prob_above', 0) * 100:.3f}%",
                        help="The calculated chance that a single part will be produced above the Upper Specification Limit.",
                    )
                    st.metric(
                        "PPM > USL",
                        f"{res.get('ppm_above', 0):,.0f}",
                        help="The expected number of defective parts per million that will fall above the Upper Specification Limit.",
                    )
                st.metric(
                    "P(x < Tₘ)",
                    f"{res.get('prob_below_target', 0) * 100:.1f}%",
                    help="The probability that a single measurement will fall below the Target Mean. If your process is centered on the target, this should be 50%.",
                )
        else:
            st.info("Run analysis to see calculated results.")

    # --- Column 3: Summary & Interpretation ---
    with main_cols[2]:
        st.header("III. Summary & Interpretation")

        summary = st.session_state.summary

        if summary:
            verdict = summary.get("verdict", "ASSESSMENT PENDING")
            color = summary.get("verdict_color", "grey")
            st.markdown(
                f"""
                <div style="padding: 1rem; border-radius: 0.5rem; text-align: center; font-weight: 800; color: white; background-color: {color}; font-size: 1.25rem;">
                    {verdict}
                </div>
                """,
                unsafe_allow_html=True,
            )

            with st.container(border=True):
                st.markdown("<b>1. Process Centering:</b>", unsafe_allow_html=True)
                st.markdown(summary.get("centering", "..."), unsafe_allow_html=True)

                st.markdown(
                    "<br><b>2. Process Capability & Robustness:</b>",
                    unsafe_allow_html=True,
                )
                st.markdown(summary.get("capability", "..."), unsafe_allow_html=True)
                st.markdown(
                    f"<b>{summary.get('robustness', '...')}</b>", unsafe_allow_html=True
                )  # Style is harder here

                st.markdown("<br><b>3. Tolerance Adequacy:</b>", unsafe_allow_html=True)
                st.markdown(summary.get("tolerance", "..."), unsafe_allow_html=True)

                st.markdown(
                    "<br><b>4. Hypothesis Test (μ vs Tₘ):</b>", unsafe_allow_html=True
                )
                st.markdown(summary.get("hypothesis", "..."), unsafe_allow_html=True)

                st.divider()
                st.markdown("<b>Recommendations:</b>", unsafe_allow_html=True)
                st.markdown(
                    f"<ul>{''.join(summary.get('recommendations', []))}</ul>",
                    unsafe_allow_html=True,
                )
        else:
            st.info("Run analysis to see the summary and recommendations.")

# --- Tab 2: Data Worksheet ---
with tab_data:
    st.header("Data Worksheet")
    st.markdown(
        """
    <p style="font-size: 0.9rem; font-style: italic; color: inherit; opacity: 0.7;">
    Define multiple characteristics with tolerances and enter measurement values in the grid below.
    Each column represents one measurement characteristic.
    </p>
    """,
        unsafe_allow_html=True,
    )

    # --- 1. Characteristic Management ---
    with st.container(border=True):
        st.subheader("1. Characteristics & Tolerances")
        st.caption(
            "Add/remove characteristics and set their tolerances. Each characteristic gets its own column in the data grid."
        )

        # Add / Delete controls
        mgmt_cols = st.columns([3, 1, 1])
        with mgmt_cols[0]:
            new_char_name = st.text_input(
                "New Characteristic Name",
                key="new_characteristic_name_input",
                placeholder="e.g., Length, Width, Bore Diameter...",
                label_visibility="collapsed",
            )
        with mgmt_cols[1]:
            if st.button("➕ Add Characteristic", use_container_width=True):
                if new_char_name and new_char_name.strip():
                    ok, msg = create_characteristic(new_char_name.strip())
                    if ok:
                        st.rerun()
                    else:
                        st.warning(msg)
                else:
                    st.warning("Enter a name for the new characteristic.")
        with mgmt_cols[2]:
            if st.button("🗑️ Delete Active", use_container_width=True):
                if len(st.session_state.characteristics) > 1:
                    ok, msg = delete_active_characteristic()
                    if ok:
                        st.rerun()
                    else:
                        st.warning(msg)
                else:
                    st.warning("At least one characteristic must remain.")

        # Metadata table
        metadata_df = build_characteristic_metadata()
        edited_metadata = st.data_editor(
            metadata_df,
            num_rows="fixed",
            use_container_width=True,
            hide_index=True,
            column_config={
                "Characteristic": st.column_config.TextColumn(
                    "Characteristic Name",
                    help="Name of the measurement characteristic.",
                    disabled=True,
                ),
                "Description": st.column_config.TextColumn(
                    "Description",
                    help="Short engineering note (e.g., 'Outer diameter before plating').",
                ),
                "Target Mean": st.column_config.NumberColumn(
                    "Target Mean (Tₘ)",
                    help="The desired center value for this dimension.",
                    format="%.4f",
                    step=0.001,
                ),
                "LSL": st.column_config.NumberColumn(
                    "LSL",
                    help="Lower Specification Limit.",
                    format="%.4f",
                    step=0.001,
                ),
                "USL": st.column_config.NumberColumn(
                    "USL",
                    help="Upper Specification Limit.",
                    format="%.4f",
                    step=0.001,
                ),
            },
            key="metadata_editor",
        )
        save_characteristic_metadata(edited_metadata)

    # --- 2. Data Import ---
    with st.container(border=True):
        st.subheader("2. Data Import")
        upload_cols = st.columns([2, 1, 1])

        with upload_cols[0]:
            uploaded_file = st.file_uploader(
                "Upload CSV or Excel file",
                type=["csv", "xlsx", "xls"],
                help="Each numeric column becomes a characteristic. A DMC/Serial/Part column is auto-detected.",
            )

        with upload_cols[1]:
            st.markdown("**Paste options:**")
            paste_mode = st.radio(
                "Paste format",
                ["Comma separated", "Newline separated", "Tab separated"],
                horizontal=False,
                label_visibility="collapsed",
            )

        with upload_cols[2]:
            st.markdown("**Quick actions:**")


            def _make_template():
                """Generate a pre-formatted .xlsx template with columns matching current characteristics."""
                wb = Workbook()
                ws = wb.active
                ws.title = "SPC Data"

                char_names = list(st.session_state.characteristics.keys())
                n_chars = max(len(char_names), 2)  # At least 2 columns

                # Column widths
                ws.column_dimensions["A"].width = 8
                ws.column_dimensions["B"].width = 28
                for ci in range(n_chars):
                    col_letter = get_column_letter(ci + 3)
                    ws.column_dimensions[col_letter].width = 20

                hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
                hdr_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
                sub_hdr_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
                sub_hdr_font = Font(name="Calibri", bold=True, size=10, color="94a3b8")
                hdr_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin", color="334155"),
                    right=Side(style="thin", color="334155"),
                    top=Side(style="thin", color="334155"),
                    bottom=Side(style="thin", color="334155"),
                )

                # --- Row 1: Main headers ---
                headers = ["#", "DMC / Serial Number"]
                for ci in range(n_chars):
                    name = char_names[ci] if ci < len(char_names) else f"Measurement {ci + 1}"
                    headers.append(name)
                for col_idx, header in enumerate(headers, start=1):
                    c = ws.cell(row=1, column=col_idx, value=header)
                    c.font = hdr_font
                    c.fill = hdr_fill
                    c.alignment = hdr_align
                    c.border = thin_border

                # --- Row 2: Tolerance sub-header (Tm | LSL | USL) ---
                tol_c1 = ws.cell(row=2, column=1, value="")
                tol_c1.fill = sub_hdr_fill
                tol_c1.border = thin_border
                tol_c2 = ws.cell(row=2, column=2, value="Tolerances →")
                tol_c2.font = sub_hdr_font
                tol_c2.fill = sub_hdr_fill
                tol_c2.alignment = hdr_align
                tol_c2.border = thin_border
                for ci in range(n_chars):
                    char_state = {}
                    if ci < len(char_names):
                        char_state = st.session_state.characteristics.get(char_names[ci], {})
                    tm = char_state.get("tm", 10.0)
                    lsl = char_state.get("lsl", 9.9)
                    usl = char_state.get("usl", 10.1)
                    tol_text = f"Tm={tm} | LSL={lsl} | USL={usl}"
                    tc = ws.cell(row=2, column=ci + 3, value=tol_text)
                    tc.font = sub_hdr_font
                    tc.fill = sub_hdr_fill
                    tc.alignment = hdr_align
                    tc.border = thin_border

                # --- Sample data rows ---
                sample_rng = np.random.default_rng(99)
                input_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                alt_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
                for i in range(1, 11):
                    r = i + 2
                    ws.cell(row=r, column=1, value=i).border = thin_border
                    ws.cell(row=r, column=1).alignment = hdr_align
                    ws.cell(row=r, column=2, value=f"DMC-2024-{i:03d}").border = thin_border
                    fill = alt_fill if i % 2 == 0 else input_fill
                    ws.cell(row=r, column=2).fill = fill
                    for ci in range(n_chars):
                        char_state = {}
                        if ci < len(char_names):
                            char_state = st.session_state.characteristics.get(char_names[ci], {})
                        center = char_state.get("tm", 10.0)
                        spread = (char_state.get("usl", center + 0.1) - char_state.get("lsl", center - 0.1)) / 6
                        val = round(sample_rng.normal(center, max(spread, 0.01)), 4)
                        cell = ws.cell(row=r, column=ci + 3, value=val)
                        cell.border = thin_border
                        cell.number_format = "0.0000"
                        cell.fill = fill

                # --- Empty rows up to 10,000 ---
                for i in range(11, 10001):
                    r = i + 2
                    ws.cell(row=r, column=1, value=i).border = thin_border
                    ws.cell(row=r, column=1).alignment = hdr_align
                    ws.cell(row=r, column=2).border = thin_border
                    fill = alt_fill if i % 2 == 0 else input_fill
                    ws.cell(row=r, column=2).fill = fill
                    for ci in range(n_chars):
                        cell = ws.cell(row=r, column=ci + 3)
                        cell.border = thin_border
                        cell.number_format = "0.0000"
                        cell.fill = fill

                note_row = 10004
                ws.cell(row=note_row, column=2, value="💡 Replace sample data with your actual measurements.").font = Font(
                    name="Calibri", size=9, italic=True, color="6B7280")
                ws.cell(row=note_row + 1, column=2, value="📤 Each numeric column becomes a separate characteristic.").font = Font(
                    name="Calibri", size=9, italic=True, color="6B7280")
                ws.cell(row=note_row + 2, column=2, value=f"📊 Template generated with {n_chars} characteristics, up to 10,000 DMC rows.").font = Font(
                    name="Calibri", size=9, italic=True, color="6B7280")

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            st.download_button(
                "📥 Download Template",
                data=_make_template(),
                file_name="SPC_Data_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Download a pre-formatted Excel template with multiple measurement columns.",
            )
            if st.button("Clear All Data", use_container_width=True):
                st.session_state.last_uploaded_signature = None
                for char_name in st.session_state.characteristics:
                    st.session_state.characteristics[char_name]["worksheet_data"] = pd.DataFrame({"Value": [None] * 100})
                    st.session_state.characteristics[char_name]["original_worksheet_data"] = pd.DataFrame({"Value": [None] * 100})
                    st.session_state.characteristics[char_name]["raw_data"] = ""
                    st.session_state.characteristics[char_name]["transform_dirty"] = False
                st.session_state.part_ids = []
                sync_global_from_characteristic(st.session_state.active_characteristic_name)
                st.rerun()
            if st.button("Sample Data (All)", use_container_width=True):
                st.session_state.last_uploaded_signature = None
                sample_rng = np.random.default_rng(42)
                for idx, char_name in enumerate(st.session_state.characteristics):
                    char_state = st.session_state.characteristics[char_name]
                    center = char_state.get("tm", 10.0)
                    spread = (char_state.get("usl", center + 0.1) - char_state.get("lsl", center - 0.1)) / 6
                    if spread <= 0:
                        spread = 0.02
                    sample_data = sample_rng.normal(center, spread, 10000).round(4)
                    char_state["worksheet_data"] = pd.DataFrame({"Value": list(sample_data)})
                    char_state["original_worksheet_data"] = pd.DataFrame({"Value": list(sample_data)})
                    char_state["raw_data"] = ""  # Skip string storage for large datasets
                    char_state["transform_dirty"] = False
                sync_global_from_characteristic(st.session_state.active_characteristic_name)
                st.rerun()

    # Process uploaded file (multi-column support)
    if uploaded_file is not None:
        upload_signature = (
            uploaded_file.name,
            getattr(uploaded_file, "size", None),
        )
        if st.session_state.get("last_uploaded_signature") != upload_signature:
            try:
                if uploaded_file.name.endswith(".csv"):
                    df_uploaded = pd.read_csv(uploaded_file)
                else:
                    df_uploaded = pd.read_excel(uploaded_file)

                # Detect DMC column
                potential_dmc_cols = [
                    column
                    for column in df_uploaded.columns
                    if "dmc" in str(column).lower()
                    or "serial" in str(column).lower()
                    or "part" in str(column).lower()
                ]
                if potential_dmc_cols:
                    dmc_values = (
                        df_uploaded[potential_dmc_cols[0]].fillna("").astype(str).tolist()
                    )
                    st.session_state.part_ids = dmc_values

                # Each numeric column becomes a characteristic
                numeric_cols = df_uploaded.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    imported_count = 0
                    for col_name in numeric_cols:
                        char_name = str(col_name).strip()
                        if not char_name:
                            char_name = f"Measurement {imported_count + 1}"
                        values = df_uploaded[col_name].tolist()
                        valid_count = len(pd.Series(values).dropna())

                        if char_name not in st.session_state.characteristics:
                            st.session_state.characteristics[char_name] = default_characteristic_state(char_name)

                        st.session_state.characteristics[char_name]["worksheet_data"] = pd.DataFrame(
                            {"Value": values}
                        )
                        st.session_state.characteristics[char_name]["original_worksheet_data"] = pd.DataFrame(
                            {"Value": values}
                        )
                        st.session_state.characteristics[char_name]["raw_data"] = ", ".join(
                            map(str, pd.Series(values).dropna().tolist())
                        )
                        st.session_state.characteristics[char_name]["transform_dirty"] = False
                        imported_count += 1

                    st.session_state.last_uploaded_signature = upload_signature
                    # Set active to first imported characteristic
                    first_char = str(numeric_cols[0]).strip() or "Measurement 1"
                    if first_char in st.session_state.characteristics:
                        set_active_characteristic(first_char)
                    st.success(
                        f"✅ Imported **{imported_count}** characteristic(s) from `{uploaded_file.name}` with up to {len(df_uploaded)} rows each."
                    )
                else:
                    st.error("No numeric columns found in the uploaded file.")
            except Exception as e:
                st.error(f"Error reading file: {e}")

    # --- 3. Multi-Column Measurement Grid ---
    st.subheader("3. Parts Worksheet")
    st.caption(
        "Each row is one part. `DMC` is the part identifier. Each additional column is a measurement characteristic."
    )

    ensure_part_ids()
    matrix_df = build_characteristic_matrix()

    # Build column config dynamically
    column_config = {
        "DMC": st.column_config.TextColumn(
            "DMC / Serial Number",
            help="Data Matrix Code or unique part identifier.",
        ),
    }
    for char_name in st.session_state.characteristics:
        char_state = st.session_state.characteristics[char_name]
        desc = char_state.get("description", "") or "Measured actual value."
        column_config[char_name] = st.column_config.NumberColumn(
            char_name,
            help=f"{desc} | Tₘ={char_state.get('tm', '')}, LSL={char_state.get('lsl', '')}, USL={char_state.get('usl', '')}",
            format="%.4f",
            step=0.0001,
        )

    edited_matrix_df = st.data_editor(
        matrix_df,
        num_rows="dynamic",
        use_container_width=True,
        height=480,
        hide_index=True,
        column_config=column_config,
        key="parts_matrix_editor",
    )

    # Save edited matrix back to characteristic states
    save_characteristic_matrix(edited_matrix_df)

    # --- Status bar ---
    status_parts = []
    for char_name in st.session_state.characteristics:
        char_state = st.session_state.characteristics[char_name]
        ws = char_state.get("worksheet_data")
        if isinstance(ws, pd.DataFrame) and "Value" in ws.columns:
            count = len(coerce_valid_numeric_values(ws["Value"].dropna().tolist()))
            if count > 0:
                status_parts.append(f"`{char_name}`: **{count}** pts")
    if status_parts:
        st.success(
            f"✅ Data ready — {', '.join(status_parts)}. Go to **Analysis & Report** → select **Use Data Worksheet** mode → click **ANALYZE ALL**."
        )
    else:
        st.info("Enter data in the grid above or upload a file to get started.")




# --- Tab 3: Visualization ---
with tab_viz:
    st.header("Visualization")
    st.markdown(
        """
    <p style="font-size: 0.9rem; color: inherit; opacity: 0.7;">
    Interactive charts for each characteristic. Select a sub-tab to view charts.
    </p>
    """,
        unsafe_allow_html=True,
    )

    # Chart settings
    show_annotations = st.checkbox(
        "Show Annotations", value=True, key="show_annotations"
    )

    # Dynamic sub-tabs for each characteristic
    viz_char_names = list(st.session_state.characteristics.keys())
    if len(viz_char_names) == 0:
        st.info("No characteristics defined. Go to Data Worksheet to add one.")
    else:
        viz_sub_tabs = st.tabs(viz_char_names)

        for viz_tab_idx, viz_char_name in enumerate(viz_char_names):
            with viz_sub_tabs[viz_tab_idx]:
                viz_char_state = st.session_state.characteristics[viz_char_name]
                figs = viz_char_state.get("figs", {})
                res = viz_char_state.get("results", {})

                # Get visualization data from this characteristic's worksheet
                viz_data = []
                viz_ws = viz_char_state.get("worksheet_data")
                if isinstance(viz_ws, pd.DataFrame) and "Value" in viz_ws.columns:
                    viz_data = coerce_valid_numeric_values(
                        viz_ws["Value"].dropna().tolist()
                    )

                if len(viz_data) >= 2:
                    st.subheader(f"Worksheet Distribution — {viz_char_name}")
                    preview_cols = st.columns(2)

                    with preview_cols[0]:
                        fig_hist_preview = go.Figure()
                        fig_hist_preview.add_trace(
                            go.Histogram(
                                x=viz_data,
                                nbinsx=20,
                                marker_color="#3B82F6",
                                opacity=0.75,
                                name="Data",
                            )
                        )
                        fig_hist_preview.update_layout(
                            title="Distribution Histogram",
                            height=300,
                            margin=dict(l=40, r=20, t=50, b=40),
                            showlegend=False,
                            template="plotly_white",
                        )
                        st.plotly_chart(
                            fig_hist_preview,
                            use_container_width=True,
                            config=PlotManager.PLOT_CONFIG,
                            key=f"viz_hist_preview_{viz_char_name}",
                        )

                    with preview_cols[1]:
                        fig_box = go.Figure()
                        fig_box.add_trace(
                            go.Box(
                                y=viz_data,
                                marker_color="#10B981",
                                boxpoints="outliers",
                                name="Values",
                            )
                        )
                        fig_box.update_layout(
                            title="Box Plot",
                            height=300,
                            margin=dict(l=40, r=20, t=50, b=40),
                            showlegend=False,
                            template="plotly_white",
                        )
                        st.plotly_chart(
                            fig_box, use_container_width=True, config=PlotManager.PLOT_CONFIG,
                            key=f"viz_box_{viz_char_name}",
                        )

                if figs and figs.get("before") and figs.get("after"):
                    viz_cols = st.columns(2)
                    with viz_cols[0]:
                        st.plotly_chart(
                            figs["before"], use_container_width=True, config=PlotManager.PLOT_CONFIG,
                            key=f"viz_before_{viz_char_name}",
                        )
                    with viz_cols[1]:
                        st.plotly_chart(
                            figs["after"], use_container_width=True, config=PlotManager.PLOT_CONFIG,
                            key=f"viz_after_{viz_char_name}",
                        )

                    if figs.get("hist"):
                        st.subheader("Data Distribution Analysis")
                        st.plotly_chart(
                            figs["hist"], use_container_width=True, config=PlotManager.PLOT_CONFIG,
                            key=f"viz_hist_{viz_char_name}",
                        )

                    # --- Control Charts (I-Chart + MR-Chart with Filter) ---
                    # Use importedData from results if available, otherwise fall back to worksheet data
                    control_chart_data = res.get("importedData", []) if res else []
                    if not control_chart_data and viz_data:
                        control_chart_data = viz_data
                    if len(control_chart_data) >= 5:
                        st.subheader("📊 Control Charts")

                        data_points_all = control_chart_data
                        total_n = len(data_points_all)

                        # --- Filter control ---
                        ctrl_cols = st.columns([1, 2, 1])
                        with ctrl_cols[0]:
                            filter_options = [10, 25, 50, 100, 250, 500, "All"]
                            valid_options = [opt for opt in filter_options
                                             if opt == "All" or (isinstance(opt, int) and opt <= total_n)]
                            if not valid_options or valid_options[-1] != "All":
                                valid_options.append("All")
                            default_idx = min(2, len(valid_options) - 1)
                            show_n = st.selectbox(
                                "Show Points",
                                valid_options,
                                index=default_idx,
                                key=f"ctrl_chart_filter_{viz_char_name}",
                                help="Filter the number of data points displayed in control charts",
                            )
                        with ctrl_cols[1]:
                            effective_n = total_n if show_n == "All" else int(show_n)
                            st.info(f"Showing **{min(effective_n, total_n)}** of **{total_n}** data points")
                        with ctrl_cols[2]:
                            show_warnings = st.checkbox("Show Warning Limits (±2σ)", value=True, key=f"show_uwl_{viz_char_name}")

                        # Slice data
                        data_points = data_points_all[:effective_n]
                        n = len(data_points)
                        x_bar = float(np.mean(data_points))
                        s = float(np.std(data_points, ddof=1)) if n >= 2 else 0.0

                        # ±1σ zone lines
                        plus_1s = x_bar + 1 * s
                        minus_1s = x_bar - 1 * s

                        # I-MR constants
                        ucl = x_bar + 3 * s
                        lcl = x_bar - 3 * s
                        uwl = x_bar + 2 * s
                        lwl = x_bar - 2 * s

                        # Specification lines from characteristic state
                        _lsl = float(viz_char_state.get("lsl", 0))
                        _usl = float(viz_char_state.get("usl", 0))
                        _tm = float(viz_char_state.get("tm", 0))

                        # Moving Range
                        mr_values = [abs(data_points[i] - data_points[i - 1]) for i in range(1, n)]
                        mr_bar = float(np.mean(mr_values)) if mr_values else 0.0
                        mr_ucl = 3.267 * mr_bar  # D4 for n=2

                        # ====== I-CHART ======
                        fig_control = go.Figure()

                        fig_control.add_trace(
                            go.Scatter(
                                x=list(range(1, n + 1)),
                                y=data_points,
                                mode="lines+markers",
                                name="Individual Value",
                                line=dict(color="#3B82F6", width=2),
                                marker=dict(size=5, color="#3B82F6"),
                                hovertemplate="Sample %{x}<br>Value: %{y:.4f}<extra></extra>",
                            )
                        )

                        # Center line
                        fig_control.add_trace(
                            go.Scatter(
                                x=[1, n], y=[x_bar, x_bar],
                                mode="lines", name=f"CL x̄ = {x_bar:.4f}",
                                line=dict(color="#10B981", width=2, dash="solid"),
                            )
                        )
                        # UCL / LCL (±3σ)
                        fig_control.add_trace(
                            go.Scatter(
                                x=[1, n], y=[ucl, ucl],
                                mode="lines", name=f"UCL x̄+3σ = {ucl:.4f}",
                                line=dict(color="#EF4444", width=1.5, dash="dash"),
                            )
                        )
                        fig_control.add_trace(
                            go.Scatter(
                                x=[1, n], y=[lcl, lcl],
                                mode="lines", name=f"LCL x̄−3σ = {lcl:.4f}",
                                line=dict(color="#EF4444", width=1.5, dash="dash"),
                            )
                        )

                        # ±2σ Warning limits
                        if show_warnings:
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[1, n], y=[uwl, uwl],
                                    mode="lines", name=f"+2σ = {uwl:.4f}",
                                    line=dict(color="#F59E0B", width=1, dash="dot"),
                                )
                            )
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[1, n], y=[lwl, lwl],
                                    mode="lines", name=f"−2σ = {lwl:.4f}",
                                    line=dict(color="#F59E0B", width=1, dash="dot"),
                                )
                            )

                        # ±1σ zone lines
                        fig_control.add_trace(
                            go.Scatter(
                                x=[1, n], y=[plus_1s, plus_1s],
                                mode="lines", name=f"+1σ = {plus_1s:.4f}",
                                line=dict(color="#A78BFA", width=1, dash="dot"),
                                visible="legendonly",
                            )
                        )
                        fig_control.add_trace(
                            go.Scatter(
                                x=[1, n], y=[minus_1s, minus_1s],
                                mode="lines", name=f"−1σ = {minus_1s:.4f}",
                                line=dict(color="#A78BFA", width=1, dash="dot"),
                                visible="legendonly",
                            )
                        )

                        # Specification lines (LSL / USL / Target)
                        if _usl > _lsl:
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[1, n], y=[_usl, _usl],
                                    mode="lines", name=f"USL = {_usl:.3f}",
                                    line=dict(color="#059669", width=2, dash="dashdot"),
                                )
                            )
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[1, n], y=[_lsl, _lsl],
                                    mode="lines", name=f"LSL = {_lsl:.3f}",
                                    line=dict(color="#059669", width=2, dash="dashdot"),
                                )
                            )
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[1, n], y=[_tm, _tm],
                                    mode="lines", name=f"Target = {_tm:.3f}",
                                    line=dict(color="#8B5CF6", width=1.5, dash="longdash"),
                                )
                            )

                        # Out-of-control points
                        ooc_indices = [i for i, v in enumerate(data_points) if v > ucl or v < lcl]
                        if ooc_indices:
                            fig_control.add_trace(
                                go.Scatter(
                                    x=[i + 1 for i in ooc_indices],
                                    y=[data_points[i] for i in ooc_indices],
                                    mode="markers", name="Out of Control",
                                    marker=dict(size=12, color="#EF4444", symbol="circle-open", line=dict(width=2)),
                                )
                            )

                        _fc = "#8b95a5"

                        # Right-side zone annotations
                        zone_annotations = [
                            dict(x=1.02, y=ucl, xref="paper", yref="y", text="UCL (x̄+3σ)", showarrow=False,
                                 font=dict(size=9, color="#EF4444"), xanchor="left"),
                            dict(x=1.02, y=lcl, xref="paper", yref="y", text="LCL (x̄−3σ)", showarrow=False,
                                 font=dict(size=9, color="#EF4444"), xanchor="left"),
                            dict(x=1.02, y=x_bar, xref="paper", yref="y", text="CL (x̄)", showarrow=False,
                                 font=dict(size=9, color="#10B981"), xanchor="left"),
                        ]
                        if show_warnings:
                            zone_annotations.extend([
                                dict(x=1.02, y=uwl, xref="paper", yref="y", text="Zone B (+2σ)", showarrow=False,
                                     font=dict(size=8, color="#F59E0B"), xanchor="left"),
                                dict(x=1.02, y=lwl, xref="paper", yref="y", text="Zone B (−2σ)", showarrow=False,
                                     font=dict(size=8, color="#F59E0B"), xanchor="left"),
                            ])
                        # Zone labels in the middle of chart
                        if s > 0:
                            zone_annotations.extend([
                                dict(x=0.98, y=(x_bar + plus_1s) / 2, xref="paper", yref="y", text="Zone C",
                                     showarrow=False, font=dict(size=8, color="rgba(128,128,128,0.5)"), xanchor="right"),
                                dict(x=0.98, y=(plus_1s + uwl) / 2, xref="paper", yref="y", text="Zone B",
                                     showarrow=False, font=dict(size=8, color="rgba(128,128,128,0.5)"), xanchor="right"),
                                dict(x=0.98, y=(uwl + ucl) / 2, xref="paper", yref="y", text="Zone A",
                                     showarrow=False, font=dict(size=8, color="rgba(128,128,128,0.5)"), xanchor="right"),
                            ])

                        _ctrl_layout = dict(
                            height=420,
                            margin=dict(t=55, b=65, l=55, r=70),
                            hovermode="x unified",
                            xaxis=dict(title=dict(text="Sample Number", font=dict(color=_fc, size=11)),
                                       tickfont=dict(size=10, color=_fc),
                                       gridcolor="rgba(128,128,128,0.15)"),
                            yaxis=dict(title=dict(text="Value", font=dict(color=_fc, size=11)),
                                       tickfont=dict(size=10, color=_fc),
                                       gridcolor="rgba(128,128,128,0.15)"),
                            legend=dict(orientation="h", y=-0.22, x=0.5, xanchor="center",
                                        bgcolor="rgba(128,128,128,0.08)", font=dict(size=9, color=_fc)),
                            hoverlabel=dict(font_size=11, bgcolor="rgba(30,41,59,0.92)",
                                            font_color="#e2e8f0", bordercolor="rgba(128,128,128,0.3)"),
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color=_fc),
                        )

                        fig_control.update_layout(
                            title=dict(text=f"I-Chart — {viz_char_name} ({n} points)", font=dict(size=12, color=_fc)),
                            annotations=zone_annotations,
                            **_ctrl_layout,
                        )

                        st.plotly_chart(fig_control, use_container_width=True, config=PlotManager.PLOT_CONFIG,
                                        key=f"viz_ichart_{viz_char_name}")

                        # Alert
                        if ooc_indices:
                            st.warning(
                                f"⚠️ {len(ooc_indices)} point(s) outside control limits at samples: {', '.join(map(str, [i + 1 for i in ooc_indices[:20]]))}"
                                + (f" ... and {len(ooc_indices) - 20} more" if len(ooc_indices) > 20 else "")
                            )
                        else:
                            st.success("✅ All points within control limits — process is in statistical control")

                        # ====== STATISTICS SUMMARY PANEL ======
                        st.markdown("---")
                        st.subheader(f"📋 Control Chart Statistics — {viz_char_name}")

                        _cp = ((_usl - _lsl) / (6 * s)) if s > 0 and _usl > _lsl else float("inf")
                        _cpk = min((_usl - x_bar) / (3 * s), (x_bar - _lsl) / (3 * s)) if s > 0 and _usl > _lsl else float("inf")
                        _ppm_above = sum(1 for v in data_points if v > _usl)
                        _ppm_below = sum(1 for v in data_points if v < _lsl)
                        _zone_a = sum(1 for v in data_points if v > uwl or v < lwl)
                        _zone_b = sum(1 for v in data_points if (uwl >= v > plus_1s) or (lwl <= v < minus_1s))
                        _zone_c = sum(1 for v in data_points if minus_1s <= v <= plus_1s)
                        _sigma_level = abs(x_bar - _tm) / s if s > 0 else 0.0

                        stat_cols = st.columns(4)
                        with stat_cols[0]:
                            st.markdown("**📊 Central Tendency**")
                            st.markdown(f"""
| Metric | Value |
|--------|-------|
| x̄ (Mean) | `{x_bar:.5f}` |
| Target (Tₘ) | `{_tm:.3f}` |
| Shift (Δ) | `{x_bar - _tm:.5f}` |
| σ | `{s:.5f}` |
| n | `{n}` |
""")

                        with stat_cols[1]:
                            st.markdown("**📏 Control Limits**")
                            st.markdown(f"""
| Limit | Value |
|-------|-------|
| UCL (x̄+3σ) | `{ucl:.5f}` |
| +2σ | `{uwl:.5f}` |
| +1σ | `{plus_1s:.5f}` |
| CL (x̄) | `{x_bar:.5f}` |
| −1σ | `{minus_1s:.5f}` |
| −2σ | `{lwl:.5f}` |
| LCL (x̄−3σ) | `{lcl:.5f}` |
""")

                        with stat_cols[2]:
                            st.markdown("**🎯 Capability**")
                            cp_display = f"{_cp:.3f}" if _cp < 999 else "∞"
                            cpk_display = f"{_cpk:.3f}" if _cpk < 999 else "∞"
                            st.markdown(f"""
| Metric | Value |
|--------|-------|
| Cp | `{cp_display}` |
| Cpk | `{cpk_display}` |
| 6σ Spread | `{6*s:.5f}` |
| 8σ Spread | `{8*s:.5f}` |
| LSL | `{_lsl:.3f}` |
| USL | `{_usl:.3f}` |
| Tolerance | `{_usl - _lsl:.3f}` |
""")

                        with stat_cols[3]:
                            st.markdown("**🔍 Zone Analysis**")
                            st.markdown(f"""
| Zone | Count | % |
|------|-------|---|
| Zone A (±2-3σ) | `{_zone_a}` | `{_zone_a/n*100:.1f}%` |
| Zone B (±1-2σ) | `{_zone_b}` | `{_zone_b/n*100:.1f}%` |
| Zone C (±1σ) | `{_zone_c}` | `{_zone_c/n*100:.1f}%` |
| OOC (>3σ) | `{len(ooc_indices)}` | `{len(ooc_indices)/n*100:.1f}%` |
| > USL | `{_ppm_above}` | `{_ppm_above/n*100:.2f}%` |
| < LSL | `{_ppm_below}` | `{_ppm_below/n*100:.2f}%` |
| MR̄ | `{mr_bar:.5f}` | — |
""")

                        st.markdown("---")

                        # ====== MR-CHART ======
                        fig_mr = go.Figure()

                        fig_mr.add_trace(
                            go.Scatter(
                                x=list(range(2, n + 1)),
                                y=mr_values,
                                mode="lines+markers",
                                name="Moving Range",
                                line=dict(color="#F97316", width=2),
                                marker=dict(size=5, color="#F97316"),
                                hovertemplate="Sample %{x}<br>MR: %{y:.4f}<extra></extra>",
                            )
                        )
                        fig_mr.add_trace(
                            go.Scatter(
                                x=[2, n], y=[mr_bar, mr_bar],
                                mode="lines", name=f"MR̄ ({mr_bar:.4f})",
                                line=dict(color="#10B981", width=2, dash="solid"),
                            )
                        )
                        fig_mr.add_trace(
                            go.Scatter(
                                x=[2, n], y=[mr_ucl, mr_ucl],
                                mode="lines", name=f"MR UCL ({mr_ucl:.4f})",
                                line=dict(color="#EF4444", width=1.5, dash="dash"),
                            )
                        )

                        # MR out-of-control
                        mr_ooc = [i for i, v in enumerate(mr_values) if v > mr_ucl]
                        if mr_ooc:
                            fig_mr.add_trace(
                                go.Scatter(
                                    x=[i + 2 for i in mr_ooc],
                                    y=[mr_values[i] for i in mr_ooc],
                                    mode="markers", name="MR Out of Control",
                                    marker=dict(size=12, color="#EF4444", symbol="circle-open", line=dict(width=2)),
                                )
                            )

                        # MR annotations
                        mr_annotations = [
                            dict(x=1.02, y=mr_ucl, xref="paper", yref="y", text="MR UCL", showarrow=False,
                                 font=dict(size=9, color="#EF4444"), xanchor="left"),
                            dict(x=1.02, y=mr_bar, xref="paper", yref="y", text="MR̄", showarrow=False,
                                 font=dict(size=9, color="#10B981"), xanchor="left"),
                        ]

                        fig_mr.update_layout(
                            title=dict(text=f"MR-Chart — {viz_char_name} ({n-1} ranges)", font=dict(size=12, color=_fc)),
                            annotations=mr_annotations,
                            **{**_ctrl_layout,
                               "xaxis": dict(title=dict(text="Sample Number", font=dict(color=_fc, size=11)),
                                             tickfont=dict(size=10, color=_fc),
                                             gridcolor="rgba(128,128,128,0.15)"),
                               "yaxis": dict(title=dict(text="Moving Range |Xᵢ − Xᵢ₋₁|", font=dict(color=_fc, size=11)),
                                             tickfont=dict(size=10, color=_fc),
                                             gridcolor="rgba(128,128,128,0.15)")},
                        )

                        st.plotly_chart(fig_mr, use_container_width=True, config=PlotManager.PLOT_CONFIG,
                                        key=f"viz_mrchart_{viz_char_name}")

                elif not figs and len(viz_data) < 2:
                    st.info(
                        f"No analysis results for **{viz_char_name}** yet. Run analysis on the 'Analysis & Report' tab."
                    )



# --- Tab 4: History ---
with tab_history:
    st.header("Analysis History (Last 250 Runs)")
    st.caption('History is logged only when you click the "ANALYZE & PLOT" button.')

    if not st.session_state.history:
        st.info("No history available. Run an analysis to log it here.")
    else:
        # ---------- Summary metrics ----------
        hist_df_all = pd.DataFrame(st.session_state.history)
        cpk_vals = pd.to_numeric(hist_df_all.get("CpkCurrent", pd.Series()), errors="coerce").dropna()
        sum_cols = st.columns(5)
        with sum_cols[0]:
            st.metric("Total Runs", len(st.session_state.history))
        with sum_cols[1]:
            st.metric("Avg Cpk", f"{cpk_vals.mean():.3f}" if len(cpk_vals) else "—")
        with sum_cols[2]:
            st.metric("Best Cpk", f"{cpk_vals.max():.3f}" if len(cpk_vals) else "—")
        with sum_cols[3]:
            st.metric("Worst Cpk", f"{cpk_vals.min():.3f}" if len(cpk_vals) else "—")
        with sum_cols[4]:
            good_count = sum(1 for e in st.session_state.history if "GOOD" in e.get("verdict", ""))
            st.metric("Pass Rate", f"{good_count / len(st.session_state.history) * 100:.0f}%")

        # ---------- Cpk trend chart ----------
        if len(cpk_vals) >= 2:
            st.subheader("📈 Cpk Trend Over Time")
            trend_df = hist_df_all[["id", "CpkCurrent"]].copy()
            trend_df["CpkCurrent"] = pd.to_numeric(trend_df["CpkCurrent"], errors="coerce")
            trend_df = trend_df.dropna(subset=["CpkCurrent"])
            trend_df["Timestamp"] = trend_df["id"].apply(
                lambda v: datetime.datetime.fromisoformat(v) if v else None)
            trend_df = trend_df.dropna(subset=["Timestamp"]).sort_values("Timestamp")

            fig_trend = go.Figure()
            fig_trend.add_trace(go.Scatter(
                x=trend_df["Timestamp"], y=trend_df["CpkCurrent"],
                mode="lines+markers", name="Cpk",
                line=dict(color="#3B82F6", width=2),
                marker=dict(size=6, color="#3B82F6"),
                hovertemplate="%{x|%Y-%m-%d %H:%M}<br>Cpk: %{y:.3f}<extra></extra>",
            ))
            # Threshold lines
            fig_trend.add_hline(y=1.33, line_dash="dash", line_color="#F59E0B",
                                annotation_text="Min (1.33)", annotation_position="bottom right")
            fig_trend.add_hline(y=1.67, line_dash="dash", line_color="#10B981",
                                annotation_text="Target (1.67)", annotation_position="top right")
            _fc = "#8b95a5"
            fig_trend.update_layout(
                height=280, margin=dict(t=30, b=50, l=50, r=30),
                xaxis=dict(title="Run Date", tickfont=dict(size=10, color=_fc), gridcolor="rgba(128,128,128,0.15)"),
                yaxis=dict(title="Cpk", tickfont=dict(size=10, color=_fc), gridcolor="rgba(128,128,128,0.15)"),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font=dict(color=_fc),
                showlegend=False,
            )
            st.plotly_chart(fig_trend, use_container_width=True, config=PlotManager.PLOT_CONFIG)

        st.divider()

        # ---------- Filters ----------
        hist_filter_cols = st.columns([1.4, 1, 1])
        with hist_filter_cols[0]:
            filter_name = st.text_input("Filter by Name")
        with hist_filter_cols[1]:
            filter_verdict = st.selectbox(
                "Filter by Verdict",
                ["all", "PROCESS HEALTH: GOOD", "MARGINAL", "ACTION REQUIRED", "INVALID INPUTS"],
            )
        with hist_filter_cols[2]:
            filter_characteristic = st.selectbox(
                "Filter by Characteristic",
                ["all"] + sorted({entry.get("characteristic_name", entry.get("measurement_name", "")) for entry in st.session_state.history}),
            )

        history_df = pd.DataFrame(st.session_state.history)
        if "characteristic_name" not in history_df.columns:
            history_df["characteristic_name"] = history_df.get("measurement_name", "")

        filtered_history = history_df
        if filter_name:
            filtered_history = filtered_history[
                filtered_history["measurement_name"].str.contains(filter_name, case=False, na=False)
            ]
        if filter_verdict != "all":
            filtered_history = filtered_history[filtered_history["verdict"] == filter_verdict]
        if filter_characteristic != "all":
            filtered_history = filtered_history[
                filtered_history["characteristic_name"].fillna(filtered_history["measurement_name"]) == filter_characteristic
            ]

        display_cols = [
            "id", "characteristic_name", "measurement_name", "verdict",
            "Cp", "CpkCurrent", "shiftValue", "tm", "lsl", "usl",
            "x_bar", "s", "n_samples", "ppm_below", "ppm_above",
        ]
        rename_map = {
            "characteristic_name": "Characteristic", "measurement_name": "Name",
            "verdict": "Verdict", "CpkCurrent": "Cpk", "shiftValue": "Shift (Δ)",
            "tm": "Tₘ", "x_bar": "Mean (x̄)", "s": "StdDev (σ)",
            "n_samples": "n", "ppm_below": "PPM < LSL", "ppm_above": "PPM > USL",
        }

        display_df = filtered_history[display_cols].copy()
        display_df.insert(0, "Select", False)
        display_df["Timestamp"] = display_df["id"].apply(
            lambda value: datetime.datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M:%S") if value else ""
        )
        display_df.rename(columns=rename_map, inplace=True)

        format_config = {
            "Select": st.column_config.CheckboxColumn("Select", help="Choose rows to include in the export."),
            "Timestamp": st.column_config.TextColumn(),
            "Cp": st.column_config.NumberColumn(format="%.3f"),
            "Cpk": st.column_config.NumberColumn(format="%.3f"),
            "Shift (Δ)": st.column_config.NumberColumn(format="%.3f"),
            "Tₘ": st.column_config.NumberColumn(format="%.3f"),
            "LSL": st.column_config.NumberColumn(format="%.3f"),
            "USL": st.column_config.NumberColumn(format="%.3f"),
            "Mean (x̄)": st.column_config.NumberColumn(format="%.3f"),
            "StdDev (σ)": st.column_config.NumberColumn(format="%.4f"),
            "PPM < LSL": st.column_config.NumberColumn(format="%d"),
            "PPM > USL": st.column_config.NumberColumn(format="%d"),
        }

        st.markdown(f"Showing **{len(filtered_history)}** of **{len(st.session_state.history)}** runs. Select rows to export:")
        selection_df = st.data_editor(
            display_df, column_config=format_config,
            hide_index=True, use_container_width=True,
            disabled=["Timestamp", "Characteristic", "Name", "Verdict", "Cp", "Cpk",
                       "Shift (Δ)", "Tₘ", "LSL", "USL", "Mean (x̄)", "StdDev (σ)", "n", "PPM < LSL", "PPM > USL"],
            column_order=["Select", "Timestamp", "Characteristic", "Name", "Verdict", "Cp", "Cpk",
                          "Shift (Δ)", "Tₘ", "LSL", "USL", "Mean (x̄)", "StdDev (σ)", "n", "PPM < LSL", "PPM > USL"],
            key="history_selection_editor",
        )

        selected_ids = selection_df.loc[selection_df["Select"], "id"].tolist()

        # Action buttons
        btn_cols = st.columns([1, 1, 1, 2])
        with btn_cols[0]:
            if selected_ids:
                try:
                    selected_history_data = [e for e in st.session_state.history if e.get("id") in selected_ids]
                    history_buffer = exporter.export_selected_history(selected_history_data)
                    st.download_button(
                        label=f"📥 Export Selected ({len(selected_ids)})",
                        data=history_buffer,
                        file_name=f"Capability_History_Selection_{datetime.date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Could not generate history export: {e}")
            else:
                st.button("📥 Export Selected (0)", use_container_width=True, disabled=True)

        with btn_cols[1]:
            # CSV download of full history
            csv_data = pd.DataFrame(st.session_state.history).to_csv(index=False)
            st.download_button(
                "📄 Download CSV",
                data=csv_data,
                file_name=f"SPC_History_{datetime.date.today()}.csv",
                mime="text/csv",
                use_container_width=True,
            )
        with btn_cols[2]:
            if st.button("🗑 Clear History", use_container_width=True):
                st.session_state.history = []
                st.rerun()


# --- Tab 5: Reference ---
with tab_ref:
    st.header("Reference Guide & Chatbot")

    ref_cols = st.columns([2, 1])

    with ref_cols[0]:
        st.subheader("Quick Reference Tables")

        # --- Cpk Thresholds ---
        with st.expander("📏 Cpk / Capability Thresholds", expanded=True):
            st.markdown("""
| Cpk Range | Rating | PPM (approx.) | Action |
|-----------|--------|---------------|--------|
| < 1.00 | ❌ Not Capable | > 2,700 | Process redesign or tighter controls needed |
| 1.00 – 1.33 | ⚠️ Marginal | 63 – 2,700 | Improvement required, monitor closely |
| 1.33 – 1.67 | ✅ Capable | 0.6 – 63 | Meets most industry standards |
| 1.67 – 2.00 | ✅ Highly Capable | < 0.6 | Meets automotive/safety-critical requirements |
| ≥ 2.00 | 🏆 Six Sigma | < 0.002 | World-class capability |
""")

        # --- Sigma Level Table ---
        with st.expander("📊 Sigma Level & PPM Table"):
            st.markdown("""
| Sigma Level | Yield (%) | DPMO (PPM) | Cpk |
|-------------|-----------|------------|-----|
| 1σ | 30.85% | 691,462 | 0.33 |
| 2σ | 69.15% | 308,538 | 0.67 |
| 3σ | 93.32% | 66,807 | 1.00 |
| 4σ | 99.3790% | 6,210 | 1.33 |
| 5σ | 99.97670% | 233 | 1.67 |
| 6σ | 99.99966% | 3.4 | 2.00 |

> *PPM values include the standard 1.5σ long-term shift assumption.*
""")

        # --- Core SPC Formulas ---
        with st.expander("🔬 Core SPC Formulas"):
            st.markdown("""
| Formula | Expression | Purpose |
|---------|-----------|---------|
| **Cp** | (USL − LSL) / 6σ | Potential capability (centered) |
| **Cpk** | min[(USL − x̄)/3σ, (x̄ − LSL)/3σ] | Actual capability (with shift) |
| **Pp** | (USL − LSL) / 6σ_overall | Long-term potential performance |
| **Ppk** | min[(USL − x̄)/3σ_overall, (x̄ − LSL)/3σ_overall] | Long-term actual performance |
| **Shift (Δ)** | Tₘ − x̄ | Required mean adjustment |
| **Z-score** | (x̄ − Tₘ) / (σ / √n) | Hypothesis test statistic |
| **UCL** | x̄ + 3σ | Upper control limit |
| **LCL** | x̄ − 3σ | Lower control limit |
| **MR̄** | Σ|Xᵢ − Xᵢ₋₁| / (n−1) | Average moving range |
| **MR UCL** | 3.267 × MR̄ | MR chart upper control limit |
""")

        # --- Control Chart Zones ---
        with st.expander("🎯 Control Chart Zones (I-MR)"):
            st.markdown("""
| Zone | Range | Expected % | Color |
|------|-------|-----------|-------|
| **Zone C** | x̄ ± 1σ | 68.27% | Green (normal) |
| **Zone B** | x̄ ± 1σ to ± 2σ | 27.18% | Amber (caution) |
| **Zone A** | x̄ ± 2σ to ± 3σ | 4.28% | Red (warning) |
| **Outside** | Beyond ± 3σ | 0.27% | Out of Control |

**Western Electric Rules for detecting out-of-control conditions:**
1. Any single point beyond ±3σ
2. Two of three consecutive points beyond ±2σ (same side)
3. Four of five consecutive points beyond ±1σ (same side)
4. Eight consecutive points on the same side of CL
5. Six consecutive points trending up or down
""")

        # --- Industry Standards ---
        with st.expander("🏭 Industry Standard Requirements"):
            st.markdown("""
| Standard | Application | Cpk Requirement |
|----------|------------|-----------------|
| **IATF 16949** | Automotive production | ≥ 1.33 (ongoing), ≥ 1.67 (new process) |
| **VDA 6.1** | German automotive | ≥ 1.33 (Cmk), ≥ 1.67 (critical) |
| **AS9100** | Aerospace | ≥ 1.33 (typical), ≥ 1.5 (critical) |
| **ISO 13485** | Medical devices | ≥ 1.33 (critical features) |
| **Six Sigma** | General manufacturing | ≥ 2.0 (6σ target) |
| **ISO 22514** | Capability study standard | Defines Cm/Cmk/Pp/Ppk procedures |
""")

        st.divider()

        # --- Detailed Reference Sections ---
        with st.expander("📖 Application Context & Usage Guide"):
            st.markdown("""
#### Technical Application: Process Centering and Root Cause Analysis
This tool is primarily utilized in **Six Sigma and SPC** environments for **Process Centering and Tolerance Verification**:
-   **Quantify Process Drift:** Calculate the exact **Δ = Tₘ − x̄** needed to re-center the process.
-   **Predict Initial State:** Infer the required initial dimension/setting to achieve target after process variables.
-   **Verify Tolerance Adequacy:** Determine the minimum **Required Tolerance** for desired Cpk/Ppk.

#### Step-by-Step Usage
1.  **Data Worksheet:** Enter measurement label, Tₘ, LSL, USL, and measured values.
2.  **Load data:** Type, upload CSV/Excel, or use '📥 Download Template' → fill → upload.
3.  **Analysis & Report:** Select mode → click **ANALYZE & PLOT**.
4.  **Visualization:** Histogram, box plot, I-Chart, MR-Chart with filter, stats summary.
5.  **History:** Compare runs over time, export as Excel or CSV.
""")

        with st.expander("📖 Hypothesis Testing & Confidence Level"):
            st.markdown("""
This tool performs a **Z-test** to determine if μ ≠ Tₘ:
-   **H₀: μ = Tₘ** (on target) vs. **H₁: μ ≠ Tₘ** (shifted)
-   **p-value < 0.05** → Reject H₀ (significant shift detected)
-   **p-value ≥ 0.05** → Fail to Reject H₀ (no significant evidence of shift)

| Test Type | When to Use |
|-----------|-------------|
| **Two-Sided** | Both directions matter (e.g., diameter) |
| **Upper-Sided** | Only concerned if mean is ABOVE target |
| **Lower-Sided** | Only concerned if mean is BELOW target |
""")

        with st.expander("📖 Additional Metric Definitions"):
            st.markdown("""
| Metric | Definition |
|--------|-----------|
| **x̄ ± 4σ Spread** | Contains ~99.9937% of output (conservative) |
| **P(x < LSL)** | Probability a part is below lower spec |
| **P(x > USL)** | Probability a part is above upper spec |
| **P(x < Tₘ)** | Should be 50% if centered; measures bias |
| **PPM** | Defect rate in parts per million |
| **MR̄** | Average moving range (point-to-point variation) |
| **Required Shift (Δ)** | Tₘ − x̄ (positive = increase, negative = decrease) |
| **Required Tolerance** | Target Index × 6σ (minimum spec band needed) |
""")

    with ref_cols[1]:
        st.subheader("Guide Chatbot")
        st.info("Ask me a question about the reference guide!")

        # Chat message history
        if "chat_messages" not in st.session_state:
            st.session_state.chat_messages = []

        for message in st.session_state.chat_messages:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])

        # Chat input
        if prompt := st.chat_input("Ask about 'Cp', 'Cpk', 'PPM', 'hypothesis' ..."):
            # Add user message to chat history
            st.session_state.chat_messages.append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)

            # Get bot response
            response = bot.get_response(prompt)

            # Add bot response to chat history
            with st.chat_message("assistant"):
                st.markdown(response)
            st.session_state.chat_messages.append(
                {"role": "assistant", "content": response}
            )

# --- Floating Sigma Assistant (Clippy-style) ---
# This renders as a fixed position widget in the bottom-right corner of the page
# Using st.markdown to inject directly into Streamlit's DOM for TRUE fixed positioning
mascot_html = SigmaAssistant.render_fixed(
    state=st.session_state.get("mascot_state", "idle"),
    message=st.session_state.get("mascot_message", None),
)
st.markdown(mascot_html, unsafe_allow_html=True)
